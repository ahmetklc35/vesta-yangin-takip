from __future__ import annotations

import io
import json
import os
import smtplib
import tempfile
import re
import unicodedata
import uuid
import base64
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from functools import wraps

import fitz
import qrcode
from qrcode.constants import ERROR_CORRECT_H
from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont
from playwright.sync_api import sync_playwright
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader, simpleSplit
from reportlab.platypus import Flowable, Image as PdfImage, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from sqlalchemy import (
    Boolean,
    Column,
    Float,
    ForeignKey,
    Integer,
    MetaData,
    String,
    Table,
    text,
    create_engine,
    desc,
    insert,
    select,
    update,
    UniqueConstraint,
)
from sqlalchemy.engine import Engine
from sqlalchemy.exc import IntegrityError
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from xhtml2pdf import pisa


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SQLITE_URL = f"sqlite:///{(BASE_DIR / 'database.db').as_posix()}"
BRAND_NAME = "Vesta Yangin"
LOGO_PATH = BASE_DIR / "static" / "vesta qr.png"
VESTA_HEADER_LOGO_PATH = BASE_DIR / "static" / "vesta-form-amblem.png"
TSE_HYB_LOGO_PATH = BASE_DIR / "static" / "tse-form-amblem.png"
CONTROL_FORM_TEMPLATE_PATH = BASE_DIR / "assets" / "control-form-template.pdf"
CONTROL_FORM_TEMPLATE_IMAGE_PATH = BASE_DIR / "assets" / "control-form-template.png"
CONTROL_FORM_EXCEL_TEMPLATE_PATH = BASE_DIR / "assets" / "control-form-template.xlsx"
ELECTRICAL_REPORT_TEMPLATE_PATH = BASE_DIR / "assets" / "electrical-installation-template.pdf"
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USERNAME)
ALERT_WINDOW_DAYS = int(os.environ.get("ALERT_WINDOW_DAYS", "7"))
CONTROL_FORM_METHOD_TEXT = "İEKSGŞY, TS ISO 11602-2, TS 862-7 EN 3-7 + A1 ve TS EN 1866-1 standartlarına göre kontrol edilmiştir."
CONTROL_FORM_NOTES = [
    "NOT 1: Yangın söndürücünün kontrolü Madde a) ve b) bendindeki gibi listelenmiş koşullarda, bir eksikliği ortaya çıkardığı zaman, acil düzeltici faaliyet yapılmalıdır.",
    "NOT 2: Yangın söndürücünün kontrolü Madde c), d), e), f) veya g) bendindeki koşullarından herhangi birinde bir eksikliği ortaya çıkardığı zaman, söndürücü uygun bakım işlemlerine VESTA YANGIN tarafından tabi tutulmalıdır.",
    "NOT 3: Madde c), d), e), f) veya g) bendindeki koşullarından herhangi birinde, doldurulmayan tozlu söndürücünün kontrolü bir eksikliği ortaya çıkardığı zaman, bu söndürücü hizmetten kaldırılmalıdır.",
    "NOT 4: Bu muayene raporundaki bulgular muayene tarihindeki işletme koşulları için geçerlidir. Bu rapor 2 nüsha basılmıştır. Muayene raporu VESTA YANGIN onayı olmaksızın kopya edilemez",
]
SCBA_METHOD_TEXT = "İEKSGŞY, TS ISO 11602-2, TS 862-7 EN 3-7 + A1 ve TS EN 1866-1 standartlarına göre kontrol edilmiştir."
SCBA_NOTES = [
    "NOT 1: Bağımsız solunum cihazı kontrolü Madde a), b), c), d), e), f) bendindeki gibi listelenmiş koşullarda, bir eksikliği ortaya çıkardığı zaman, acil düzeltici faaliyet yapılmalıdır.",
    "NOT 2: Bağımsız solunum cihazı kontrolü Madde d), e), f) veya g) bendindeki koşullarından herhangi birinde bir eksikliği ortaya çıkardığı zaman, söndürücü uygun bakım işlemlerine VESTA YANGIN tarafından tabi tutulmalıdır.",
    "NOT 3: Madde a), b), c), d) veya e) bendindeki koşullarından herhangi birinde, bağımsız solunum cihazı kontrolü bir eksikliği ortaya çıkardığı zaman, bu solunum cihazı hizmetten kaldırılmalıdır.",
    "NOT 4: Bu muayene raporundaki bulgular muayene tarihindeki işletme koşulları için geçerlidir. Bu rapor 2 nüsha basılmıştır. Muayene raporu VESTA YANGIN onayı olmaksızın kopya edilemez.",
]
SPECIAL_CATEGORY_FORM_CONFIGS = {
    "SCBA": {
        "form_code": "F-17",
        "subject": "BAGIMSIZ SOLUNUM CIHAZI (SELF-CONTAINED BREATHING APPARATUS) Kontrol Formu",
        "section_title": "BAGIMSIZ SOLUNUM CIHAZI (SELF-CONTAINED BREATHING APPARATUS) BILGILERI",
        "method_text": "IEKSGSY, TS ISO 11602-2, TS 862-7 EN 3-7 + A1 ve TS EN 1866-1 standartlarina gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Bagimsiz solunum cihazi kontrolu Madde a), b), c), d), e), f) bendindeki gibi listelenmis kosullarda, bir eksikligi ortaya cikardigi zaman, acil duzeltici faaliyet uygulanmalidir.",
            "NOT 2: Bagimsiz solunum cihazi kontrolu Madde d), e), f) veya g) bendindeki kosullardan herhangi birinde bir eksiklik ortaya cikardigi zaman, cihaz VESTA YANGIN tarafindan bakima alinmalidir.",
            "NOT 3: Madde a), b), c), d) veya e) bendindeki kosullardan herhangi birinde bir eksiklik ortaya cikardigi zaman, bu cihaz hizmetten kaldirilmalidir.",
            "NOT 4: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir. Bu rapor 2 nusha basilmistir. Muayene raporu VESTA YANGIN onayi olmaksizin kopya edilemez.",
        ],
    },
    "EEBD": {
        "form_code": "F-18",
        "subject": "ACIL KACIS SETI (EMERGENCY ESCAPE BREATHING DEVICE SET) Kontrol Formu",
        "section_title": "ACIL KACIS SETI (EEBD) BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: EEBD kontrolunde uygunsuzluk tespit edilirse ekipman kullanima verilmeden once duzeltici faaliyet uygulanmalidir.",
            "NOT 2: Solunum, regulator, silindir veya kemer bilesenlerinde uygunsuzluk varsa ekipman VESTA YANGIN tarafindan bakima alinmalidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Hava Tupu": {
        "form_code": "F-19",
        "subject": "BASINCLI HAVA SOLUNUM TUPU (COMPRESSED AIR BREATHING CYLINDERS) Kontrol Formu",
        "section_title": "BASINCLI HAVA SOLUNUM TUPU BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Valf, silindir veya servis etiketinde uygunsuzluk tespit edilirse tup hizmete verilmeden once duzeltici faaliyet uygulanmalidir.",
            "NOT 2: Hidrostatik test ve dolum bilgileri guncel tutulmalidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Yangin Elbisesi": {
        "form_code": "F-20",
        "subject": "YANGIN ELBISESI Kontrol Formu",
        "section_title": "YANGIN ELBISESI BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Kumas, fermuar, bant, astar veya dikislerde uygunsuzluk tespit edilirse ekipman kullanima verilmeden once bakima alinmalidir.",
            "NOT 2: Servis etiketi guncel olmayan ekipman hizmete verilmemelidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Yangin Bareti": {
        "form_code": "F-21",
        "subject": "YANGIN BARETI Kontrol Formu",
        "section_title": "YANGIN BARETI BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Baretin dis yuzu, ic yapi, vizor veya boyun koruyucusunda uygunsuzluk varsa ekipman bakima alinmalidir.",
            "NOT 2: Servis etiketi guncel olmayan ekipman hizmete verilmemelidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Yangin Baltasi": {
        "form_code": "F-22",
        "subject": "YANGIN BALTASI Kontrol Formu",
        "section_title": "YANGIN BALTASI BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Sap, metal kisim, agiz veya ulasilabilirlikte uygunsuzluk varsa ekipman kullanima verilmeden once duzeltilmelidir.",
            "NOT 2: Servis etiketi guncel olmayan ekipman hizmete verilmemelidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Yangin Sondurme Dolabi": {
        "form_code": "F-23",
        "subject": "YANGIN SONDURME DOLABI Periyodik Kontrol Formu",
        "section_title": "YANGIN SONDURME DOLABI BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Dolap, hortum, lans, vana ve basinc kontrol kriterlerinden herhangi birinde uygunsuzluk tespit edilirse dolap bakima alinmalidir.",
            "NOT 2: Erisebilirlik ve levha uygunsuzluklari acil duzeltici faaliyet gerektirir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Kopuklu Yangin Sondurme Dolabi": {
        "form_code": "F-24",
        "subject": "KOPUKLU YANGIN SONDURME DOLABI Periyodik Kontrol Formu",
        "section_title": "KOPUKLU YANGIN SONDURME DOLABI BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Dolap, hortum, vana, oranlayici, basinc ve lans kriterlerinden herhangi birinde uygunsuzluk varsa sistem bakima alinmalidir.",
            "NOT 2: Kopuk doluluk ve karisim ayarlari guncel tutulmalidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Yangin Hidranti": {
        "form_code": "F-25",
        "subject": "YANGIN HIDRANTI Periyodik Kontrol Formu",
        "section_title": "YANGIN HIDRANTI BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Hidrant ulasilabilirligi, kapak, vana sizdirmazligi ve cikis agizlarinda uygunsuzluk tespit edilirse sistem bakima alinmalidir.",
            "NOT 2: Erisebilirlik ve gorunurluk uygunsuzluklari acil duzeltici faaliyet gerektirir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Kasik Sedye": {
        "form_code": "F-26",
        "subject": "KASIK SEDYE Periyodik Kontrol Formu",
        "section_title": "KASIK SEDYE BILGILERI",
        "method_text": "IEKSGSY ve firma kontrol kriterlerine gore kontrol edilmistir.",
        "notes": [
            "NOT 1: Aluminyum yuzeylerde catlak, egilme veya deformasyon tespit edilirse sedye hizmetten kaldirilmalidir.",
            "NOT 2: Acilma-kapanma mekanizmasi, kilitler veya sabitleme pimleri calismiyorsa sedye kullanima verilmemelidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki isletme kosullari icin gecerlidir.",
        ],
    },
    "Yangin Cizmesi": {
        "form_code": "F-27",
        "subject": "YANGIN ÇİZMESİ Kontrol Formu", 
        "section_title": "YANGIN ÇİZMESİ BİLGİLERİ", 
        "method_text": "İEKSGŞY ve firma kontrol kriterlerine göre kontrol edilmiştir.",
        "notes": [
            "NOT 1: Çizme yüzeyinde veya tabanında koruyucu özelliği yitiren hasar tespit edilirse ekipman değiştirilmelidir.",
            "NOT 2: Hijyen ve astar bütünlüğü kullanıcı sağlığı için periyodik olarak denetlenmelidir.",
            "NOT 3: Bu muayene raporundaki bulgular muayene tarihindeki işletme koşulları için geçerlidir.",
        ],
    },
    "Geri Sarimli Dussundurucu": {
        "form_code": "F-28",
        "subject": "GERİ SARIMLI DÜŞÜŞ DURDURUCU Kontrol Formu",
        "section_title": "EKİPMAN BİLGİLERİ",
        "method_text": "İEKSGŞY ve üretici kriterlerine göre kontrol edilmiştir.",
        "notes": ["NOT: Kilitleme mekanizmasında gecikme olan cihazlar derhal servis dışı bırakılmalıdır."],
    },
    "Omurga Tahtasi": {
        "form_code": "F-29",
        "subject": "OMURGA TAHTASI Periyodik Kontrol Formu",
        "section_title": "SEDYE BİLGİLERİ",
        "method_text": "Firma kontrol kriterlerine göre kontrol edilmiştir.",
        "notes": ["NOT: Yüzeyde derin çatlak tespit edilen tahtalar taşıma kapasitesini yitirmiş sayılır."],
    },
    "Parasut Tipi Kemer": {
        "form_code": "F-30",
        "subject": "PARAŞÜT TİPİ EMNİYET KEMERİ Kontrol Formu",
        "section_title": "KEMER BİLGİLERİ",
        "method_text": "İEKSGŞY ve ilgili standartlara göre kontrol edilmiştir.",
        "notes": ["NOT: Tekstil aksamda erime, kimyasal yanık veya kesik varsa kemer imha edilmelidir."],
    },
    "Sedye Tasima Sapani": {
        "form_code": "F-31",
        "subject": "SEDYE TAŞIMA SAPANI Kontrol Formu",
        "section_title": "SAPAN BİLGİLERİ",
        "method_text": "Firma güvenlik prosedürlerine göre kontrol edilmiştir.",
        "notes": ["NOT: Dengesiz yükleme riskine karşı bağlantı halkaları periyodik kontrol edilmelidir."],
    },
    "Sepet Sedye": {
        "form_code": "F-32",
        "subject": "SEPET SEDYE Kontrol Formu",
        "section_title": "SEDYE BİLGİLERİ",
        "method_text": "Firma kontrol kriterlerine göre kontrol edilmiştir.",
        "notes": ["NOT: Halat ve zincir bağlantılarında korozyon kontrolü hayati önem taşır."],
    },
}

EQUIPMENT_OPTIONS = [
    "Kuru Kimyevi Toz",
    "CO2",
    "Kopuk",
]
ASSET_CATEGORIES = [
    {
        "slug": "yangin-sondurme-cihazi",
        "label": "Yangin Sondurme Cihazi",
        "description": "Yangin tup ve sondurme cihazlarinin tamamini goruntule.",
    },
    {
        "slug": "yangin-elbisesi",
        "label": "Yangin Elbisesi",
        "description": "Yangin elbisesi kontrollerini ve kayitlarini listele.",
    },
    {
        "slug": "yangin-bareti",
        "label": "Yangin Bareti",
        "description": "Yangin baretlerine ait kontrolleri ve kayitlari goruntule.",
    },
    {
        "slug": "yangin-baltasi",
        "label": "Yangin Baltasi",
        "description": "Yangin baltasi ekipmanlarini listele.",
    },
    {
        "slug": "scba",
        "label": "SCBA",
        "description": "Bagimsiz solunum cihazi kayitlarini goruntule.",
    },
    {
        "slug": "eebd",
        "label": "EEBD",
        "description": "Acil kacis solunum setlerini goruntule.",
    },
    {
        "slug": "hava-tupu",
        "label": "Hava Tupu",
        "description": "Basincli hava solunum tupu kayitlarini goruntule.",
    },
    {
        "slug": "yangin-sondurme-dolabi",
        "label": "Yangin Sondurme Dolabi",
        "description": "Yangin dolabi periyodik kontrol kayitlarini listele.",
    },
    {
        "slug": "kopuklu-yangin-sondurme-dolabi",
        "label": "Kopuklu Yangin Sondurme Dolabi",
        "description": "Kopuklu dolap kontrol kayitlarini listele.",
    },
    {
        "slug": "yangin-hidranti",
        "label": "Yangin Hidranti",
        "description": "Yangin hidrant periyodik kontrol kayitlarini goruntule.",
    },
    {
        "slug": "elektrik-ic-tesisati",
        "label": "Elektrik Ic Tesisati",
        "description": "Elektrik ic tesisati periyodik kontrol raporlarini goruntule.",
    },
    {
        "slug": "kasik-sedye",
        "label": "Kasik Sedye",
        "description": "Kasik sedye periyodik kontrol kayitlarini goruntule.",
    },
    {
        "slug": "yangin-cizmesi",
        "label": "Yangin Cizmesi",
        "description": "Yangin cizmesi periyodik kontrol kayitlarini goruntule.",
    },
    {
        "slug": "geri-sarimli-dusus-durdurucu",
        "label": "Geri Sarımlı Düşüş Durdurucu",
        "description": "Geri sarımlı düşüş durdurucu ekipmanlarının periyodik kontrollerini yönet.",
    },
    {
        "slug": "omurga-tahtasi",
        "label": "Omurga Tahtası",
        "description": "Omurga tahtası (backboard) ekipmanlarına ait kayıtları görüntüle.",
    },
    {
        "slug": "parasut-tipi-emniyet-kemeri",
        "label": "Paraşüt Tipi Emniyet Kemeri",
        "description": "Paraşüt tipi emniyet kemerlerinin teknik kontrollerini listele.",
    },
    {
        "slug": "sedye-tasima-sapani",
        "label": "Sedye Taşıma Sapanı",
        "description": "Sedye taşıma sapanı ekipmanlarının periyodik muayene kayıtları.",
    },
    {
        "slug": "sepet-sedye",
        "label": "Sepet Sedye",
        "description": "Sepet sedye ekipmanlarına ait periyodik kontrol formlarını yönet.",
    },
]
DEFAULT_ASSET_CATEGORY = ASSET_CATEGORIES[0]["label"]
ASSET_CATEGORY_BY_SLUG = {item["slug"]: item for item in ASSET_CATEGORIES}
ASSET_CATEGORY_BY_LABEL = {item["label"]: item for item in ASSET_CATEGORIES}
REGISTRATION_GROUPS = [
    {
        "slug": "yangin-sondurme-cihazi",
        "label": "Yangin Sondurme Cihazi",
        "description": "Mevcut calisan YSC kayit ekranini kullanir.",
        "status": "active",
    },
    {
        "slug": "yangin-elbisesi",
        "label": "Yangin Elbisesi",
        "description": "Yangin elbisesi kontrol formuna uygun yeni kayit akisi.",
        "status": "active",
    },
    {
        "slug": "yangin-bareti",
        "label": "Yangin Bareti",
        "description": "Yangin bareti kontrol formuna uygun kayit akisi.",
        "status": "active",
    },
    {
        "slug": "yangin-baltasi",
        "label": "Yangin Baltasi",
        "description": "Yangin baltasi kontrol formuna uygun kayit akisi.",
        "status": "active",
    },
    {
        "slug": "scba",
        "label": "SCBA",
        "description": "Bagimsiz solunum cihazi kontrol formuna uygun kayit akisi.",
        "status": "active",
    },
    {
        "slug": "eebd",
        "label": "EEBD",
        "description": "Acil kacis seti kontrol formuna uygun kayit akisi.",
        "status": "active",
    },
    {
        "slug": "hava-tupu",
        "label": "Hava Tupu",
        "description": "Basincli hava solunum tupu kontrol formuna uygun kayit akisi.",
        "status": "active",
    },
    {
        "slug": "yangin-sondurme-dolabi",
        "label": "Yangin Sondurme Dolabi",
        "description": "Yangin sondurme dolabi periyodik kontrol kayit akisi.",
        "status": "active",
    },
    {
        "slug": "kopuklu-yangin-sondurme-dolabi",
        "label": "Kopuklu Yangin Sondurme Dolabi",
        "description": "Kopuklu yangin sondurme dolabi kontrol kayit akisi.",
        "status": "active",
    },
    {
        "slug": "yangin-hidranti",
        "label": "Yangin Hidranti",
        "description": "Yangin hidranti periyodik kontrol kayit akisi.",
        "status": "active",
    },
    {
        "slug": "elektrik-ic-tesisati",
        "label": "Elektrik Ic Tesisati",
        "description": "Elektrik ic tesisati gozle kontrol ve fonksiyon testleri icin trial kayit akisi.",
        "status": "active",
    },
    {
        "slug": "kasik-sedye",
        "label": "Kasik Sedye",
        "description": "Kasik sedye periyodik kontrol kayit akisi.",
        "status": "active",
    },
    {
        "slug": "yangin-cizmesi",
        "label": "Yangin Cizmesi",
        "description": "Yangin cizmesi kontrol formuna uygun kayit akisi.",
        "status": "active",
    },
    {
        "slug": "geri-sarimli-dusus-durdurucu",
        "label": "Geri Sarimli Dusus Durdurucu",
        "description": "Düşüş durdurucu sistemler için kayıt akışı.",
        "status": "active",
    },
    {
        "slug": "omurga-tahtasi",
        "label": "Omurga Tahtasi",
        "description": "Omurga tahtası sedyeler için kayıt akışı.",
        "status": "active",
    },
    {
        "slug": "parasut-tipi-emniyet-kemeri",
        "label": "Parasut Tipi Emniyet Kemeri",
        "description": "Emniyet kemerleri için kayıt akışı.",
        "status": "active",
    },
    {
        "slug": "sedye-tasima-sapani",
        "label": "Sedye Tasima Sapani",
        "description": "Taşıma sapanları için kayıt akışı.",
        "status": "active",
    },
    {
        "slug": "sepet-sedye",
        "label": "Sepet Sedye",
        "description": "Sepet sedyeler için kayıt akışı.",
        "status": "active",
    },
]
REGISTRATION_GROUP_BY_SLUG = {item["slug"]: item for item in REGISTRATION_GROUPS}
EQUIPMENT_PRESETS = {
    "Kopuk": {
        "title": "Taşınılabilir Yangın Söndürücü, 9 Litre Köpük",
        "summary": "YANGIN SINIFI: 21 A 183 B - 9 litrelik köpük yangın söndürücü , -30°C ila +60°C sıcaklık aralığında, UNI EN 3-7 standardına uygun olarak üretilmiş , MED 2014/90/EU Denizcilik Ekipmanları Direktifi onaylı, PED 2014/68/EU Basınçlı Ekipman Direktifi'ne göre sertifikalandırılmıştır . Üretim kontrolleri EN 3-10 standardına uygun olarak yapılmıştır . - Elektrik voltajı 1.000 V'a kadar olan yangınlarda, en az 1 metre mesafede kullanıma uygundur.",
        "features": "Özellikler: - SİLİNDİR: Yüksek mukavemetli alaşımlı çelikten, 3 parçalı klipsli, dış yüzeyi RAL 3000 Kırmızı toz boya ile boyanmış. -  SÖNDÜRÜCÜ MADDE: Su bazlı köpük. - İTİCİ AKIŞKAN : Nemi alınmış hava veya azot (N2). - VALF: M. 30x1.5, pirinç gövde, kolları RAL 6029 yeşil boya ile boyanmış. - KULLANIM: AB yangın sınıfı (katı maddeler, yanıcı sıvılar).",
        "technical_details": "Teknik Özellikler: -  YANGIN DAYANIKLILIĞI: 21 A 183 B -  SÖNDÜRÜCÜ MADDE: Köpük EW-30 (%100 Premix MG6-30) -  İTİCİ AKIŞKAN:  Nemi alınmış hava veya Azot (N2), 20°C'de 15 bar -  SICAKLIK ARALIĞI: -30°C / +60°C -  NOMİNAL HACİM:  9 Litre -  TOPLAM AĞIRLIK: ~ 15,1 Kg -  BOYUTLAR: Yükseklik (taban - valf) 635 +/- 5 mm - Çap (silindir) 170 +/- 2 mm -  BOŞALTMA SÜRESİ: ~ 34,8 saniye -  VALF SIKMA TORKU: Minimum 45 Nm, Maksimum 68 Nm -  SİLİNDİR BASINÇ  TESTİ : PT 27 bar -  SİLİNDİR HACMİ: 10,7 L. -  GÜVENLİK CİHAZI: Ayarlanmış 20 ile 26 bar arasında -  SİLİNDİR MALZEMESİ: Alaşımlı çelik -  İŞLEM: Dış yüzey: Kum püskürtme ve toz boyama, RAL 3000 - İç yüzey: Plastik kaplama",
        "image": "equipment/kopuk.png",
    },
    "CO2": {
        "title": "Taşınılabilir Yangın Söndürücü, 5 Kg CO₂",
        "summary": "YANGIN SINIFI: 3 7 113 B  - 5 kg Karbondioksitli Yangın Söndürücü, çalışma sıcaklığı -30°C ile +60°C arasında, UNI EN 3-7 (DM 7.1.2005) standardına uygun olarak üretilmiş, Denizcilik Ekipmanları Direktifi MED 2014/90/UE tarafından onaylanmış, Basınçlı Ekipman Direktifi PED 2014/68/EU'ya göre sertifikalandırılmıştır. EN 3-10 ile mutabık kalınan üretim kontrollerine göre üretilmiştir. - Elektrik voltajı 1.000 V'a kadar olan yangınlarda, en az 1 metre mesafede kullanıma uygundur.",
        "features": "Özellikler: - SİLİNDİR: Alaşımlı çelik, dış yüzey toz boya kaplama, renk Kırmızı RAL 3000. - SÖNDÜRÜCÜ MADDE: Karbondioksit %99,99 (CO2) - VALF: Levian M. 25x2, pirinç gövde, kol kırmızı RAL 3000. - KULLANIM: Yangın dayanıklılık sınıfı B (yanıcı sıvılar).",
        "technical_details": "Teknik Özellikler: • YANGIN DAYANIKLILIĞI 113 B • SÖNDÜRÜCÜ MADDE Karbondioksit %99,99 (CO2) • SICAKLIK ARALIĞI -30°C/ +60°C • 5 kg için nominal ücret. • TOPLAM AĞIRLIK ~ 13,65 Kg • BOYUTLAR Yükseklik (taban - valf) 740 +/- 10 mm - Çap (silindir) 136 +/- 2 mm • Boşalma süresi  ~ 15,6 saniye",
        "image": "equipment/co2.png",
    },
    "Kuru Kimyevi Toz": {
        "title": "Taşınılabilir Yangın Söndürücü, 6 Kg ABC Tozu",
        "summary": "YANGIN SINIFI: 34 A 233 B C  - 6 kg tozlu yangın söndürücü, -30°C ila +60°C sıcaklık aralığında, UNI EN 3-7 (DM 7.1.2005) standardına uygun olarak üretilmiş, PED 2014/68/EU basınçlı ekipman direktifine göre sertifikalandırılmıştır. Üretim kontrolleri EN 3-10 standardına uygun olarak gerçekleştirilmiştir. - Elektrik voltajı 1.000 V'a kadar olan yangınlarda, en az 1 metre mesafede kullanıma uygundur.",
        "features": "Özellikler: - Silindir: Alaşımlı çelik, derin çekme, klipsli, epoksipolyester toz boya, Kırmızı RAL 3000. - Söndürücü Madde: ABC Tozu - MAP %20. - İtici Gaz: Nemi alınmış hava veya N2. - Vana: M. 58x2, hafif alüminyum alaşım AA6061 gövde, harici çek valfli, kırmızı RAL 3000 boyalı kollar. - Kullanım: ABC yangın sınıflandırması (katı maddeler, yanıcı sıvılar, yanıcı gazlar).",
        "technical_details": "Teknik Özellikler: - Yangın Dayanımı: 34 A 233 BC - Söndürücü Madde: EPW 18462 (ABC Favorit Tertia) - ABC tozu - MAP %20 - İtici Gaz: Nemi alınmış hava veya N2, 20°C'de 15 Bar - Sıcaklık Aralığı: -30°C / +60°C - Nominal Şarj: 6 Kg - Tam Ağırlık: ~ 9,4 Kg - Boyutlar: Yükseklik 550 mm, Çap 160 mm - Boşaltma Süresi:  16,5 sn. - VALF SIKMA TORKU: Minimum 40 Nm, Maksimum 60 Nm - Silindir Basınç Testi: PT 27 bar - Silindir Hacmi: 7,8 L. - Emniyet Valfi: 22 ile 27 bar arasında ayarlanmıştır - Silindir Malzemesi: Alaşımlı çelik - Dış/İç İşlem: Kum püskürtme ve epoksipolyester toz boya, Kırmızı Ral 3000 rengi",
        "image": "equipment/kuru-kimyevi-toz.png",
    },
}
MONTHLY_CONTROL_ITEMS = [
    ("item_1", "17.M.1001.A.1 YSC Konumu Değiştirilmemiş (belirlenen yerde duruyor)"),
    ("item_2", "17.M.1001.A.2 YSC Kullanım Talimatı Kolay Okunabiliyor"),
    ("item_3", "17.M.1001.A.3 YSC Mühür ve Basınç Göstergesinin uygunluğu"),
    ("item_4", "17.M.1001.A.4 YSC Doluluk Durumu (el ile tartarak kontrol edilmeli)"),
    ("item_5", "17.M.1001.A.5 YSC Paslanmamış ve Nozulda Tıkanıklık veya sızdırma yok"),
    ("item_6", "17.M.1001.A.6 YSC Manometresinden Okunan Basınç Kabul Edilebilir aralıkta"),
]
CONTROL_FORM_ITEMS = [
    ("check_a", "a) BELİRLENEN YERE KONULDUĞU"),
    ("check_b", "b) KULLANMA TALİMATLARI GÖRÜNÜMÜ"),
    ("check_c", "c) KULLANMA TALİMATI OKUNURLUĞU"),
    ("check_d", "d) MÜHÜR VE BASINÇ GÖSTERGESİ UYGUNLUĞU"),
    ("check_e", "e) DOLULUK DURUMU (TARTARAK VEYA ELLE)"),
    ("check_f", "f) NOZUL UYGUNLUĞU (PASLANMA VB.)"),
    ("check_g", "g) BASINÇ GÖSTERGESİ İŞLEVSELLİĞİ"),
]
ELECTRICAL_NOTE_SECTIONS = [
    (
        "Firma Bilgileri",
        [
            "ISG-KATIP Sozlesme ID",
            "SGK Sicil Numarasi",
            "Periyodik Kontrol Metodu ve Kapsami",
        ],
    ),
    (
        "Detay Bilgiler",
        [
            "Tesise ait proje var mi",
            "Tek hat semasi var mi",
            "Kontrol nedeni",
            "Topraklayici tipi",
            "Yapi cinsi",
            "Son kontrol tarihi",
            "Faz iletkenlerinin sayisi ve tipi",
            "Temel topraklama direnci",
            "Ilave topraklama elektrotu detaylari",
            "Sistem topraklama iletkeni ve kesiti",
            "Ana espotansiyel iletkeni ve kesiti",
            "Besleme kaynagi karakteristikleri",
            "Ana RCD anma akimi",
            "Ana kesici karakteristikleri",
            "Ana RCD test akimi ve suresi",
        ],
    ),
    (
        "Tespit Edilen Bilgiler",
        [
            "Tesisatta kapsamli degisiklik var mi (>%20)",
            "Asiri gerilim koruma cihazlari kullanilmis mi",
            "Dogrudan dokunmaya karsi koruma onlemleri",
            "Bir onceki periyodik kontrol etiketi var mi",
            "Tespit edilen bilgiler",
        ],
    ),
    (
        "Termal Kamera ve Olcum Aletleri",
        [
            "Termal Kamera 1",
            "Termal Kamera 2",
            "Olcum Aleti 1",
            "Olcum Aleti 2",
            "Termal fotograf tarihi",
            "Termal fotograf no",
            "Kontak gevsakligi isinmasi",
            "Asiri yuk isinmasi",
        ],
    ),
    (
        "Test ve Sonuc",
        [
            "Kontrol Kriterleri ve Testler",
            "Olcum ve Dogrulama Metodu",
            "6.1 Notlari",
            "6.2 Notlari",
            "6.3 Notlari",
            "Kusur Aciklamalari",
            "Ekipman Fotograflari",
            "Genel Notlar",
            "Sonuc ve Kanaat",
            "Yetkili Kisi",
            "Nusha Sayisi",
        ],
    ),
]
FIRE_SUIT_CONTROL_ITEMS = [
    ("item_1", "17. Y .1001.A.1 Gorsel Kumas Kontrol yapildi"),
    ("item_2", "17. Y .1001.A.2 Fonksiyonel Fermuar, Cirt cirtlar ve Dugmeler kontrol edildi"),
    ("item_3", "17. Y .1001.A.3 Yansitici Bantlar Kontrol Edildi"),
    ("item_4", "17. Y .1001.A.4 Elbiselerin Temizligi Kontrol Edildi"),
    ("item_5", "17. Y .1001.A.5 Ic Astarin Durumu, Yalitim Ozelligi ve Dikisler Kontrol Edildi"),
    ("item_6", "17. Y .1001.A.6 Servis Etiketi Ekipmana Yapistirildi"),
]
HELMET_CONTROL_ITEMS = [
    ("item_1", "17.B.1001.A.1 Dis kisimda gorsel kontrol yapildi"),
    ("item_2", "17.B.1001.A.2 Ic kisimda gorsel kontrol yapildi"),
    ("item_3", "17.B.1001.A.3 Ayar mekanizmasinda kontrol yapildi"),
    ("item_4", "17.B.1001.A.4 Vizor ve goz korumasinda kontrol yapildi"),
    ("item_5", "17.B.1001.A.5 Boyun koruyucu kontrolu yapildi"),
    ("item_6", "17.B.1001.A.6 Servis etiketi ekipmana yapistirildi"),
]
AXE_CONTROL_ITEMS = [
    ("item_1", "17.YB.1001.A.1 Baltanin ahsap veya yalitkan sapi kontrol edildi"),
    ("item_2", "17.YB.1001.A.2 Metal kismi kontrol edildi"),
    ("item_3", "17.YB.1001.A.3 Agiz kismi kontrol edildi"),
    ("item_4", "17.YB.1001.A.4 Bulundugu yerde ulasilabilir durumda mi"),
    ("item_5", "17.YB.1001.A.5 Servis etiketi ekipmana yapistirildi"),
]
SCBA_CONTROL_ITEMS = [
    ("item_1", "17.S.1001.A.1 Yüz Maskesi Kontrol Edildi"),
    ("item_2", "17.S.1001.A.2 Solunum Valfi Kontrol Edildi"),
    ("item_3", "17.S.1001.A.3 Regülatör Ünitesi Kontrol Edildi"),
    ("item_4", "17.S.1001.A.4 Kemer Kontrol Edildi"),
    ("item_5", "17.S.1001.A.5 Silindir Kontrol Edildi"),
    ("item_6", "17.S.1001.A.6 NOZUL UYGUNLUĞU (PASLANMA VB.)"),
    ("item_7", "17.S.1001.A.7 Servis Etiketi Ekipmana Yapıştırıldı"),
]
EEBD_CONTROL_ITEMS = [
    ("item_1", "17.E.1001.A.1 Yuz maskesi kontrol edildi"),
    ("item_2", "17.E.1001.A.2 Solunum valfi kontrol edildi"),
    ("item_3", "17.E.1001.A.3 Regulator unitesi kontrol edildi"),
    ("item_4", "17.E.1001.A.4 Kemer kontrol edildi"),
    ("item_5", "17.E.1001.A.5 Silindir kontrol edildi"),
    ("item_6", "17.E.1001.A.6 Servis etiketi ekipmana yapistirildi"),
]
AIR_CYLINDER_CONTROL_ITEMS = [
    ("item_1", "17.C.1001.A.1 Valfi kontrol edildi"),
    ("item_2", "17.C.1001.A.2 Silindir kontrol edildi"),
    ("item_3", "17.C.1001.A.3 Servis etiketi ekipmana yapistirildi"),
]
CABINET_CONTROL_ITEMS = [
    ("item_1", "17.YD.1001.A.1 Erisebilirlik Kontrol Edildi (Dolap onu acik mi, istif veya engel malz. var mi)"),
    ("item_2", "17.YD.1001.A.2 Levhalar Kontrol Edildi (Yangin dolabi isareti ve talimati mevcut mu)"),
    ("item_3", "17.YD.1001.A.3 Kapak ve Kilit Kontrolu Yapildi (Kapak rahat aciliyor mu, kilit saglam mi)"),
    ("item_4", "17.YD.1001.A.4 Dolap Dis Yuzey Kontrol Edildi (Paslanma veya boya kabarmasi var mi)"),
    ("item_5", "17.YD.1001.A.5 Makara Kontrol Edildi (Kolayca acilabiliyor mu)"),
    ("item_6", "17.YD.1001.A.6 Hortum Kontrol Edildi (Catlama, kirilma, sertlesme veya kacak var mi)"),
    ("item_7", "17.YD.1001.A.7 Baglanti Rekorlari Kontrol Edildi (Hortum vana ve lans baglantilari siki mi)"),
    ("item_8", "17.YD.1001.A.8 Dolap ici ve disi temizlik kontrolu yapildi"),
    ("item_9", "17.YD.1001.A.9 Vana kontrol edildi (Vana kolu rahat donuyor mu, kacak veya sizdirma var mi)"),
    ("item_10", "17.YD.1001.A.10 Basinc Kontrol Edildi (Statik ve dinamik basinc degerleri uygun mu) (min.4 bar)"),
    ("item_11", "17.YD.1001.A.11 Lans Kontrol Edildi (Jet/Spray/Kapali konumlari islevsel mi)"),
    ("item_12", "17.YD.1001.A.12 Servis etiketi ekipmana yapistirildi"),
]
FOAM_CABINET_CONTROL_ITEMS = [
    ("item_1", "17.KYD.1001.A.1 Erisebilirlik Kontrol Edildi (Dolap onu acik mi, istif veya engel malz. var mi)"),
    ("item_2", "17.KYD.1001.A.2 Levhalar Kontrol Edildi (Kopuklu yangin dolabi isareti ve talimati mevcut mu)"),
    ("item_3", "17.KYD.1001.A.3 Kapak ve Kilit Kontrolu Yapildi (Kapak rahat aciliyor mu, kilit saglam mi)"),
    ("item_4", "17.KYD.1001.A.4 Dolap Dis Yuzey Kontrol Edildi (Paslanma veya boya kabarmasi var mi)"),
    ("item_5", "17.KYD.1001.A.5 Makara Kontrol Edildi (Kolayca acilabiliyor mu)"),
    ("item_6", "17.KYD.1001.A.6 Hortum Kontrol Edildi (Catlama, kirilma, sertlesme veya kacak var mi)"),
    ("item_7", "17.KYD.1001.A.7 Baglanti Rekorlari Kontrol Edildi (Hortum vana ve lans baglantilari siki mi)"),
    ("item_8", "17.KYD.1001.A.8 Kopuk Doluluk Orani Kontrol Edildi (Seviye gostergesi kontrolu)"),
    ("item_9", "17.KYD.1001.A.9 Dolap ici ve disi temizlik kontrolu yapildi"),
    ("item_10", "17.KYD.1001.A.10 Kopuk Oranlayici Ayarlari Kontrol Edildi (Mix ayari dogru yuzde mi %1,%3 veya %6)"),
    ("item_11", "17.KYD.1001.A.11 Vana kontrol edildi (Ana su giris vanasi ve kopuk vanasi islevsel mi)"),
    ("item_12", "17.KYD.1001.A.12 Basinc Kontrol Edildi (Sistem calisma basinci kopuk olusumu icin yeterli mi) (min.5-6 bar)"),
    ("item_13", "17.KYD.1001.A.13 Kopuk Lans Kontrol Edildi (Kopuk yapici ozel lans saglam mi)"),
    ("item_14", "17.KYD.1001.A.14 Servis etiketi ekipmana yapistirildi"),
]
HYDRANT_CONTROL_ITEMS = [
    ("item_1", "17.H.1001.A.1 Erisebilirlik Kontrol Edildi (Hidrant cevresinde arac, malzeme engeli var mi)"),
    ("item_2", "17.H.1001.A.2 Gorunurlugu kontrol edildi (Hidrantin kirmizi boyasi canli mi, yonlendirme levhalari var mi)"),
    ("item_3", "17.H.1001.A.3 Kapak Kontrolu Yapildi (Cikis agzindaki kor tapalar/kapaklar takili mi, zincirleri saglam mi)"),
    ("item_4", "17.H.1001.A.4 Genel Dis Yuzey Kontrol Yapildi (Govdede catlak, korozyon veya darbe izi var mi)"),
    ("item_5", "17.H.1001.A.5 Makara Kontrol Edildi (Kolayca acilabiliyor mu)"),
    ("item_6", "17.H.1001.A.6 Acma kapama mili kontrol edildi (Hidrant anahtari ile mil rahatca donuyor mu)"),
    ("item_7", "17.H.1001.A.7 Vana Sizdirmazligi Kontrol Edildi (Hidrant kapaliyken cikis agzindan veya govde altindan su sizintisi var mi)"),
    ("item_8", "17.H.1001.A.8 Cikis agizlari kontrol edildi (Rekor dislerinde veya tirnaklarinda asinma veya deformasyon var mi)"),
    ("item_9", "17.H.1001.A.9 Servis etiketi ekipmana yapistirildi"),
]
KASIK_SEDYE_CONTROL_ITEMS = [
    ("item_1", "17.KS.1001.A.1 Aluminyum yuzeyler catlak, egilme ve deformasyon yonunden kontrol edildi"),
    ("item_2", "17.KS.1001.A.2 Sedyenin ortadan acilma-kapanma mekanizmasi calisiyor, kilitler tam oturuyor"),
    ("item_3", "17.KS.1001.A.3 Boyaya yatirma mekanizmasi calisir durumda, sabitleme pimleri ve baglama tokalari caliisiyor"),
    ("item_4", "17.KS.1001.A.4 Sedye Hijyen kontrolu yapildi"),
    ("item_5", "17.KS.1001.A.5 Bulundugu Yerde Ulasilabilir Durumda"),
    ("item_6", "17.KS.1001.A.6 Servis Etiketi Ekipmana Yapistirildi"),
]
FIRE_BOOTS_CONTROL_ITEMS = [
    ("item_1", "17.YÇ.1001.A.1 Çizme yüzeyinde yırtık, çatlak, delik veya aşırı aşınma kontrolü"), 
    ("item_2", "17.YÇ.1001.A.2 Çizme tabanında kaymazlık özelliği, çelik burun veya ara taban deformasyon kontrolü"), 
    ("item_3", "17.YÇ.1001.A.3 Çizme içindeki astarın zarar görüp görmediği, temizliği ve hijyeni kontrol edildi"), 
    ("item_4", "17.YÇ.1001.A.4 Bulunduğu Yerde Ulaşılabilir Durumda mı"), 
    ("item_5", "17.YÇ.1001.A.5 Servis Etiketi Ekipmana Yapıştırıldı"), 
]
FALL_ARRESTER_CONTROL_ITEMS = [
    ("item_1", "17.DD.1001.A.1 Gövde kısımda; kırık, çatlak veya darbe izi var mı, muhafaza sağlam mı"),
    ("item_2", "17.DD.1001.A.2 Çelik halat/kolonda kesilme, ezilme, lif kopması, paslanma var mı"),
    ("item_3", "17.DD.1001.A.3 Hat serbestçe çekilip geri sarıyor mu, takılma veya ani boşalma var mı"),
    ("item_4", "17.DD.1001.A.4 Ani çekmede sistem kilitleniyor mu, kilitleme gecikmesi var mı"),
    ("item_5", "17.DD.1001.A.5 Bulunduğu yerde ulaşılabilir durumda ve uygun koşullarda mı"),
    ("item_6", "17.DD.1001.A.6 Servis Etiketi Ekipmana Yapıştırıldı"),
]
SPINE_BOARD_CONTROL_ITEMS = [
    ("item_1", "17.OT.1001.A.1 Tahta yüzeylerde çatlak, kırık, eğilme ve deformasyon kontrolü"),
    ("item_2", "17.OT.1001.A.2 Tutma yerleri ergonomik ve sağlam mı"),
    ("item_3", "17.OT.1001.A.3 Kayış bağlantı noktaları, aparatlar ve tokalar düzgün çalışıyor mu"),
    ("item_4", "17.OT.1001.A.4 Sedye hijyenik durumda mı"),
    ("item_5", "17.OT.1001.A.5 Bulunduğu yerde ulaşılabilir durumda mı"),
    ("item_6", "17.OT.1001.A.6 Servis etiketi ekipmana yapıştırıldı"),
]
SAFETY_HARNESS_CONTROL_ITEMS = [
    ("item_1", "17.PEK.1001.A.1 Sapan gövdesinde kesik, yırtık, aşınma ve dikiş kontrolü"),
    ("item_2", "17.PEK.1001.A.2 Metal akşamlarda çatlak, paslanma, deformasyon ve kanca emniyet kilidi kontrolü"),
    ("item_3", "17.PEK.1001.A.3 Sırt, göğüs ve yan bağlantı noktaları sağlam mı"),
    ("item_4", "17.PEK.1001.A.4 Kemer temiz ve kullanılabilir durumda mı"),
    ("item_5", "17.PEK.1001.A.5 Bulunduğu yerde ulaşılabilir ve uygun saklama koşullarında mı"),
    ("item_6", "17.PEK.1001.A.6 Servis etiketi ekipmana yapıştırıldı"),
]
STRETCHER_SLING_CONTROL_ITEMS = [
    ("item_1", "17.TS.1001.A.1 Sapan gövdesinde kesik, yırtık, lif kopması ve aşınma kontrolü"),
    ("item_2", "17.TS.1001.A.2 Metal akşamlarda çatlak, paslanma, korozyon ve emniyet mandalı kontrolü"),
    ("item_3", "17.TS.1001.A.3 Sedye ile bağlantı noktaları uygun mu, dengesiz yükleme riski var mı"),
    ("item_4", "17.TS.1001.A.4 Sapan temiz ve kullanılabilir durumda mı"),
    ("item_5", "17.TS.1001.A.5 Bulunduğu yerde ulaşılabilir durumda mı"),
    ("item_6", "17.TS.1001.A.6 Servis etiketi ekipmana yapıştırıldı"),
]
BASKET_STRETCHER_CONTROL_ITEMS = [
    ("item_1", "17.SS.1001.A.1 Gövde, kenar korumaları ve iç yüzeyde çatlak, deformasyon veya kırık kontrolü"),
    ("item_2", "17.SS.1001.A.2 Halat bağlantı noktaları, karabina ve zincirlerde paslanma veya gevşeme kontrolü"),
    ("item_3", "17.SS.1001.A.3 Emniyet kemerleri, sabitleme kayışları ve baş sabitleyici aparat kontrolü"),
    ("item_4", "17.SS.1001.A.4 Sedye hijyenik durumda mı"),
    ("item_5", "17.SS.1001.A.5 Bulunduğu yerde ulaşılabilir durumda mı"),
    ("item_6", "17.SS.1001.A.6 Servis etiketi ekipmana yapıştırıldı"),
]
ASSET_PROFILES = {
    "Yangin Sondurme Cihazi": {
        "label": "Yangin Sondurme Cihazi",
        "type_label": "Tip",
        "class_label": "YSC Sinifi",
        "brand_label": "YSC Uretici",
        "owner_label": "Firma Yetkilisi",
        "last_service_label": "Son Bakim",
        "next_service_label": "Sonraki Bakim",
        "service_input_label": "Son bakim tarihi",
        "next_service_input_label": "Sonraki bakim tarihi",
        "show_weight": True,
        "show_pressure": True,
        "show_hydrostatic": True,
        "control_form_enabled": True,
        "fixed_type": None,
        "monthly_control_items": MONTHLY_CONTROL_ITEMS,
        "control_form_items": CONTROL_FORM_ITEMS,
    },
    "Yangin Elbisesi": {
        "label": "Yangin Elbisesi",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "size_label": "Beden (S/M/L/XL)",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Yangin Elbisesi",
        "monthly_control_items": FIRE_SUIT_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Yangin Bareti": {
        "label": "Yangin Bareti",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Yangin Bareti",
        "monthly_control_items": HELMET_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Yangin Baltasi": {
        "label": "Yangin Baltasi",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Yangin Baltasi",
        "monthly_control_items": AXE_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "SCBA": {
        "label": "SCBA",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Dolum Tarihi",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Dolum tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": True,
        "show_pressure": False,
        "show_hydrostatic": True,
        "control_form_enabled": False,
        "fixed_type": "SCBA",
        "monthly_control_items": SCBA_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "EEBD": {
        "label": "EEBD",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Dolum Tarihi",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Dolum tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": True,
        "show_pressure": False,
        "show_hydrostatic": True,
        "control_form_enabled": False,
        "fixed_type": "EEBD",
        "monthly_control_items": EEBD_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Hava Tupu": {
        "label": "Hava Tupu",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Dolum Tarihi",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Dolum tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": True,
        "show_pressure": False,
        "show_hydrostatic": True,
        "control_form_enabled": False,
        "fixed_type": "Hava Tupu",
        "monthly_control_items": AIR_CYLINDER_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Yangin Sondurme Dolabi": {
        "label": "Yangin Sondurme Dolabi",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Yangin Sondurme Dolabi",
        "monthly_control_items": CABINET_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Kopuklu Yangin Sondurme Dolabi": {
        "label": "Kopuklu Yangin Sondurme Dolabi",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Kopuklu Yangin Sondurme Dolabi",
        "monthly_control_items": FOAM_CABINET_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Yangin Hidranti": {
        "label": "Yangin Hidranti",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Yangin Hidranti",
        "monthly_control_items": HYDRANT_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Elektrik Ic Tesisati": {
        "label": "Elektrik Ic Tesisati",
        "type_label": "Rapor Tipi",
        "class_label": "Sebeke Tipi",
        "brand_label": "Enerji Saglayan Kurulus",
        "owner_label": "Rapor Numarasi",
        "last_service_label": "Rapor Tarihi",
        "next_service_label": "Bir Sonraki Kontrol",
        "service_input_label": "Rapor tarihi",
        "next_service_input_label": "Bir sonraki periyodik kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Elektrik Ic Tesisati",
        "monthly_control_items": [],
        "control_form_items": [],
    },
    "Kasik Sedye": {
        "label": "Kasik Sedye",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": True,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": False,
        "fixed_type": "Kasik Sedye",
        "monthly_control_items": KASIK_SEDYE_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Yangin Cizmesi": {
        "label": "Yangin Cizmesi", 
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi", 
        "brand_label": "Marka", 
        "owner_label": "Ekipman Yetkilisi", 
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "service_input_label": "Kontrol tarihi",
        "next_service_input_label": "Sonraki kontrol tarihi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": True,
        "fixed_type": "Yangin Cizmesi",
        "monthly_control_items": FIRE_BOOTS_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Geri Sarimli Dusus Durdurucu": {
        "label": "Geri Sarimli Dusus Durdurucu",
        "type_label": "Ekipman Tipi",
        "class_label": "Kategori / Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": True,
        "fixed_type": "Geri Sarimli Dussundurucu",
        "monthly_control_items": FALL_ARRESTER_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Omurga Tahtasi": {
        "label": "Omurga Tahtasi",
        "type_label": "Sedye Tipi",
        "class_label": "Kategori",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "last_service_label": "Son Kontrol",
        "next_service_label": "Sonraki Kontrol",
        "show_weight": True, # Belgeye göre ağırlık görünebilir
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": True,
        "fixed_type": "Omurga Tahtasi",
        "monthly_control_items": SPINE_BOARD_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Parasut Tipi Emniyet Kemeri": {
        "label": "Parasut Tipi Emniyet Kemeri",
        "type_label": "Kemer Tipi",
        "class_label": "Cinsi",
        "brand_label": "Marka",
        "owner_label": "Ekipman Yetkilisi",
        "show_weight": False,
        "show_pressure": False,
        "show_hydrostatic": False,
        "control_form_enabled": True,
        "fixed_type": "Parasut Tipi Kemer",
        "monthly_control_items": SAFETY_HARNESS_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Sedye Tasima Sapani": {
        "label": "Sedye Tasima Sapani",
        "type_label": "Sapan Tipi",
        "brand_label": "Marka",
        "show_weight": False,
        "show_pressure": False,
        "control_form_enabled": True,
        "fixed_type": "Sedye Tasima Sapani",
        "monthly_control_items": STRETCHER_SLING_CONTROL_ITEMS,
        "control_form_items": [],
    },
    "Sepet Sedye": {
        "label": "Sepet Sedye",
        "type_label": "Sedye Tipi",
        "brand_label": "Marka",
        "show_weight": True,
        "show_pressure": False,
        "control_form_enabled": True,
        "fixed_type": "Sepet Sedye",
        "monthly_control_items": BASKET_STRETCHER_CONTROL_ITEMS,
        "control_form_items": [],
    }
}
MONTH_LABELS = [
    (1, "OCAK"),
    (2, "ŞUBAT"),
    (3, "MART"),
    (4, "NİSAN"),
    (5, "MAYIS"),
    (6, "HAZİRAN"),
    (7, "TEMMUZ"),
    (8, "AĞUSTOS"),
    (9, "EYLÜL"),
    (10, "EKİM"),
    (11, "KASIM"),
    (12, "ARALIK"),
]


def build_database_url() -> str:
    raw_url = os.environ.get("DATABASE_URL", DEFAULT_SQLITE_URL)
    if raw_url.startswith("postgres://"):
        return raw_url.replace("postgres://", "postgresql+psycopg://", 1)
    if raw_url.startswith("postgresql://"):
        return raw_url.replace("postgresql://", "postgresql+psycopg://", 1)
    return raw_url


DATABASE_URL = build_database_url()
CONNECT_ARGS = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine: Engine = create_engine(
    DATABASE_URL,
    future=True,
    pool_pre_ping=True,
    connect_args=CONNECT_ARGS,
)

metadata = MetaData()

users = Table(
    "users",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("username", String(128), nullable=False, unique=True),
    Column("full_name", String(255), nullable=False),
    Column("password_hash", String(255), nullable=False),
    Column("is_admin", Boolean, nullable=False, default=False),
    Column("is_active", Boolean, nullable=False, default=True),
    Column("created_at", String(32), nullable=False),
    Column("updated_at", String(32), nullable=False),
)

companies = Table(
    "companies",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("public_id", String(32), nullable=False, unique=True),
    Column("slug", String(255), nullable=False, unique=True),
    Column("name", String(255), nullable=False, unique=True),
    Column("address", String(255), nullable=False),
    Column("contact_name", String(255), nullable=False),
    Column("phone", String(64), nullable=False, default="-"),
    Column("email", String(255), nullable=False, default="-"),
    Column("created_at", String(32), nullable=False),
    Column("updated_at", String(32), nullable=False),
)

extinguishers = Table(
    "extinguishers",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("public_id", String(32), nullable=False, unique=True),
    Column("serial_number", String(128), nullable=False),
    Column("company_id", Integer, ForeignKey("companies.id")),
    Column("company_name", String(255), nullable=False),
    Column("location_detail", String(255), nullable=False),
    Column("weight_kg", Float, nullable=False),
    Column("asset_category", String(128), nullable=False, default=DEFAULT_ASSET_CATEGORY),
    Column("extinguisher_type", String(255), nullable=False),
    Column("fire_class", String(255)),
    Column("manufacturer", String(255)),
    Column("hydrostatic_test_date", String(32)),
    Column("company_address", String(255)),
    Column("company_contact", String(255)),
    Column("pressure_status", String(255)),
    Column("notes", String),
    Column("last_service_date", String(32)),
    Column("next_service_date", String(32)),
    Column("created_at", String(32), nullable=False),
    Column("updated_at", String(32), nullable=False),
    UniqueConstraint("asset_category", "serial_number", name="uq_extinguishers_asset_category_serial_number"),
)

service_logs = Table(
    "service_logs",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("extinguisher_id", Integer, ForeignKey("extinguishers.id"), nullable=False),
    Column("service_date", String(32), nullable=False),
    Column("technician_name", String(255), nullable=False),
    Column("operation_summary", String, nullable=False),
    Column("pressure_status", String(255)),
    Column("notes", String),
    Column("created_at", String(32), nullable=False),
)

monthly_inspections = Table(
    "monthly_inspections",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("extinguisher_id", Integer, ForeignKey("extinguishers.id"), nullable=False),
    Column("inspection_date", String(32), nullable=False),
    Column("inspector_name", String(255), nullable=False),
    Column("item_1", Boolean, nullable=False),
    Column("item_2", Boolean, nullable=False),
    Column("item_3", Boolean, nullable=False),
    Column("item_4", Boolean, nullable=False),
    Column("item_5", Boolean, nullable=False),
    Column("item_6", Boolean, nullable=False),
    Column("item_7", Boolean, nullable=False, default=False),
    Column("item_8", Boolean, nullable=False, default=False),
    Column("item_9", Boolean, nullable=False, default=False),
    Column("item_10", Boolean, nullable=False, default=False),
    Column("item_11", Boolean, nullable=False, default=False),
    Column("item_12", Boolean, nullable=False, default=False),
    Column("item_13", Boolean, nullable=False, default=False),
    Column("item_14", Boolean, nullable=False, default=False),
    Column("check_a", Boolean, nullable=False, default=False),
    Column("check_b", Boolean, nullable=False, default=False),
    Column("check_c", Boolean, nullable=False, default=False),
    Column("check_d", Boolean, nullable=False, default=False),
    Column("check_e", Boolean, nullable=False, default=False),
    Column("check_f", Boolean, nullable=False, default=False),
    Column("check_g", Boolean, nullable=False, default=False),
    Column("notes", String),
    Column("created_at", String(32), nullable=False),
)

notification_logs = Table(
    "notification_logs",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("extinguisher_id", Integer, ForeignKey("extinguishers.id"), nullable=False),
    Column("notification_type", String(64), nullable=False),
    Column("target_date", String(32), nullable=False),
    Column("recipient_email", String(255), nullable=False),
    Column("created_at", String(32), nullable=False),
    UniqueConstraint("extinguisher_id", "notification_type", "target_date", name="uq_notification_once_per_target"),
)

metadata.create_all(engine)


def coerce_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "strftime"):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except Exception:
        try:
            return datetime.strptime(str(value), "%d.%m.%Y").date()
        except Exception:
            return None


def is_real_email(value: str | None) -> bool:
    email = (value or "").strip()
    return bool(email and email != "-" and "@" in email)


def send_email_message(subject: str, body: str, recipient: str) -> bool:
    if not (SMTP_USERNAME and SMTP_PASSWORD and SMTP_FROM and recipient):
        return False

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = SMTP_FROM
    message["To"] = recipient
    message.set_content(body)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
        smtp.send_message(message)
    return True


def get_due_soon_assets(window_days: int = ALERT_WINDOW_DAYS) -> list[dict]:
    today = datetime.now().date()
    deadline = today + timedelta(days=window_days)
    rows = fetch_all(
        select(
            extinguishers,
            companies.c.email.label("company_email"),
            companies.c.contact_name.label("company_contact_name"),
        )
        .select_from(extinguishers.join(companies, extinguishers.c.company_id == companies.c.id))
        .where(extinguishers.c.next_service_date.is_not(None))
        .order_by(extinguishers.c.next_service_date.asc(), extinguishers.c.company_name.asc())
    )
    due_assets = []
    for row in rows:
        target_date = coerce_date(row.get("next_service_date"))
        if not target_date:
            continue
        if not (today <= target_date <= deadline):
            continue
        enriched = dict(row)
        enriched["days_left"] = (target_date - today).days
        enriched["target_date"] = target_date.isoformat()
        due_assets.append(enriched)
    return due_assets


def process_due_soon_notifications(window_days: int = ALERT_WINDOW_DAYS) -> list[dict]:
    due_assets = get_due_soon_assets(window_days)
    if not due_assets:
        return []

    now = datetime.now().isoformat(timespec="seconds")
    with engine.begin() as connection:
        for asset in due_assets:
            recipient = (asset.get("company_email") or "").strip()
            if not is_real_email(recipient):
                continue

            notification_type = f"next_service_{window_days}_days"
            existing = connection.execute(
                select(notification_logs.c.id)
                .where(notification_logs.c.extinguisher_id == asset["id"])
                .where(notification_logs.c.notification_type == notification_type)
                .where(notification_logs.c.target_date == asset["target_date"])
                .limit(1)
            ).fetchone()
            if existing:
                continue

            subject = f"Vesta Yangin - Yaklasan kontrol bildirimi ({asset['serial_number']})"
            body = (
                f"Merhaba {asset.get('company_contact_name') or asset.get('company_contact') or 'Yetkili'},\n\n"
                f"{asset.get('company_name')} firmasina ait {asset.get('asset_category') or 'ekipman'} kaydinin "
                f"sonraki kontrol tarihi {asset.get('next_service_date')} olarak gorunuyor.\n"
                f"Cihaz seri no: {asset.get('serial_number')}\n"
                f"Bulundugu yer: {asset.get('location_detail') or '-'}\n"
                f"Kalan sure: {asset['days_left']} gun\n\n"
                "Kontrol planlamasi icin bizimle iletisime gecebilirsiniz.\n\n"
                "Vesta Yangin"
            )

            try:
                sent = send_email_message(subject, body, recipient)
            except Exception:
                sent = False

            if sent:
                connection.execute(
                    insert(notification_logs).values(
                        extinguisher_id=asset["id"],
                        notification_type=notification_type,
                        target_date=asset["target_date"],
                        recipient_email=recipient,
                        created_at=now,
                    )
                )

    return due_assets


def ensure_monthly_inspection_columns() -> None:
    extra_boolean_columns = [
        "item_7",
        "item_8",
        "item_9",
        "item_10",
        "item_11",
        "item_12",
        "item_13",
        "item_14",
    ]
    with engine.begin() as connection:
        if engine.dialect.name == "sqlite":
            existing_rows = connection.execute(text("PRAGMA table_info(monthly_inspections)")).fetchall()
            existing = {row[1] for row in existing_rows}
            for column_name in extra_boolean_columns:
                if column_name not in existing:
                    connection.execute(
                        text(
                            f"ALTER TABLE monthly_inspections ADD COLUMN {column_name} BOOLEAN NOT NULL DEFAULT 0"
                        )
                    )
        else:
            existing_rows = connection.execute(
                text(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = 'monthly_inspections'"
                )
            ).fetchall()
            existing = {row[0] for row in existing_rows}
            for column_name in extra_boolean_columns:
                if column_name not in existing:
                    connection.execute(
                        text(
                            f"ALTER TABLE monthly_inspections ADD COLUMN {column_name} BOOLEAN NOT NULL DEFAULT FALSE"
                        )
                    )


ensure_monthly_inspection_columns()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-key")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)  # type: ignore[assignment]
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "vestaadmin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "degistir-beni-2026")
DEFAULT_USER_PASSWORD = os.environ.get("DEFAULT_USER_PASSWORD", "Vesta123!")
DEFAULT_USERS = [
    {
        "username": ADMIN_USERNAME,
        "full_name": "Admin",
        "password": ADMIN_PASSWORD,
        "is_admin": True,
    },
    {
        "username": "yunus.emre.duman",
        "full_name": "Yunus Emre Duman",
        "password": DEFAULT_USER_PASSWORD,
        "is_admin": False,
    },
    {
        "username": "atil.bati",
        "full_name": "Atıl Batı",
        "password": DEFAULT_USER_PASSWORD,
        "is_admin": False,
    },
    {
        "username": "mustafa.kilic",
        "full_name": "Mustafa Kiliç",
        "password": DEFAULT_USER_PASSWORD,
        "is_admin": False,
    },
]


class RotatedParagraph(Flowable):
    def __init__(self, text: str, style, width: float, height: float) -> None:
        super().__init__()
        self.text = text
        self.style = style
        self.width = width
        self.height = height
        self._paragraph = Paragraph(text, style)

    def wrap(self, availWidth, availHeight):
        return self.width, self.height

    def draw(self):
        self.canv.saveState()
        self.canv.translate(self.width, 0)
        self.canv.rotate(90)
        paragraph_width = self.height - 4
        paragraph_height = self.width - 4
        self._paragraph.wrapOn(self.canv, paragraph_width, paragraph_height)
        self._paragraph.drawOn(self.canv, 2, 2)
        self.canv.restoreState()


def seed_default_users() -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with engine.begin() as connection:
        existing = {
            row[0]
            for row in connection.execute(select(users.c.username)).all()
        }
        for user in DEFAULT_USERS:
            if user["username"] in existing:
                continue
            connection.execute(
                insert(users).values(
                    username=user["username"],
                    full_name=user["full_name"],
                    password_hash=generate_password_hash(user["password"]),
                    is_admin=user["is_admin"],
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                )
            )


def run_schema_migrations() -> None:
    if DATABASE_URL.startswith("sqlite"):
        with engine.begin() as connection:
            company_tables = {
                row[0]
                for row in connection.execute(
                    text("SELECT name FROM sqlite_master WHERE type='table' AND name='companies'")
                ).fetchall()
            }
            if "companies" not in company_tables:
                connection.execute(
                    text(
                        """
                        CREATE TABLE companies (
                            id INTEGER PRIMARY KEY,
                            public_id VARCHAR(32) NOT NULL UNIQUE,
                            slug VARCHAR(255) NOT NULL UNIQUE,
                            name VARCHAR(255) NOT NULL UNIQUE,
                            address VARCHAR(255) NOT NULL,
                            contact_name VARCHAR(255) NOT NULL,
                            phone VARCHAR(64) NOT NULL DEFAULT '-',
                            email VARCHAR(255) NOT NULL DEFAULT '-',
                            created_at VARCHAR(32) NOT NULL,
                            updated_at VARCHAR(32) NOT NULL
                        )
                        """
                    )
                )
            columns = {
                row[1] for row in connection.execute(text("PRAGMA table_info(users)")).fetchall()
            }
            if "is_active" not in columns:
                connection.execute(text("ALTER TABLE users ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT 1"))
            extinguisher_columns = {
                row[1] for row in connection.execute(text("PRAGMA table_info(extinguishers)")).fetchall()
            }
            for column_name, column_def in [
                ("company_id", "INTEGER"),
                ("fire_class", "TEXT"),
                ("manufacturer", "TEXT"),
                ("hydrostatic_test_date", "TEXT"),
                ("company_address", "TEXT"),
                ("company_contact", "TEXT"),
                ("asset_category", "TEXT"),
            ]:
                if column_name not in extinguisher_columns:
                    connection.execute(text(f"ALTER TABLE extinguishers ADD COLUMN {column_name} {column_def}"))
            company_columns = {
                row[1] for row in connection.execute(text("PRAGMA table_info(companies)")).fetchall()
            }
            if "public_id" not in company_columns:
                connection.execute(text("ALTER TABLE companies ADD COLUMN public_id TEXT"))
            if "slug" not in company_columns:
                connection.execute(text("ALTER TABLE companies ADD COLUMN slug TEXT"))
            if "phone" not in company_columns:
                connection.execute(text("ALTER TABLE companies ADD COLUMN phone TEXT NOT NULL DEFAULT '-'"))
            if "email" not in company_columns:
                connection.execute(text("ALTER TABLE companies ADD COLUMN email TEXT NOT NULL DEFAULT '-'"))

            inspection_columns = {
                row[1] for row in connection.execute(text("PRAGMA table_info(monthly_inspections)")).fetchall()
            }
            for column_name in ["check_a", "check_b", "check_c", "check_d", "check_e", "check_f", "check_g"]:
                if column_name not in inspection_columns:
                    connection.execute(
                        text(f"ALTER TABLE monthly_inspections ADD COLUMN {column_name} BOOLEAN NOT NULL DEFAULT 0")
                    )
            connection.execute(
                text(
                    """
                    UPDATE monthly_inspections
                    SET
                        check_a = item_1,
                        check_b = item_2,
                        check_c = item_2,
                        check_d = item_3,
                        check_e = item_4,
                        check_f = item_5,
                        check_g = item_6
                    WHERE extinguisher_id IN (
                        SELECT id FROM extinguishers WHERE asset_category = :ysc_category
                    )
                    """
                ),
                {"ysc_category": DEFAULT_ASSET_CATEGORY},
            )

            connection.execute(
                text(
                    "UPDATE extinguishers SET asset_category = :default_category "
                    "WHERE asset_category IS NULL OR TRIM(asset_category) = ''"
                ),
                {"default_category": DEFAULT_ASSET_CATEGORY},
            )
            extinguisher_sql_row = connection.execute(
                text("SELECT sql FROM sqlite_master WHERE type='table' AND name='extinguishers'")
            ).fetchone()
            extinguisher_sql = (extinguisher_sql_row[0] or "") if extinguisher_sql_row else ""
            needs_serial_migration = "serial_number TEXT NOT NULL UNIQUE" in extinguisher_sql
            if needs_serial_migration:
                connection.execute(text("PRAGMA foreign_keys=OFF"))
                connection.execute(
                    text(
                        """
                        CREATE TABLE extinguishers_new (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            public_id TEXT NOT NULL UNIQUE,
                            serial_number TEXT NOT NULL,
                            company_name TEXT NOT NULL,
                            location_detail TEXT NOT NULL,
                            weight_kg REAL NOT NULL,
                            extinguisher_type TEXT NOT NULL,
                            pressure_status TEXT,
                            notes TEXT,
                            last_service_date TEXT,
                            next_service_date TEXT,
                            created_at TEXT NOT NULL,
                            updated_at TEXT NOT NULL,
                            fire_class TEXT,
                            manufacturer TEXT,
                            hydrostatic_test_date TEXT,
                            company_address TEXT,
                            company_contact TEXT,
                            company_id INTEGER,
                            asset_category TEXT NOT NULL,
                            CONSTRAINT uq_extinguishers_asset_category_serial_number UNIQUE (asset_category, serial_number)
                        )
                        """
                    )
                )
                connection.execute(
                    text(
                        """
                        INSERT INTO extinguishers_new (
                            id, public_id, serial_number, company_name, location_detail, weight_kg,
                            extinguisher_type, pressure_status, notes, last_service_date, next_service_date,
                            created_at, updated_at, fire_class, manufacturer, hydrostatic_test_date,
                            company_address, company_contact, company_id, asset_category
                        )
                        SELECT
                            id, public_id, serial_number, company_name, location_detail, weight_kg,
                            extinguisher_type, pressure_status, notes, last_service_date, next_service_date,
                            created_at, updated_at, fire_class, manufacturer, hydrostatic_test_date,
                            company_address, company_contact, company_id,
                            COALESCE(NULLIF(asset_category, ''), :default_category)
                        FROM extinguishers
                        """
                    ),
                    {"default_category": DEFAULT_ASSET_CATEGORY},
                )
                connection.execute(text("DROP TABLE extinguishers"))
                connection.execute(text("ALTER TABLE extinguishers_new RENAME TO extinguishers"))
                connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_extinguishers_public_id ON extinguishers(public_id)"))
                connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_extinguishers_asset_serial ON extinguishers(asset_category, serial_number)"))
                connection.execute(text("PRAGMA foreign_keys=ON"))
    else:
        with engine.begin() as connection:
            table_exists = connection.execute(
                text(
                    """
                    SELECT table_name
                    FROM information_schema.tables
                    WHERE table_name = 'companies'
                    """
                )
            ).fetchone()
            if table_exists is None:
                connection.execute(
                    text(
                        """
                        CREATE TABLE companies (
                            id SERIAL PRIMARY KEY,
                            public_id VARCHAR(32) NOT NULL UNIQUE,
                            slug VARCHAR(255) NOT NULL UNIQUE,
                            name VARCHAR(255) NOT NULL UNIQUE,
                            address VARCHAR(255) NOT NULL,
                            contact_name VARCHAR(255) NOT NULL,
                            phone VARCHAR(64) NOT NULL DEFAULT '-',
                            email VARCHAR(255) NOT NULL DEFAULT '-',
                            created_at VARCHAR(32) NOT NULL,
                            updated_at VARCHAR(32) NOT NULL
                        )
                        """
                    )
                )
            result = connection.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = 'users' AND column_name = 'is_active'
                    """
                )
            ).fetchone()
            if result is None:
                connection.execute(
                    text("ALTER TABLE users ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT TRUE")
                )
            for column_name in ["company_id", "fire_class", "manufacturer", "hydrostatic_test_date", "company_address", "company_contact", "asset_category"]:
                result = connection.execute(
                    text(
                        f"""
                        SELECT column_name
                        FROM information_schema.columns
                        WHERE table_name = 'extinguishers' AND column_name = '{column_name}'
                        """
                    )
                ).fetchone()
                if result is None:
                    column_type = "INTEGER" if column_name == "company_id" else "TEXT"
                    connection.execute(text(f"ALTER TABLE extinguishers ADD COLUMN {column_name} {column_type}"))
            result = connection.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = 'companies' AND column_name = 'public_id'
                    """
                )
            ).fetchone()
            if result is None:
                connection.execute(text("ALTER TABLE companies ADD COLUMN public_id TEXT"))
            result = connection.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = 'companies' AND column_name = 'slug'
                    """
                )
            ).fetchone()
            if result is None:
                connection.execute(text("ALTER TABLE companies ADD COLUMN slug TEXT"))
            result = connection.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = 'companies' AND column_name = 'phone'
                    """
                )
            ).fetchone()
            if result is None:
                connection.execute(text("ALTER TABLE companies ADD COLUMN phone TEXT NOT NULL DEFAULT '-'"))
            result = connection.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = 'companies' AND column_name = 'email'
                    """
                )
            ).fetchone()
            if result is None:
                connection.execute(text("ALTER TABLE companies ADD COLUMN email TEXT NOT NULL DEFAULT '-'"))
            for column_name in ["check_a", "check_b", "check_c", "check_d", "check_e", "check_f", "check_g"]:
                result = connection.execute(
                    text(
                        f"""
                        SELECT column_name
                        FROM information_schema.columns
                        WHERE table_name = 'monthly_inspections' AND column_name = '{column_name}'
                        """
                    )
                ).fetchone()
                if result is None:
                    connection.execute(
                        text(f"ALTER TABLE monthly_inspections ADD COLUMN {column_name} BOOLEAN NOT NULL DEFAULT FALSE")
                    )
            connection.execute(
                text(
                    """
                    UPDATE monthly_inspections mi
                    SET
                        check_a = item_1,
                        check_b = item_2,
                        check_c = item_2,
                        check_d = item_3,
                        check_e = item_4,
                        check_f = item_5,
                        check_g = item_6
                    FROM extinguishers e
                    WHERE mi.extinguisher_id = e.id
                      AND e.asset_category = :ysc_category
                    """
                ),
                {"ysc_category": DEFAULT_ASSET_CATEGORY},
            )
            connection.execute(
                text(
                    "UPDATE extinguishers SET asset_category = :default_category "
                    "WHERE asset_category IS NULL OR BTRIM(asset_category) = ''"
                ),
                {"default_category": DEFAULT_ASSET_CATEGORY},
            )
            connection.execute(
                text(
                    """
                    DO $$
                    DECLARE
                        drop_sql TEXT;
                    BEGIN
                        FOR drop_sql IN
                            SELECT 'ALTER TABLE extinguishers DROP CONSTRAINT ' || quote_ident(c.conname)
                            FROM pg_constraint c
                            JOIN pg_class t ON t.oid = c.conrelid
                            JOIN pg_namespace n ON n.oid = t.relnamespace
                            WHERE t.relname = 'extinguishers'
                              AND n.nspname = current_schema()
                              AND c.contype = 'u'
                              AND (
                                SELECT string_agg(att.attname, ',' ORDER BY att.attname)
                                FROM unnest(c.conkey) AS colnum
                                JOIN pg_attribute att
                                  ON att.attrelid = c.conrelid
                                 AND att.attnum = colnum
                              ) = 'serial_number'
                        LOOP
                            EXECUTE drop_sql;
                        END LOOP;

                        FOR drop_sql IN
                            SELECT 'DROP INDEX IF EXISTS ' || quote_ident(indexname)
                            FROM pg_indexes
                            WHERE schemaname = current_schema()
                              AND tablename = 'extinguishers'
                              AND indexdef ILIKE '%UNIQUE%'
                              AND indexdef ILIKE '%(serial_number)%'
                              AND indexdef NOT ILIKE '%(asset_category, serial_number)%'
                        LOOP
                            EXECUTE drop_sql;
                        END LOOP;
                    END $$;
                    """
                )
            )
            connection.execute(
                text(
                    """
                    CREATE UNIQUE INDEX IF NOT EXISTS uq_extinguishers_asset_category_serial_number
                    ON extinguishers (asset_category, serial_number)
                    """
                )
            )


def seed_companies_from_extinguishers() -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with engine.begin() as connection:
        company_rows = {
            row.name: {"id": row.id, "public_id": row.public_id, "slug": row.slug}
            for row in connection.execute(select(companies.c.id, companies.c.name, companies.c.public_id, companies.c.slug)).all()
        }
        extinguisher_rows = connection.execute(
            select(
                extinguishers.c.id,
                extinguishers.c.company_id,
                extinguishers.c.company_name,
                extinguishers.c.company_address,
                extinguishers.c.company_contact,
            )
        ).mappings().all()
        for row in extinguisher_rows:
            company_name = (row["company_name"] or "").strip()
            if not company_name:
                continue
            company_info = company_rows.get(company_name)
            if company_info is None:
                result = connection.execute(
                    insert(companies).values(
                        public_id=uuid.uuid4().hex[:12],
                        slug=build_unique_company_slug(connection, company_name),
                        name=company_name,
                        address=(row.get("company_address") or "-").strip() or "-",
                        contact_name=(row.get("company_contact") or "-").strip() or "-",
                        created_at=now,
                        updated_at=now,
                    )
                )
                company_id = result.inserted_primary_key[0]
                company_rows[company_name] = {"id": company_id, "public_id": None, "slug": None}
            else:
                company_id = company_info["id"]
            if row.get("company_id") != company_id:
                connection.execute(
                    update(extinguishers)
                    .where(extinguishers.c.id == row["id"])
                    .values(company_id=company_id)
                )
        existing_companies = connection.execute(
            select(companies.c.id, companies.c.public_id, companies.c.slug, companies.c.name)
        ).mappings().all()
        for company in existing_companies:
            values = {}
            if not company.get("public_id"):
                values["public_id"] = uuid.uuid4().hex[:12]
            if not company.get("slug"):
                values["slug"] = build_unique_company_slug(connection, company["name"], exclude_id=company["id"])
            if values:
                values["updated_at"] = now
                connection.execute(
                    update(companies)
                    .where(companies.c.id == company["id"])
                    .values(**values)
                )


def seed_asset_categories() -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with engine.begin() as connection:
        rows = connection.execute(
            select(extinguishers.c.id, extinguishers.c.asset_category)
        ).mappings().all()
        for row in rows:
            if row.get("asset_category"):
                continue
            connection.execute(
                update(extinguishers)
                .where(extinguishers.c.id == row["id"])
                .values(asset_category=DEFAULT_ASSET_CATEGORY, updated_at=now)
            )


def is_authenticated() -> bool:
    return session.get("authenticated") is True


def current_user_full_name() -> str:
    return session.get("full_name", "")


def is_admin_user() -> bool:
    return session.get("is_admin") is True


def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if not is_authenticated():
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped_view


def admin_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if not is_authenticated():
            return redirect(url_for("login", next=request.path))
        if not is_admin_user():
            abort(403)
        return view_func(*args, **kwargs)

    return wrapped_view


@app.before_request
def protect_private_routes():
    allowed_endpoints = {
        "login",
        "logout",
        "public_detail",
        "public_control_form_pdf",
        "public_category_report_pdf",
        "public_electrical_report_pdf",
        "public_electrical_report_preview",
        "public_company_portal",
        "public_company_assets",
        "health",
        "static",
    }
    if request.endpoint in allowed_endpoints or request.endpoint is None:
        return None
    if not is_authenticated():
        return redirect(url_for("login", next=request.path))
    return None


def fetch_all(statement):
    with engine.connect() as connection:
        result = connection.execute(statement)
        return [dict(row._mapping) for row in result]


def fetch_one(statement):
    with engine.connect() as connection:
        row = connection.execute(statement).mappings().first()
        return dict(row) if row else None


def parse_float(value: str, field_name: str) -> float:
    try:
        return float(value.replace(",", "."))
    except ValueError as exc:
        raise ValueError(f"{field_name} sayi olmali.") from exc


def parse_required_form(form: dict[str, str]) -> dict[str, str]:
    return {key: value.strip() for key, value in form.items()}


def parse_structured_notes(notes: str | None) -> dict[str, str]:
    parsed: dict[str, str] = {}
    if not notes:
        return parsed
    for raw_line in notes.splitlines():
        line = raw_line.strip()
        if not line or ":" not in line:
            continue
        key, value = line.split(":", 1)
        parsed[key.strip()] = value.strip()
    return parsed


def build_electrical_note_sections(notes: str | None) -> list[dict]:
    parsed = parse_structured_notes(notes)
    sections: list[dict] = []
    for title, keys in ELECTRICAL_NOTE_SECTIONS:
        items = []
        for key in keys:
            value = parsed.get(key)
            if value:
                items.append({"label": key, "value": value})
        if items:
            sections.append({"title": title, "items": items})
    return sections


def parse_electrical_operation_summary(summary: str | None) -> dict[str, str]:
    if not summary:
        return {}
    parsed: dict[str, str] = {}
    start_match = re.search(r"Baslangic:\s*(.+?)\s*/\s*Bitis:", summary)
    end_match = re.search(r"Bitis:\s*(.+?)\s*/\s*Kullanim amaci:", summary)
    usage_match = re.search(r"Kullanim amaci:\s*(.+)$", summary)
    if start_match:
        parsed["control_start"] = start_match.group(1).strip()
    if end_match:
        parsed["control_end"] = end_match.group(1).strip()
    if usage_match:
        parsed["equipment_usage_purpose"] = usage_match.group(1).strip()
    return parsed


def build_electrical_final_conclusion(status: str | None) -> str:
    normalized = (status or "").strip().lower()
    if normalized == "uygundur":
        suffix = "uygundur."
    elif normalized == "uygun degildir":
        suffix = "uygun degildir."
    else:
        return "-"
    return f"Periyodik kontrol tarihi itibari ile mevcut sartlar altinda kullanimi {suffix}"


def build_electrical_structured_notes(form: dict[str, str]) -> dict[str, str]:
    final_conclusion_status = form.get("final_conclusion_status", "")
    return {
        "isg_katip_id": form.get("isg_katip_id", ""),
        "sgk_number": form.get("sgk_number", ""),
        "control_method": form.get("control_method", ""),
        "project_exists": form.get("project_exists", ""),
        "single_line_schema": form.get("single_line_schema", ""),
        "control_reason": form.get("control_reason", ""),
        "grounder_type": form.get("grounder_type", ""),
        "structure_type": form.get("structure_type", ""),
        "last_control_date": form.get("last_control_date", ""),
        "phase_conductor_type": form.get("phase_conductor_type", ""),
        "ground_resistance": form.get("ground_resistance", ""),
        "additional_ground_details": form.get("additional_ground_details", ""),
        "system_ground_conductor": form.get("system_ground_conductor", ""),
        "equipotential_conductor": form.get("equipotential_conductor", ""),
        "supply_characteristics": form.get("supply_characteristics", ""),
        "main_rcd_nominal": form.get("main_rcd_nominal", ""),
        "main_switch_characteristics": form.get("main_switch_characteristics", ""),
        "main_rcd_test": form.get("main_rcd_test", ""),
        "major_installation_change": form.get("major_installation_change", ""),
        "spd_used": form.get("spd_used", ""),
        "direct_contact_protections": form.get("direct_contact_protections", ""),
        "previous_control_label": form.get("previous_control_label", ""),
        "findings": form.get("findings", ""),
        "thermal_camera_1": form.get("thermal_camera_1", ""),
        "thermal_camera_2": form.get("thermal_camera_2", ""),
        "thermal_calibration_1": form.get("thermal_calibration_1", ""),
        "thermal_calibration_2": form.get("thermal_calibration_2", ""),
        "thermal_validity_1": form.get("thermal_validity_1", ""),
        "thermal_validity_2": form.get("thermal_validity_2", ""),
        "thermal_serial_1": form.get("thermal_serial_1", ""),
        "thermal_serial_2": form.get("thermal_serial_2", ""),
        "thermal_calibration_no_1": form.get("thermal_calibration_no_1", ""),
        "thermal_calibration_no_2": form.get("thermal_calibration_no_2", ""),
        "measurement_device_1": form.get("measurement_device_1", ""),
        "measurement_device_2": form.get("measurement_device_2", ""),
        "measurement_calibration_1": form.get("measurement_calibration_1", ""),
        "measurement_calibration_2": form.get("measurement_calibration_2", ""),
        "measurement_validity_1": form.get("measurement_validity_1", ""),
        "measurement_validity_2": form.get("measurement_validity_2", ""),
        "measurement_serial_1": form.get("measurement_serial_1", ""),
        "measurement_serial_2": form.get("measurement_serial_2", ""),
        "measurement_calibration_no_1": form.get("measurement_calibration_no_1", ""),
        "measurement_calibration_no_2": form.get("measurement_calibration_no_2", ""),
        "control_criteria_notes": form.get("control_criteria_notes", ""),
        "measurement_method": form.get("measurement_method", ""),
        "thermal_photo_date": form.get("thermal_photo_date", ""),
        "thermal_photo_number": form.get("thermal_photo_number", ""),
        "thermal_loose_contact_heating": form.get("thermal_loose_contact_heating", ""),
        "thermal_overload_heating": form.get("thermal_overload_heating", ""),
        "section_61_notes": form.get("section_61_notes", ""),
        "section_62_notes": form.get("section_62_notes", ""),
        "section_63_notes": form.get("section_63_notes", ""),
        "fault_notes": form.get("fault_notes", ""),
        "equipment_photos_notes": form.get("equipment_photos_notes", ""),
        "general_notes": form.get("general_notes", ""),
        "final_conclusion_status": final_conclusion_status,
        "final_conclusion": build_electrical_final_conclusion(final_conclusion_status),
        "authorized_person_name": form.get("authorized_person_name", ""),
        "authorized_person_job": form.get("authorized_person_job", ""),
        "authorized_person_registry": form.get("authorized_person_registry", ""),
        "copy_count": form.get("copy_count", ""),
    }


def build_electrical_note_lines(structured_notes: dict[str, str]) -> list[str]:
    return [
        f"ISG-KATIP Sozlesme ID: {structured_notes['isg_katip_id'] or '-'}",
        f"SGK Sicil Numarasi: {structured_notes['sgk_number'] or '-'}",
        f"Periyodik Kontrol Metodu ve Kapsami: {structured_notes['control_method'] or '-'}",
        f"Tesise ait proje var mi: {structured_notes['project_exists'] or '-'}",
        f"Tek hat semasi var mi: {structured_notes['single_line_schema'] or '-'}",
        f"Kontrol nedeni: {structured_notes['control_reason'] or '-'}",
        f"Topraklayici tipi: {structured_notes['grounder_type'] or '-'}",
        f"Yapi cinsi: {structured_notes['structure_type'] or '-'}",
        f"Son kontrol tarihi: {structured_notes['last_control_date'] or '-'}",
        f"Faz iletkenlerinin sayisi ve tipi: {structured_notes['phase_conductor_type'] or '-'}",
        f"Temel topraklama direnci: {structured_notes['ground_resistance'] or '-'}",
        f"Ilave topraklama elektrotu detaylari: {structured_notes['additional_ground_details'] or '-'}",
        f"Sistem topraklama iletkeni ve kesiti: {structured_notes['system_ground_conductor'] or '-'}",
        f"Ana espotansiyel iletkeni ve kesiti: {structured_notes['equipotential_conductor'] or '-'}",
        f"Besleme kaynagi karakteristikleri: {structured_notes['supply_characteristics'] or '-'}",
        f"Ana RCD anma akimi: {structured_notes['main_rcd_nominal'] or '-'}",
        f"Ana kesici karakteristikleri: {structured_notes['main_switch_characteristics'] or '-'}",
        f"Ana RCD test akimi ve suresi: {structured_notes['main_rcd_test'] or '-'}",
        f"Tesisatta kapsamli degisiklik var mi (>%20): {structured_notes['major_installation_change'] or '-'}",
        f"Asiri gerilim koruma cihazlari kullanilmis mi: {structured_notes['spd_used'] or '-'}",
        f"Dogrudan dokunmaya karsi koruma onlemleri: {structured_notes['direct_contact_protections'] or '-'}",
        f"Bir onceki periyodik kontrol etiketi var mi: {structured_notes['previous_control_label'] or '-'}",
        f"Tespit edilen bilgiler: {structured_notes['findings'] or '-'}",
        f"Termal Kamera 1: {structured_notes['thermal_camera_1'] or '-'} / Seri: {structured_notes['thermal_serial_1'] or '-'} / Kal.No: {structured_notes['thermal_calibration_no_1'] or '-'}",
        f"Termal Kamera 2: {structured_notes['thermal_camera_2'] or '-'} / Seri: {structured_notes['thermal_serial_2'] or '-'} / Kal.No: {structured_notes['thermal_calibration_no_2'] or '-'}",
        f"Olcum Aleti 1: {structured_notes['measurement_device_1'] or '-'} / Seri: {structured_notes['measurement_serial_1'] or '-'} / Kal.No: {structured_notes['measurement_calibration_no_1'] or '-'}",
        f"Olcum Aleti 2: {structured_notes['measurement_device_2'] or '-'} / Seri: {structured_notes['measurement_serial_2'] or '-'} / Kal.No: {structured_notes['measurement_calibration_no_2'] or '-'}",
        f"Kontrol Kriterleri ve Testler: {structured_notes['control_criteria_notes'] or '-'}",
        f"Olcum ve Dogrulama Metodu: {structured_notes['measurement_method'] or '-'}",
        f"Termal fotograf tarihi: {structured_notes['thermal_photo_date'] or '-'}",
        f"Termal fotograf no: {structured_notes['thermal_photo_number'] or '-'}",
        f"Kontak gevsakligi isinmasi: {structured_notes['thermal_loose_contact_heating'] or '-'}",
        f"Asiri yuk isinmasi: {structured_notes['thermal_overload_heating'] or '-'}",
        f"6.1 Notlari: {structured_notes['section_61_notes'] or '-'}",
        f"6.2 Notlari: {structured_notes['section_62_notes'] or '-'}",
        f"6.3 Notlari: {structured_notes['section_63_notes'] or '-'}",
        f"Kusur Aciklamalari: {structured_notes['fault_notes'] or '-'}",
        f"Ekipman Fotograflari: {structured_notes['equipment_photos_notes'] or '-'}",
        f"Genel Notlar: {structured_notes['general_notes'] or '-'}",
        f"Sonuc Durumu: {structured_notes['final_conclusion_status'] or '-'}",
        f"Sonuc ve Kanaat: {structured_notes['final_conclusion'] or '-'}",
        f"Yetkili Kisi: {structured_notes['authorized_person_name'] or '-'} / Meslek: {structured_notes['authorized_person_job'] or '-'} / Kayit No: {structured_notes['authorized_person_registry'] or '-'}",
        f"Nusha Sayisi: {structured_notes['copy_count'] or '-'}",
    ]


def build_electrical_form_state(extinguisher: dict, latest_log: dict | None = None) -> dict[str, str]:
    notes = parse_structured_notes(extinguisher.get("notes"))
    summary = parse_electrical_operation_summary(latest_log.get("operation_summary") if latest_log else "")
    return {
        "company_id": str(extinguisher.get("company_id") or ""),
        "report_number": extinguisher.get("serial_number") or "",
        "company_address": extinguisher.get("company_address") or "",
        "report_date": extinguisher.get("last_service_date") or "",
        "isg_katip_id": notes.get("ISG-KATIP Sozlesme ID", "") or notes.get("isg_katip_id", ""),
        "control_start": summary.get("control_start", ""),
        "control_end": summary.get("control_end", ""),
        "sgk_number": notes.get("SGK Sicil Numarasi", "") or notes.get("sgk_number", ""),
        "next_service_date": extinguisher.get("next_service_date") or "",
        "control_method": notes.get("Periyodik Kontrol Metodu ve Kapsami", "") or notes.get("control_method", ""),
        "energy_provider": extinguisher.get("manufacturer") or "",
        "grid_type": extinguisher.get("fire_class") or "",
        "grid_voltage": notes.get("Sebeke gerilimi", "") or notes.get("grid_voltage", ""),
        "project_exists": notes.get("Tesise ait proje var mi", "") or notes.get("project_exists", ""),
        "single_line_schema": notes.get("Tek hat semasi var mi", "") or notes.get("single_line_schema", ""),
        "control_reason": notes.get("Kontrol nedeni", "") or notes.get("control_reason", ""),
        "grounder_type": notes.get("Topraklayici tipi", "") or notes.get("grounder_type", ""),
        "structure_type": notes.get("Yapi cinsi", "") or notes.get("structure_type", ""),
        "equipment_usage_purpose": summary.get("equipment_usage_purpose", ""),
        "last_control_date": notes.get("Son kontrol tarihi", "") or notes.get("last_control_date", ""),
        "phase_conductor_type": notes.get("Faz iletkenlerinin sayisi ve tipi", "") or notes.get("phase_conductor_type", ""),
        "ground_resistance": notes.get("Temel topraklama direnci", "") or notes.get("ground_resistance", ""),
        "additional_ground_details": notes.get("Ilave topraklama elektrotu detaylari", "") or notes.get("additional_ground_details", ""),
        "system_ground_conductor": notes.get("Sistem topraklama iletkeni ve kesiti", "") or notes.get("system_ground_conductor", ""),
        "equipotential_conductor": notes.get("Ana espotansiyel iletkeni ve kesiti", "") or notes.get("equipotential_conductor", ""),
        "supply_characteristics": notes.get("Besleme kaynagi karakteristikleri", "") or notes.get("supply_characteristics", ""),
        "main_rcd_nominal": notes.get("Ana RCD anma akimi", "") or notes.get("main_rcd_nominal", ""),
        "main_switch_characteristics": notes.get("Ana kesici karakteristikleri", "") or notes.get("main_switch_characteristics", ""),
        "main_rcd_test": notes.get("Ana RCD test akimi ve suresi", "") or notes.get("main_rcd_test", ""),
        "major_installation_change": notes.get("Tesisatta kapsamli degisiklik var mi (>%20)", "") or notes.get("major_installation_change", ""),
        "spd_used": notes.get("Asiri gerilim koruma cihazlari kullanilmis mi", "") or notes.get("spd_used", ""),
        "direct_contact_protections": notes.get("Dogrudan dokunmaya karsi koruma onlemleri", "") or notes.get("direct_contact_protections", ""),
        "previous_control_label": notes.get("Bir onceki periyodik kontrol etiketi var mi", "") or notes.get("previous_control_label", ""),
        "findings": notes.get("Tespit edilen bilgiler", "") or notes.get("findings", ""),
        "thermal_camera_1": notes.get("thermal_camera_1", ""),
        "thermal_camera_2": notes.get("thermal_camera_2", ""),
        "thermal_calibration_1": notes.get("thermal_calibration_1", ""),
        "thermal_calibration_2": notes.get("thermal_calibration_2", ""),
        "thermal_validity_1": notes.get("thermal_validity_1", ""),
        "thermal_validity_2": notes.get("thermal_validity_2", ""),
        "thermal_serial_1": notes.get("thermal_serial_1", ""),
        "thermal_serial_2": notes.get("thermal_serial_2", ""),
        "thermal_calibration_no_1": notes.get("thermal_calibration_no_1", ""),
        "thermal_calibration_no_2": notes.get("thermal_calibration_no_2", ""),
        "measurement_device_1": notes.get("measurement_device_1", ""),
        "measurement_device_2": notes.get("measurement_device_2", ""),
        "measurement_calibration_1": notes.get("measurement_calibration_1", ""),
        "measurement_calibration_2": notes.get("measurement_calibration_2", ""),
        "measurement_validity_1": notes.get("measurement_validity_1", ""),
        "measurement_validity_2": notes.get("measurement_validity_2", ""),
        "measurement_serial_1": notes.get("measurement_serial_1", ""),
        "measurement_serial_2": notes.get("measurement_serial_2", ""),
        "measurement_calibration_no_1": notes.get("measurement_calibration_no_1", ""),
        "measurement_calibration_no_2": notes.get("measurement_calibration_no_2", ""),
        "control_criteria_notes": notes.get("Kontrol Kriterleri ve Testler", "") or notes.get("control_criteria_notes", ""),
        "measurement_method": notes.get("Olcum ve Dogrulama Metodu", "") or notes.get("measurement_method", ""),
        "thermal_photo_date": notes.get("Termal fotograf tarihi", "") or notes.get("thermal_photo_date", ""),
        "thermal_photo_number": notes.get("Termal fotograf no", "") or notes.get("thermal_photo_number", ""),
        "thermal_loose_contact_heating": notes.get("Kontak gevsakligi isinmasi", "") or notes.get("thermal_loose_contact_heating", ""),
        "thermal_overload_heating": notes.get("Asiri yuk isinmasi", "") or notes.get("thermal_overload_heating", ""),
        "section_61_notes": notes.get("6.1 Notlari", "") or notes.get("section_61_notes", ""),
        "section_62_notes": notes.get("6.2 Notlari", "") or notes.get("section_62_notes", ""),
        "section_63_notes": notes.get("6.3 Notlari", "") or notes.get("section_63_notes", ""),
        "fault_notes": notes.get("Kusur Aciklamalari", "") or notes.get("fault_notes", ""),
        "equipment_photos_notes": notes.get("Ekipman Fotograflari", "") or notes.get("equipment_photos_notes", ""),
        "general_notes": notes.get("Genel Notlar", "") or notes.get("general_notes", ""),
        "final_conclusion_status": notes.get("Sonuc Durumu", "") or notes.get("final_conclusion_status", ""),
        "authorized_person_name": notes.get("authorized_person_name", ""),
        "authorized_person_job": notes.get("authorized_person_job", ""),
        "authorized_person_registry": notes.get("authorized_person_registry", ""),
        "copy_count": notes.get("Nusha Sayisi", "") or notes.get("copy_count", ""),
        "technician_name": current_user_full_name(),
    }


def get_extinguisher(public_id: str) -> dict:
    extinguisher = fetch_one(
        select(extinguishers).where(extinguishers.c.public_id == public_id)
    )
    if extinguisher is None:
        abort(404)
    return extinguisher


def get_company(company_id: int) -> dict:
    company = fetch_one(select(companies).where(companies.c.id == company_id))
    if company is None:
        abort(404)
    return company


def get_company_by_public_id(public_id: str) -> dict:
    company = fetch_one(select(companies).where(companies.c.public_id == public_id))
    if company is None:
        abort(404)
    return company


def get_company_by_slug(slug: str) -> dict:
    company = fetch_one(select(companies).where(companies.c.slug == slug))
    if company is None:
        abort(404)
    return company


def get_company_choices() -> list[dict]:
    return fetch_all(select(companies).order_by(companies.c.name))


def get_asset_category_choices() -> list[dict]:
    return ASSET_CATEGORIES


def get_asset_category(slug: str | None = None, *, label: str | None = None) -> dict | None:
    if slug:
        return ASSET_CATEGORY_BY_SLUG.get(slug)
    if label:
        return ASSET_CATEGORY_BY_LABEL.get(label)
    return None


def slugify_company_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", normalized.lower()).strip("-")
    return slug or "firma"


def build_unique_company_slug(connection, name: str, exclude_id: int | None = None) -> str:
    base_slug = slugify_company_name(name)
    slug = base_slug
    index = 2
    while True:
        statement = select(companies.c.id).where(companies.c.slug == slug)
        row = connection.execute(statement).first()
        if row is None or (exclude_id is not None and row.id == exclude_id):
            return slug
        slug = f"{base_slug}-{index}"
        index += 1


def resolve_company_slug(connection, raw_slug: str, fallback_name: str, *, exclude_id: int | None = None) -> str:
    source = (raw_slug or "").strip() or fallback_name
    return build_unique_company_slug(connection, source, exclude_id=exclude_id)


run_schema_migrations()
seed_default_users()
seed_companies_from_extinguishers()
seed_asset_categories()


def build_company_portal_sections(company_id: int) -> list[dict]:
    assets = fetch_all(
        select(
            extinguishers.c.public_id,
            extinguishers.c.asset_category,
            extinguishers.c.serial_number,
            extinguishers.c.location_detail,
            extinguishers.c.extinguisher_type,
            extinguishers.c.last_service_date,
            extinguishers.c.next_service_date,
            extinguishers.c.pressure_status,
        )
        .where(extinguishers.c.company_id == company_id)
        .order_by(extinguishers.c.asset_category, extinguishers.c.location_detail, extinguishers.c.serial_number)
    )
    grouped: dict[str, list[dict]] = {}
    for asset in assets:
        label = asset.get("asset_category") or DEFAULT_ASSET_CATEGORY
        grouped.setdefault(label, []).append(asset)

    sections = []
    for category in ASSET_CATEGORIES:
        items = grouped.get(category["label"], [])
        sections.append(
            {
                **category,
                "count": len(items),
                "items": items,
            }
        )
    return sections


def get_registration_groups() -> list[dict]:
    return REGISTRATION_GROUPS


def get_registration_group(group_slug: str) -> dict:
    group = REGISTRATION_GROUP_BY_SLUG.get(group_slug)
    if group is None:
        abort(404)
    return group


def get_registration_group_by_label(group_label: str | None) -> dict | None:
    if not group_label:
        return None
    for group in REGISTRATION_GROUPS:
        if group["label"] == group_label:
            return group
    return None


def get_asset_profile(asset_category: str | None) -> dict:
    return ASSET_PROFILES.get(asset_category or "", ASSET_PROFILES["Yangin Sondurme Cihazi"])


def generate_prefixed_report_number(prefix: str) -> str:
    year = datetime.now().year
    like_pattern = f"{prefix}-{year}-%"
    existing = fetch_all(
        select(extinguishers.c.serial_number)
        .where(extinguishers.c.serial_number.like(like_pattern))
        .order_by(desc(extinguishers.c.serial_number))
    )
    max_number = 0
    for row in existing:
        value = row.get("serial_number") or ""
        match = re.match(rf"^{re.escape(prefix)}-{year}-(\d+)$", value)
        if match:
            max_number = max(max_number, int(match.group(1)))
    return f"{prefix}-{year}-{max_number + 1:04d}"


def render_profile_record_form(group_slug: str):
    group = get_registration_group(group_slug)
    if group_slug == "yangin-sondurme-cihazi":
        return redirect(url_for("create_extinguisher"))

    company_choices = get_company_choices()
    asset_profile = get_asset_profile(group["label"])

    if request.method == "POST":
        submit_action = request.form.get("submit_action", "save_print")
        form = parse_required_form(request.form)
        form["technician_name"] = form.get("technician_name") or current_user_full_name()
        form["asset_category"] = group["label"]
        form["extinguisher_type"] = asset_profile.get("fixed_type") or group["label"]
        try:
            form, selected_company = sync_company_payload_from_selection(form)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
            )

        required_fields = {
            "serial_number": "Seri numarasi",
            "company_id": "Cari secimi",
            "company_contact": asset_profile["owner_label"],
            "location_detail": "Bulundugu yer",
            "fire_class": asset_profile["class_label"],
            "manufacturer": asset_profile["brand_label"],
            "last_service_date": asset_profile["service_input_label"],
            "next_service_date": asset_profile["next_service_input_label"],
            "technician_name": "Teknisyen",
            "operation_summary": "Yapilan islem",
        }
        if asset_profile["show_weight"]:
            required_fields["weight_kg"] = "Kg"
        if asset_profile["show_hydrostatic"]:
            required_fields["hydrostatic_test_date"] = "Hidrostatik test tarihi"

        missing = [label for key, label in required_fields.items() if not form.get(key)]
        if missing:
            flash(f"Eksik alanlar: {', '.join(missing)}", "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
            )

        try:
            weight_kg = parse_float(form["weight_kg"], "Kg") if asset_profile["show_weight"] else 0.0
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
            )

        now = datetime.now().isoformat(timespec="seconds")
        public_id = uuid.uuid4().hex[:12]
        inspection_values = build_monthly_inspection_values(request.form, asset_profile["monthly_control_items"])
        control_values = build_control_form_values({})
        existing_same_category = fetch_one(
            select(extinguishers.c.id)
            .where(extinguishers.c.asset_category == group["label"])
            .where(extinguishers.c.serial_number == form["serial_number"])
        )
        if existing_same_category:
            flash("Bu seri numarasi bu urun grubunda zaten kayitli.", "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
            )

        try:
            with engine.begin() as connection:
                result = connection.execute(
                    insert(extinguishers).values(
                        public_id=public_id,
                        serial_number=form["serial_number"],
                        company_id=selected_company["id"],
                        company_name=form["company_name"],
                        company_address=form["company_address"],
                        company_contact=form["company_contact"],
                        asset_category=group["label"],
                        location_detail=form["location_detail"],
                        weight_kg=weight_kg,
                        extinguisher_type=form["extinguisher_type"],
                        fire_class=form["fire_class"],
                        manufacturer=form["manufacturer"],
                        hydrostatic_test_date=form.get("hydrostatic_test_date") if asset_profile["show_hydrostatic"] else None,
                        pressure_status=None,
                        notes=form.get("notes"),
                        last_service_date=form["last_service_date"],
                        next_service_date=form["next_service_date"],
                        created_at=now,
                        updated_at=now,
                    )
                )
                extinguisher_id = result.inserted_primary_key[0]
                connection.execute(
                    insert(service_logs).values(
                        extinguisher_id=extinguisher_id,
                        service_date=form["last_service_date"],
                        technician_name=form["technician_name"],
                        operation_summary=form["operation_summary"],
                        pressure_status=None,
                        notes=form.get("notes"),
                        created_at=now,
                    )
                )
                save_monthly_inspection(
                    connection=connection,
                    extinguisher_id=extinguisher_id,
                    inspection_date=form["last_service_date"],
                    inspector_name=form["technician_name"],
                    notes=form.get("notes"),
                    inspection_values=inspection_values,
                    control_values=control_values,
                    created_at=now,
                )
        except IntegrityError as exc:
            error_text = str(getattr(exc, "orig", exc))
            if "serial_number" in error_text.lower():
                flash("Bu seri numarasi bu urun grubunda zaten kayitli.", "error")
            else:
                flash(f"Kayit sirasinda veritabani hatasi olustu: {error_text[:180]}", "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
            )

        if submit_action == "save":
            flash(f"{group['label']} kaydedildi.", "success")
            return redirect(url_for("extinguisher_detail", public_id=public_id))

        flash(f"{group['label']} kaydedildi. Etiket yazdirmaya yonlendirildiniz.", "success")
        return redirect(url_for("extinguisher_label", public_id=public_id))

    return render_template(
        "create_asset_profile.html",
        form={
            "technician_name": current_user_full_name(),
            "asset_category": group["label"],
            "extinguisher_type": asset_profile.get("fixed_type") or group["label"],
        },
        companies=company_choices,
        asset_profile=asset_profile,
        group=group,
    )


@app.route("/records/new/elektrik-ic-tesisati", methods=["GET", "POST"])
@login_required
def create_electrical_installation():
    group = get_registration_group("elektrik-ic-tesisati")
    asset_profile = get_asset_profile(group["label"])
    company_choices = get_company_choices()

    if request.method == "POST":
        form = parse_required_form(request.form)
        form["technician_name"] = current_user_full_name()
        form["report_number"] = generate_prefixed_report_number("ELK")
        try:
            form, selected_company = sync_company_payload_from_selection(form)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_electrical_installation.html",
                form=form,
                companies=company_choices,
                group=group,
                asset_profile=asset_profile,
            )

        required_fields = {
            "company_id": "Cari secimi",
            "company_address": "Periyodik kontrol adresi",
            "report_date": "Rapor tarihi",
            "control_start": "Periyodik kontrol baslangic tarihi ve saati",
            "control_end": "Periyodik kontrol bitis tarihi ve saati",
            "next_service_date": "Bir sonraki periyodik kontrol tarihi",
            "energy_provider": "Enerji saglayan kurulus",
            "grid_type": "Sebeke tipi",
            "grid_voltage": "Sebeke gerilimi",
            "equipment_usage_purpose": "Ekipmanin kullanim amaci",
            "final_conclusion_status": "Sonuc ve kanaat",
            "technician_name": "Teknisyen",
        }
        missing = [label for key, label in required_fields.items() if not form.get(key)]
        if missing:
            flash(f"Eksik alanlar: {', '.join(missing)}", "error")
            return render_template(
                "create_electrical_installation.html",
                form=form,
                companies=company_choices,
                group=group,
                asset_profile=asset_profile,
            )

        final_conclusion_text = build_electrical_final_conclusion(form.get("final_conclusion_status"))

        structured_notes = {
            "isg_katip_id": form.get("isg_katip_id", ""),
            "sgk_number": form.get("sgk_number", ""),
            "control_method": form.get("control_method", ""),
            "project_exists": form.get("project_exists", ""),
            "single_line_schema": form.get("single_line_schema", ""),
            "control_reason": form.get("control_reason", ""),
            "grounder_type": form.get("grounder_type", ""),
            "structure_type": form.get("structure_type", ""),
            "last_control_date": form.get("last_control_date", ""),
            "phase_conductor_type": form.get("phase_conductor_type", ""),
            "ground_resistance": form.get("ground_resistance", ""),
            "additional_ground_details": form.get("additional_ground_details", ""),
            "system_ground_conductor": form.get("system_ground_conductor", ""),
            "equipotential_conductor": form.get("equipotential_conductor", ""),
            "supply_characteristics": form.get("supply_characteristics", ""),
            "main_rcd_nominal": form.get("main_rcd_nominal", ""),
            "main_switch_characteristics": form.get("main_switch_characteristics", ""),
            "main_rcd_test": form.get("main_rcd_test", ""),
            "major_installation_change": form.get("major_installation_change", ""),
            "spd_used": form.get("spd_used", ""),
            "direct_contact_protections": form.get("direct_contact_protections", ""),
            "previous_control_label": form.get("previous_control_label", ""),
            "findings": form.get("findings", ""),
            "thermal_camera_1": form.get("thermal_camera_1", ""),
            "thermal_camera_2": form.get("thermal_camera_2", ""),
            "thermal_calibration_1": form.get("thermal_calibration_1", ""),
            "thermal_calibration_2": form.get("thermal_calibration_2", ""),
            "thermal_validity_1": form.get("thermal_validity_1", ""),
            "thermal_validity_2": form.get("thermal_validity_2", ""),
            "thermal_serial_1": form.get("thermal_serial_1", ""),
            "thermal_serial_2": form.get("thermal_serial_2", ""),
            "thermal_calibration_no_1": form.get("thermal_calibration_no_1", ""),
            "thermal_calibration_no_2": form.get("thermal_calibration_no_2", ""),
            "measurement_device_1": form.get("measurement_device_1", ""),
            "measurement_device_2": form.get("measurement_device_2", ""),
            "measurement_calibration_1": form.get("measurement_calibration_1", ""),
            "measurement_calibration_2": form.get("measurement_calibration_2", ""),
            "measurement_validity_1": form.get("measurement_validity_1", ""),
            "measurement_validity_2": form.get("measurement_validity_2", ""),
            "measurement_serial_1": form.get("measurement_serial_1", ""),
            "measurement_serial_2": form.get("measurement_serial_2", ""),
            "measurement_calibration_no_1": form.get("measurement_calibration_no_1", ""),
            "measurement_calibration_no_2": form.get("measurement_calibration_no_2", ""),
            "control_criteria_notes": form.get("control_criteria_notes", ""),
            "measurement_method": form.get("measurement_method", ""),
            "thermal_photo_date": form.get("thermal_photo_date", ""),
            "thermal_photo_number": form.get("thermal_photo_number", ""),
            "thermal_loose_contact_heating": form.get("thermal_loose_contact_heating", ""),
            "thermal_overload_heating": form.get("thermal_overload_heating", ""),
            "section_61_notes": form.get("section_61_notes", ""),
            "section_62_notes": form.get("section_62_notes", ""),
            "section_63_notes": form.get("section_63_notes", ""),
            "fault_notes": form.get("fault_notes", ""),
            "equipment_photos_notes": form.get("equipment_photos_notes", ""),
            "general_notes": form.get("general_notes", ""),
            "final_conclusion_status": form.get("final_conclusion_status", ""),
            "final_conclusion": final_conclusion_text,
            "authorized_person_name": form.get("authorized_person_name", ""),
            "authorized_person_job": form.get("authorized_person_job", ""),
            "authorized_person_registry": form.get("authorized_person_registry", ""),
            "copy_count": form.get("copy_count", ""),
        }
        note_lines = [
            "Elektrik Ic Tesisati Tam Form Baslangic Kaydi",
            f"ISG-KATIP Sozlesme ID: {structured_notes['isg_katip_id'] or '-'}",
            f"SGK Sicil Numarasi: {structured_notes['sgk_number'] or '-'}",
            f"Periyodik Kontrol Metodu ve Kapsami: {structured_notes['control_method'] or '-'}",
            f"Tesise ait proje var mi: {structured_notes['project_exists'] or '-'}",
            f"Tek hat semasi var mi: {structured_notes['single_line_schema'] or '-'}",
            f"Kontrol nedeni: {structured_notes['control_reason'] or '-'}",
            f"Topraklayici tipi: {structured_notes['grounder_type'] or '-'}",
            f"Yapi cinsi: {structured_notes['structure_type'] or '-'}",
            f"Son kontrol tarihi: {structured_notes['last_control_date'] or '-'}",
            f"Faz iletkenlerinin sayisi ve tipi: {structured_notes['phase_conductor_type'] or '-'}",
            f"Temel topraklama direnci: {structured_notes['ground_resistance'] or '-'}",
            f"Ilave topraklama elektrotu detaylari: {structured_notes['additional_ground_details'] or '-'}",
            f"Sistem topraklama iletkeni ve kesiti: {structured_notes['system_ground_conductor'] or '-'}",
            f"Ana espotansiyel iletkeni ve kesiti: {structured_notes['equipotential_conductor'] or '-'}",
            f"Besleme kaynagi karakteristikleri: {structured_notes['supply_characteristics'] or '-'}",
            f"Ana RCD anma akimi: {structured_notes['main_rcd_nominal'] or '-'}",
            f"Ana kesici karakteristikleri: {structured_notes['main_switch_characteristics'] or '-'}",
            f"Ana RCD test akimi ve suresi: {structured_notes['main_rcd_test'] or '-'}",
            f"Tesisatta kapsamli degisiklik var mi (>%20): {structured_notes['major_installation_change'] or '-'}",
            f"Asiri gerilim koruma cihazlari kullanilmis mi: {structured_notes['spd_used'] or '-'}",
            f"Dogrudan dokunmaya karsi koruma onlemleri: {structured_notes['direct_contact_protections'] or '-'}",
            f"Bir onceki periyodik kontrol etiketi var mi: {structured_notes['previous_control_label'] or '-'}",
            f"Tespit edilen bilgiler: {structured_notes['findings'] or '-'}",
            f"Termal Kamera 1: {structured_notes['thermal_camera_1'] or '-'} / Seri: {structured_notes['thermal_serial_1'] or '-'} / Kal.No: {structured_notes['thermal_calibration_no_1'] or '-'}",
            f"Termal Kamera 2: {structured_notes['thermal_camera_2'] or '-'} / Seri: {structured_notes['thermal_serial_2'] or '-'} / Kal.No: {structured_notes['thermal_calibration_no_2'] or '-'}",
            f"Olcum Aleti 1: {structured_notes['measurement_device_1'] or '-'} / Seri: {structured_notes['measurement_serial_1'] or '-'} / Kal.No: {structured_notes['measurement_calibration_no_1'] or '-'}",
            f"Olcum Aleti 2: {structured_notes['measurement_device_2'] or '-'} / Seri: {structured_notes['measurement_serial_2'] or '-'} / Kal.No: {structured_notes['measurement_calibration_no_2'] or '-'}",
            f"Kontrol Kriterleri ve Testler: {structured_notes['control_criteria_notes'] or '-'}",
            f"Olcum ve Dogrulama Metodu: {structured_notes['measurement_method'] or '-'}",
            f"Termal fotograf tarihi: {structured_notes['thermal_photo_date'] or '-'}",
            f"Termal fotograf no: {structured_notes['thermal_photo_number'] or '-'}",
            f"Kontak gevsakligi isinmasi: {structured_notes['thermal_loose_contact_heating'] or '-'}",
            f"Asiri yuk isinmasi: {structured_notes['thermal_overload_heating'] or '-'}",
            f"6.1 Notlari: {structured_notes['section_61_notes'] or '-'}",
            f"6.2 Notlari: {structured_notes['section_62_notes'] or '-'}",
            f"6.3 Notlari: {structured_notes['section_63_notes'] or '-'}",
            f"Kusur Aciklamalari: {structured_notes['fault_notes'] or '-'}",
            f"Ekipman Fotograflari: {structured_notes['equipment_photos_notes'] or '-'}",
            f"Genel Notlar: {structured_notes['general_notes'] or '-'}",
            f"Sonuc Durumu: {structured_notes['final_conclusion_status'] or '-'}",
            f"Sonuc ve Kanaat: {structured_notes['final_conclusion'] or '-'}",
            f"Yetkili Kisi: {structured_notes['authorized_person_name'] or '-'} / Meslek: {structured_notes['authorized_person_job'] or '-'} / Kayit No: {structured_notes['authorized_person_registry'] or '-'}",
            f"Nusha Sayisi: {structured_notes['copy_count'] or '-'}",
        ]

        now = datetime.now().isoformat(timespec="seconds")
        public_id = uuid.uuid4().hex[:12]
        with engine.begin() as connection:
            result = connection.execute(
                insert(extinguishers).values(
                    public_id=public_id,
                    serial_number=form["report_number"],
                    company_id=selected_company["id"],
                    company_name=form["company_name"],
                    company_address=form["company_address"],
                    company_contact=form["report_number"],
                    asset_category=group["label"],
                    location_detail=form["company_address"],
                    weight_kg=0.0,
                    extinguisher_type="Elektrik Ic Tesisati",
                    fire_class=form["grid_type"],
                    manufacturer=form["energy_provider"],
                    hydrostatic_test_date=None,
                    pressure_status=None,
                    notes="\n".join(note_lines),
                    last_service_date=form["report_date"],
                    next_service_date=form["next_service_date"],
                    created_at=now,
                    updated_at=now,
                )
            )
            extinguisher_id = result.inserted_primary_key[0]
            connection.execute(
                insert(service_logs).values(
                    extinguisher_id=extinguisher_id,
                    service_date=form["report_date"],
                    technician_name=form["technician_name"],
                    operation_summary=(
                        f"Elektrik ic tesisati periyodik kontrol raporu olusturuldu. "
                        f"Baslangic: {form['control_start']} / Bitis: {form['control_end']} / "
                        f"Kullanim amaci: {form['equipment_usage_purpose']}"
                    ),
                    pressure_status=None,
                    notes=form.get("findings"),
                    created_at=now,
                )
            )

        flash("Elektrik ic tesisati kaydi olusturuldu.", "success")
        return redirect(url_for("extinguisher_detail", public_id=public_id))

    report_number = generate_prefixed_report_number("ELK")
    return render_template(
        "create_electrical_installation.html",
        form={
            "technician_name": current_user_full_name(),
            "report_number": report_number,
            "control_method": (
                "- TS HD 60364-4-43 Alcak Gerilim Elektrik Tesisatlari - Bolum 4: Guvenlik Icin Koruma Grup 43 - Asiri Akima Karsi Koruma\n"
                "- TS HD 60364-6 Alcak Gerilim Elektrik Tesisatlari - Bolum 6: Dogrulama\n"
                "- Is Ekipmanlarinin Kullaniminda Saglik ve Guvenlik Sartlari Yonetmeligi\n"
                "- Elektrik Ic Tesisleri Yonetmeligi\n"
                "- Elektrik Tesislerinde Topraklamalar Yonetmeligi"
            ),
            "final_conclusion_status": "uygundur",
            "copy_count": "2",
        },
        companies=company_choices,
        group=group,
        asset_profile=asset_profile,
    )


def sync_company_payload_from_selection(form: dict[str, str]) -> tuple[dict[str, str], dict]:
    company_id_raw = (form.get("company_id") or "").strip()
    if not company_id_raw:
        raise ValueError("Firma secilmeden kayit olusturulamaz.")
    try:
        company_id = int(company_id_raw)
    except ValueError as exc:
        raise ValueError("Firma secimi gecersiz.") from exc
    company = get_company(company_id)
    form["company_id"] = str(company["id"])
    form["company_name"] = company["name"]
    form["company_address"] = company["address"]
    return form, company


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf"),
        Path("C:/Windows/Fonts/segoeuib.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
    ]
    for font_path in font_candidates:
        if font_path.exists():
            try:
                return ImageFont.truetype(str(font_path), size)
            except OSError:
                continue
    return ImageFont.load_default()


def resolve_system_font(*candidates: str) -> str:
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
    raise FileNotFoundError(f"Uygun font bulunamadi: {', '.join(candidates)}")


def build_branded_qr(public_url: str, *, label_mode: bool = False) -> io.BytesIO:
    qr = qrcode.QRCode(
        version=None,
        error_correction=ERROR_CORRECT_H,
        box_size=10,
        border=2 if label_mode else 4,
    )
    qr.add_data(public_url)
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    qr_width, qr_height = qr_image.size

    if label_mode:
        draw_qr = ImageDraw.Draw(qr_image)
        center_font = load_font(max(104, qr_width // 4))
        v_text = "V"
        v_box = draw_qr.textbbox((0, 0), v_text, font=center_font)
        v_width = v_box[2] - v_box[0]
        v_height = v_box[3] - v_box[1]
        v_pad = 10
        v_bg_w = v_width + (v_pad * 2)
        v_bg_h = v_height + (v_pad * 2)
        v_x = (qr_width - v_bg_w) // 2
        v_y = (qr_height - v_bg_h) // 2
        draw_qr.rounded_rectangle(
            (v_x, v_y, v_x + v_bg_w, v_y + v_bg_h),
            radius=8,
            fill="white",
        )
        draw_qr.text(
            (v_x + v_pad, v_y + v_pad - 2),
            v_text,
            fill="black",
            font=center_font,
        )
    elif LOGO_PATH.exists():
        logo_source = LOGO_PATH
        logo = Image.open(logo_source).convert("RGBA")
        max_logo_size = int(qr_width * 0.24)
        logo.thumbnail((max_logo_size, max_logo_size))

        logo_bg_size = max(logo.width, logo.height) + 20
        logo_bg = Image.new("RGBA", (logo_bg_size, logo_bg_size), (255, 255, 255, 255))
        bg_x = (logo_bg_size - logo.width) // 2
        bg_y = (logo_bg_size - logo.height) // 2
        logo_bg.alpha_composite(logo, (bg_x, bg_y))

        logo_x = (qr_width - logo_bg_size) // 2
        logo_y = (qr_height - logo_bg_size) // 2
        qr_image.alpha_composite(logo_bg, (logo_x, logo_y))

    if label_mode:
        footer_height = 130
        canvas = Image.new("RGBA", (qr_width, qr_height + footer_height), "white")
        canvas.paste(qr_image, (0, 0))
        draw = ImageDraw.Draw(canvas)
        font = load_font(96)
        brand_text = "Vesta"
        text_bbox = draw.textbbox((0, 0), brand_text, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_x = (qr_width - text_width) // 2
        text_y = qr_height + 14
        draw.text((text_x, text_y), brand_text, fill="black", font=font)
    else:
        canvas = Image.new("RGBA", (qr_width, qr_height + 78), "white")
        canvas.paste(qr_image, (0, 0))

        draw = ImageDraw.Draw(canvas)
        font = load_font(28)
        text_bbox = draw.textbbox((0, 0), BRAND_NAME, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_x = (qr_width - text_width) // 2
        text_y = qr_height + 24
        draw.text((text_x, text_y), BRAND_NAME, fill="black", font=font)

    buffer = io.BytesIO()
    canvas.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


def build_monthly_inspection_values(form_data, items: list[tuple[str, str]] | None = None) -> dict[str, bool]:
    source_items = items or MONTHLY_CONTROL_ITEMS
    values = {f"item_{index}": False for index in range(1, 15)}
    for key, _label in source_items:
        if "servis etiketi ekipmana" in _label.lower():
            values[key] = True
        else:
            values[key] = form_data.get(key) == "on"
    return values


def build_control_form_values(form_data) -> dict[str, bool]:
    values = {key: False for key, _label in CONTROL_FORM_ITEMS}
    for key, _label in CONTROL_FORM_ITEMS:
        values[key] = form_data.get(key) == "on"
    return values


def derive_ysc_control_form_values(inspection_values: dict[str, bool], control_values: dict[str, bool]) -> dict[str, bool]:
    if any(control_values.values()):
        return control_values
    return {
        "check_a": bool(inspection_values.get("item_1")),
        "check_b": bool(inspection_values.get("item_2")),
        "check_c": bool(inspection_values.get("item_2")),
        "check_d": bool(inspection_values.get("item_3")),
        "check_e": bool(inspection_values.get("item_4")),
        "check_f": bool(inspection_values.get("item_5")),
        "check_g": bool(inspection_values.get("item_6")),
    }


def save_monthly_inspection(
    connection,
    extinguisher_id: int,
    inspection_date: str,
    inspector_name: str,
    notes: str | None,
    inspection_values: dict[str, bool],
    control_values: dict[str, bool],
    created_at: str,
) -> None:
    connection.execute(
        insert(monthly_inspections).values(
            extinguisher_id=extinguisher_id,
            inspection_date=inspection_date,
            inspector_name=inspector_name,
            notes=notes,
            created_at=created_at,
            **inspection_values,
            **control_values,
        )
    )


def with_monthly_control_labels(rows: list[dict], items: list[tuple[str, str]] | None = None) -> list[dict]:
    source_items = items or MONTHLY_CONTROL_ITEMS
    enriched_rows: list[dict] = []
    for row in rows:
        checks = []
        passed_count = 0
        for key, label in source_items:
            passed = True if "servis etiketi ekipmana" in label.lower() else bool(row.get(key))
            if passed:
                passed_count += 1
            checks.append({"key": key, "label": label, "passed": passed})
        copied = dict(row)
        copied["checks"] = checks
        copied["passed_count"] = passed_count
        copied["total_count"] = len(source_items)
        enriched_rows.append(copied)
    return enriched_rows


def get_equipment_preset(extinguisher_type: str | None) -> dict | None:
    if not extinguisher_type:
        return None
    return EQUIPMENT_PRESETS.get(extinguisher_type)


def build_monthly_table(rows: list[dict], items: list[tuple[str, str]] | None = None) -> dict:
    source_items = items or MONTHLY_CONTROL_ITEMS
    if rows:
        sorted_rows = sorted(rows, key=lambda row: row["inspection_date"], reverse=True)
        target_year = datetime.strptime(sorted_rows[0]["inspection_date"], "%Y-%m-%d").year
    else:
        target_year = datetime.now().year

    latest_by_month: dict[int, dict] = {}
    for row in sorted(rows, key=lambda row: row["inspection_date"], reverse=True):
        inspection_date = datetime.strptime(row["inspection_date"], "%Y-%m-%d")
        if inspection_date.year != target_year:
            continue
        if inspection_date.month not in latest_by_month:
            latest_by_month[inspection_date.month] = row

    month_rows = []
    for month_number, month_label in MONTH_LABELS:
        source = latest_by_month.get(month_number)
        cells = []
        for key, label in source_items:
            code = label.split(" ", 1)[0]
            if source is None:
                value = None
            elif "servis etiketi ekipmana" in label.lower():
                value = True
            else:
                value = bool(source.get(key))
            cells.append({"code": code, "value": value})
        month_rows.append(
            {
                "month_label": month_label,
                "inspection_date": source["inspection_date"] if source else None,
                "cells": cells,
            }
        )

    return {
        "year": target_year,
        "headers": [
            {"key": key, "code": label.split(" ", 1)[0]}
            for key, label in source_items
        ],
        "rows": month_rows,
    }


def build_control_form_pdf(company_name: str, extinguishers_for_company: list[dict], latest_inspections: dict[int, dict]) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Taşınabilir Yangın Söndürme Cihazı 6 Aylık Bakım Kontrol Formu", styles["Title"]))
    story.append(Spacer(1, 4 * mm))

    first = extinguishers_for_company[0] if extinguishers_for_company else {}
    general_data = [
        ["FİRMA ADI", company_name, "KONTROL TARİHİ", datetime.now().strftime("%d.%m.%Y")],
        ["MUAYENE ADRESİ", first.get("company_address") or "-", "FİRMA YETKİLİ KİŞİ", first.get("company_contact") or "-"],
        ["PERİYODİK KONTROL PERSONELİ", current_user_full_name() or "-", "AÇIKLAMALAR", "-"],
    ]
    general_table = PdfTable(general_data, colWidths=[32 * mm, 86 * mm, 42 * mm, 90 * mm])
    general_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                ("BACKGROUND", (2, 0), (2, -1), colors.lightgrey),
            ]
        )
    )
    story.append(general_table)
    story.append(Spacer(1, 6 * mm))

    header = [
        "CİHAZ NO",
        "YSC CİNSİ",
        "YSC SINIFI",
        "SERİ NO / KOD",
        "YSC ÜRETİCİ",
        "DOLUM TARİHİ",
        "HİDROSTATİK TEST TARİHİ",
        "BULUNDUĞU YER",
    ] + [label.split(")")[0] + ")" for _key, label in CONTROL_FORM_ITEMS]
    rows = [header]
    for index, extinguisher in enumerate(extinguishers_for_company, start=1):
        inspection = latest_inspections.get(extinguisher["id"], {})
        row = [
            str(index),
            extinguisher.get("extinguisher_type") or "-",
            extinguisher.get("fire_class") or "-",
            extinguisher.get("serial_number") or "-",
            extinguisher.get("manufacturer") or "-",
            extinguisher.get("last_service_date") or "-",
            extinguisher.get("hydrostatic_test_date") or "-",
            extinguisher.get("location_detail") or "-",
        ]
        for key, _label in CONTROL_FORM_ITEMS:
            value = inspection.get(key)
            row.append("✔" if value is True else "X" if value is False and inspection else "─")
        rows.append(row)

    table = PdfTable(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9d6cf")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("NOT 1: Yangın söndürücünün kontrolü Madde a) ve b) bendindeki gibi listelenmiş koşullarda, bir eksikliği ortaya çıkardığı zaman, acil düzeltici faaliyet yapılmalıdır.", styles["BodyText"]))
    story.append(Paragraph("NOT 2: Yangın söndürücünün kontrolü Madde c), d), e), f) veya g) bendindeki koşullarından herhangi birinde bir eksikliği ortaya çıkardığı zaman, söndürücü uygun bakım işlemlerine VESTA YANGIN tarafından tabi tutulmalıdır.", styles["BodyText"]))
    doc.build(story)
    buffer.seek(0)
    return buffer


def build_control_form_pdf_from_template(
    company_name: str,
    extinguishers_for_company: list[dict],
    latest_inspections: dict[int, dict],
) -> io.BytesIO:
    if not CONTROL_FORM_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Kontrol formu sablonu bulunamadi: {CONTROL_FORM_TEMPLATE_PATH}")

    template = fitz.open(CONTROL_FORM_TEMPLATE_PATH)
    output = fitz.open()

    regular_font = "vesta_regular"
    bold_font = "vesta_bold"
    regular_font_path = resolve_system_font(
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    )
    bold_font_path = resolve_system_font(
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    )
    rows_per_page = 15

    def draw_box_text(page, rect, text_value, *, fontname=regular_font, fontsize=7, align=0):
        value = str(text_value or "-").strip() or "-"
        page.insert_textbox(
            fitz.Rect(*rect),
            value,
            fontname=fontname,
            fontsize=fontsize,
            color=(0, 0, 0),
            align=align,
        )

    def draw_page(chunk: list[dict], page_index: int) -> None:
        output.insert_pdf(template, from_page=0, to_page=0)
        page = output[-1]
        page.insert_font(fontname=regular_font, fontfile=regular_font_path)
        page.insert_font(fontname=bold_font, fontfile=bold_font_path)

        first = chunk[0] if chunk else {}
        draw_box_text(page, (110, 78, 608, 92), company_name, fontname=bold_font, fontsize=8)
        draw_box_text(page, (779, 78, 828, 92), datetime.now().strftime("%d.%m.%Y"), fontsize=7)
        draw_box_text(page, (110, 95, 608, 109), first.get("company_address") or "-", fontsize=7)
        draw_box_text(page, (765, 95, 828, 109), first.get("company_contact") or "-", fontsize=7)
        draw_box_text(page, (165, 117, 450, 131), current_user_full_name() or "-", fontsize=7)

        table_top = 375
        row_height = 18.6
        column_edges = [
            (31, 52),
            (52, 96),
            (96, 133),
            (133, 214),
            (214, 278),
            (278, 321),
            (321, 383),
            (383, 524),
            (524, 559),
            (559, 594),
            (594, 628),
            (628, 662),
            (662, 697),
            (697, 731),
            (731, 765),
            (765, 798),
        ]

        for row_number, extinguisher in enumerate(chunk, start=1):
            inspection = latest_inspections.get(extinguisher["id"], {})
            top = table_top + ((row_number - 1) * row_height)
            bottom = top + row_height
            base_values = [
                str((page_index * rows_per_page) + row_number),
                extinguisher.get("extinguisher_type") or "-",
                extinguisher.get("fire_class") or "-",
                extinguisher.get("serial_number") or "-",
                extinguisher.get("manufacturer") or "-",
                extinguisher.get("last_service_date") or "-",
                extinguisher.get("hydrostatic_test_date") or "-",
                extinguisher.get("location_detail") or "-",
            ]

            for index, value in enumerate(base_values):
                x0, x1 = column_edges[index]
                align = 1 if index in {0, 5, 6} else 0
                fontsize = 6 if index in {1, 2, 3, 4, 7} else 7
                draw_box_text(page, (x0 + 2, top + 1, x1 - 2, bottom - 1), value, fontsize=fontsize, align=align)

            for check_index, (key, _label) in enumerate(CONTROL_FORM_ITEMS, start=8):
                x0, x1 = column_edges[check_index]
                symbol = ""
                if inspection:
                    symbol = "✓" if inspection.get(key) else "X"
                draw_box_text(page, (x0, top + 1, x1, bottom - 1), symbol, fontname=bold_font, fontsize=8, align=1)

    for page_index, start in enumerate(range(0, len(extinguishers_for_company), rows_per_page)):
        draw_page(extinguishers_for_company[start : start + rows_per_page], page_index)

    if not extinguishers_for_company:
        draw_page([], 0)

    buffer = io.BytesIO(output.tobytes())
    buffer.seek(0)
    output.close()
    template.close()
    return buffer


def build_control_form_pdf_exact(
    company_name: str,
    extinguishers_for_company: list[dict],
    latest_inspections: dict[int, dict],
) -> io.BytesIO:
    if not CONTROL_FORM_TEMPLATE_IMAGE_PATH.exists():
        raise FileNotFoundError(f"Kontrol formu sablon gorseli bulunamadi: {CONTROL_FORM_TEMPLATE_IMAGE_PATH}")

    page_width, page_height = landscape(A4)
    rows_per_page = 15
    background = ImageReader(str(CONTROL_FORM_TEMPLATE_IMAGE_PATH))

    regular_font_path = resolve_system_font(
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    )
    bold_font_path = resolve_system_font(
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    )
    if "VestaArial" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaArial", regular_font_path))
    if "VestaArialBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaArialBold", bold_font_path))

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))

    def fill_white_box(x0, y0, x1, y1):
        pdf.saveState()
        pdf.setFillColorRGB(1, 1, 1)
        pdf.rect(x0, page_height - y1, x1 - x0, y1 - y0, stroke=0, fill=1)
        pdf.restoreState()

    def draw_text_box(x0, y0, x1, y1, value, *, font="VestaArial", size=7, align="left"):
        text = str(value or "-").strip() or "-"
        box_width = x1 - x0
        lines = simpleSplit(text, font, size, max(box_width - 4, 10))
        max_lines = max(int((y1 - y0) / (size + 1)), 1)
        lines = lines[:max_lines]
        current_y = page_height - y0 - size - 1
        for line in lines:
            line_width = pdfmetrics.stringWidth(line, font, size)
            if align == "center":
                text_x = x0 + max((box_width - line_width) / 2, 0)
            elif align == "right":
                text_x = x1 - line_width - 2
            else:
                text_x = x0 + 2
            pdf.setFont(font, size)
            pdf.drawString(text_x, current_y, line)
            current_y -= size + 1

    def draw_page(chunk: list[dict], page_index: int) -> None:
        pdf.drawImage(background, 0, 0, width=page_width, height=page_height)
        first = chunk[0] if chunk else {}

        fill_white_box(112, 129, 610, 144)
        fill_white_box(748, 129, 812, 144)
        fill_white_box(112, 147, 610, 162)
        fill_white_box(748, 147, 812, 162)

        draw_text_box(112, 129, 610, 144, company_name, font="VestaArialBold", size=7.0)
        draw_text_box(748, 129, 812, 144, datetime.now().strftime("%d.%m.%Y"), size=5.5, align="center")
        draw_text_box(112, 147, 610, 162, first.get("company_address") or "-", size=6.2)
        draw_text_box(748, 147, 812, 162, first.get("company_contact") or "-", size=5.4, align="center")

        table_top = 264
        row_height = 18.55
        column_edges = [
            (32, 52),
            (52, 96),
            (96, 133),
            (133, 214),
            (214, 278),
            (278, 321),
            (321, 383),
            (383, 524),
            (524, 559),
            (559, 594),
            (594, 628),
            (628, 662),
            (662, 697),
            (697, 731),
            (731, 765),
            (765, 798),
        ]

        for row_number, extinguisher in enumerate(chunk, start=1):
            inspection = latest_inspections.get(extinguisher["id"], {})
            top = table_top + ((row_number - 1) * row_height)
            bottom = top + row_height
            values = [
                str((page_index * rows_per_page) + row_number),
                pdf_equipment_label(extinguisher.get("extinguisher_type")),
                extinguisher.get("fire_class") or "-",
                extinguisher.get("serial_number") or "-",
                extinguisher.get("manufacturer") or "-",
                extinguisher.get("last_service_date") or "-",
                extinguisher.get("hydrostatic_test_date") or "-",
                extinguisher.get("location_detail") or "-",
            ]

            for index, value in enumerate(values):
                x0, x1 = column_edges[index]
                align = "center" if index in {0, 5, 6} else "left"
                size = 5.4 if index in {1, 2, 3, 4, 7} else 6.3
                draw_text_box(x0, top, x1, bottom, value, size=size, align=align)

            for check_index, (key, _label) in enumerate(CONTROL_FORM_ITEMS, start=8):
                x0, x1 = column_edges[check_index]
                symbol = ""
                if inspection:
                    symbol = "✓" if inspection.get(key) else "X"
                draw_text_box(x0, top, x1, bottom, symbol, font="VestaArialBold", size=8, align="center")

        pdf.showPage()

    for page_index, start in enumerate(range(0, len(extinguishers_for_company), rows_per_page)):
        draw_page(extinguishers_for_company[start : start + rows_per_page], page_index)

    if not extinguishers_for_company:
        draw_page([], 0)

    pdf.save()
    buffer.seek(0)
    return buffer


def build_company_filename(company_name: str) -> str:
    safe = "".join(char if char.isalnum() else "-" for char in company_name.lower()).strip("-")
    return safe or "kontrol-formu"


def pdf_equipment_label(extinguisher_type: str | None) -> str:
    mapping = {
        "Kuru Kimyevi Toz": "Kuru Kimyevi Toz",
        "CO2": "CO2",
        "Kopuk": "Köpük",
    }
    return mapping.get(extinguisher_type or "", extinguisher_type or "-")


def build_control_form_document_data(public_id: str) -> dict:
    extinguisher = get_extinguisher(public_id)
    company_name = extinguisher["company_name"]
    company_extinguishers = fetch_all(
        select(extinguishers)
        .where(extinguishers.c.company_name == company_name)
        .where(extinguishers.c.asset_category == "Yangin Sondurme Cihazi")
        .order_by(extinguishers.c.location_detail, extinguishers.c.serial_number)
    )
    extinguisher_ids = [row["id"] for row in company_extinguishers]
    latest_inspections: dict[int, dict] = {}
    if extinguisher_ids:
        inspection_rows = fetch_all(
            select(monthly_inspections)
            .where(monthly_inspections.c.extinguisher_id.in_(extinguisher_ids))
            .order_by(
                monthly_inspections.c.extinguisher_id,
                desc(monthly_inspections.c.inspection_date),
                desc(monthly_inspections.c.id),
            )
        )
        for row in inspection_rows:
            latest_inspections.setdefault(row["extinguisher_id"], row)

    latest_log = fetch_one(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
    )
    latest_inspection = latest_inspections.get(extinguisher["id"])
    starter_dates = [
        coerce_date(extinguisher.get("last_service_date")),
        coerce_date((latest_inspection or {}).get("inspection_date")),
        coerce_date((latest_log or {}).get("service_date")),
    ]
    starter_dates = [value for value in starter_dates if value]
    control_date = (
        starter_dates[0].strftime("%d.%m.%Y")
        if starter_dates
        else datetime.now().strftime("%d.%m.%Y")
    )

    control_rows = []
    for index, row in enumerate(company_extinguishers, start=1):
        inspection = latest_inspections.get(row["id"], {})
        control_rows.append(
            {
                "device_no": index,
                "extinguisher_type": pdf_equipment_label(row.get("extinguisher_type")),
                "fire_class": row.get("fire_class") or "-",
                "serial_number": row.get("serial_number") or "-",
                "manufacturer": row.get("manufacturer") or "-",
                "service_date": row.get("last_service_date") or "-",
                "hydrostatic_test_date": row.get("hydrostatic_test_date") or "-",
                "location_detail": row.get("location_detail") or "-",
                "checks": [
                    "✓" if inspection.get(key) else "X" if inspection else "-"
                    for key, _label in CONTROL_FORM_ITEMS
                ],
            }
        )

    return {
        "company_name": company_name,
        "company_address": extinguisher.get("company_address") or "-",
        "company_contact": extinguisher.get("company_contact") or "-",
        "control_date": control_date,
        "inspector_name": current_user_full_name() or "-",
        "method_text": CONTROL_FORM_METHOD_TEXT,
        "rows": control_rows,
        "check_headers": [label for _key, label in CONTROL_FORM_ITEMS],
        "notes": CONTROL_FORM_NOTES,
        "public_id": public_id,
    }


def fetch_company_category_assets(extinguisher: dict) -> list[dict]:
    statement = (
        select(extinguishers)
        .where(extinguishers.c.asset_category == (extinguisher.get("asset_category") or DEFAULT_ASSET_CATEGORY))
        .order_by(extinguishers.c.location_detail, extinguishers.c.serial_number)
    )
    if extinguisher.get("company_id"):
        statement = statement.where(extinguishers.c.company_id == extinguisher["company_id"])
    else:
        statement = statement.where(extinguishers.c.company_name == extinguisher["company_name"])
    return fetch_all(statement)


def build_company_category_report_document_data(public_id: str) -> dict:
    extinguisher = get_extinguisher(public_id)
    asset_profile = get_asset_profile(extinguisher.get("asset_category"))
    category_assets = fetch_company_category_assets(extinguisher)

    columns = [
        {"label": "Seri No", "key": "serial_number", "width": 20 * mm},
        {"label": asset_profile["type_label"], "key": "extinguisher_type", "width": 26 * mm},
        {"label": asset_profile["class_label"], "key": "fire_class", "width": 24 * mm},
        {"label": asset_profile["brand_label"], "key": "manufacturer", "width": 24 * mm},
        {"label": "Bulundugu Yer", "key": "location_detail", "width": 34 * mm},
        {"label": asset_profile["owner_label"], "key": "company_contact", "width": 24 * mm},
        {"label": asset_profile["last_service_label"], "key": "last_service_date", "width": 20 * mm},
        {"label": asset_profile["next_service_label"], "key": "next_service_date", "width": 20 * mm},
    ]
    if asset_profile["show_weight"]:
        columns.insert(5, {"label": "Kg", "key": "weight_kg", "width": 12 * mm})
    if asset_profile["show_pressure"]:
        columns.append({"label": "Basinç", "key": "pressure_status", "width": 18 * mm})
    if asset_profile["show_hydrostatic"]:
        columns.append({"label": "Hidrostatik Test", "key": "hydrostatic_test_date", "width": 20 * mm})

    rows = []
    for asset in category_assets:
        row = {}
        for column in columns:
            value = asset.get(column["key"])
            if column["key"] == "weight_kg":
                row[column["key"]] = f"{value:.1f}" if isinstance(value, (int, float)) else "-"
            else:
                row[column["key"]] = value or "-"
        rows.append(row)

    return {
        "company_name": extinguisher["company_name"],
        "company_address": extinguisher.get("company_address") or "-",
        "asset_category": extinguisher.get("asset_category") or DEFAULT_ASSET_CATEGORY,
        "asset_profile": asset_profile,
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "columns": columns,
        "rows": rows,
    }


def build_scba_company_document_data(public_id: str) -> dict:
    extinguisher = get_extinguisher(public_id)
    category_assets = fetch_company_category_assets(extinguisher)
    asset_ids = [row["id"] for row in category_assets]
    latest_inspections: dict[int, dict] = {}
    if asset_ids:
        inspection_rows = fetch_all(
            select(monthly_inspections)
            .where(monthly_inspections.c.extinguisher_id.in_(asset_ids))
            .order_by(
                monthly_inspections.c.extinguisher_id,
                desc(monthly_inspections.c.inspection_date),
                desc(monthly_inspections.c.id),
            )
        )
        for row in inspection_rows:
            latest_inspections.setdefault(row["extinguisher_id"], row)

    latest_service_log = fetch_one(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
    )
    latest_direct_inspection = fetch_one(
        select(monthly_inspections)
        .where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(monthly_inspections.c.inspection_date), desc(monthly_inspections.c.id))
        .limit(1)
    )
    periodic_control_person = (
        (latest_service_log or {}).get("technician_name")
        or (latest_direct_inspection or {}).get("inspector_name")
        or "-"
    )

    rows = []
    for index, row in enumerate(category_assets, start=1):
        inspection = latest_inspections.get(row["id"], {})
        checks = []
        for key, _label in SCBA_CONTROL_ITEMS:
            if key == "item_7":
                checks.append("✓")
            elif inspection:
                checks.append("✓" if inspection.get(key) else "X")
            else:
                checks.append("✓")
        rows.append(
            {
                "device_no": index,
                "category": row.get("fire_class") or "-",
                "serial_number": row.get("serial_number") or "-",
                "manufacturer": row.get("manufacturer") or "-",
                "service_date": row.get("last_service_date") or "-",
                "hydrostatic_test_date": row.get("hydrostatic_test_date") or "-",
                "location_detail": row.get("location_detail") or "-",
                "checks": checks,
            }
        )
        rows[-1]["checks"] = checks

    return {
        "company_name": extinguisher["company_name"],
        "company_address": extinguisher.get("company_address") or "-",
        "company_contact": extinguisher.get("company_contact") or "-",
        "periodic_control_person": periodic_control_person,
        "approval_name": "Mustafa Kilic",
        "asset_category": "SCBA",
        "control_date": control_date,
        "method_text": SCBA_METHOD_TEXT,
        "rows": rows,
        "notes": SCBA_NOTES,
    }


def build_special_category_company_document_data(public_id: str) -> dict:
    extinguisher = get_extinguisher(public_id)
    category = extinguisher.get("asset_category") or DEFAULT_ASSET_CATEGORY
    config = SPECIAL_CATEGORY_FORM_CONFIGS[category]
    asset_profile = get_asset_profile(category)
    category_assets = fetch_company_category_assets(extinguisher)
    asset_ids = [row["id"] for row in category_assets]
    latest_inspections: dict[int, dict] = {}
    latest_logs: dict[int, dict] = {}
    if asset_ids:
        inspection_rows = fetch_all(
            select(monthly_inspections)
            .where(monthly_inspections.c.extinguisher_id.in_(asset_ids))
            .order_by(
                monthly_inspections.c.extinguisher_id,
                desc(monthly_inspections.c.inspection_date),
                desc(monthly_inspections.c.id),
            )
        )
        for row in inspection_rows:
            latest_inspections.setdefault(row["extinguisher_id"], row)
        service_rows = fetch_all(
            select(service_logs)
            .where(service_logs.c.extinguisher_id.in_(asset_ids))
            .order_by(
                service_logs.c.extinguisher_id,
                desc(service_logs.c.service_date),
                desc(service_logs.c.id),
            )
        )
        for row in service_rows:
            latest_logs.setdefault(row["extinguisher_id"], row)

    latest_log = latest_logs.get(extinguisher["id"]) or fetch_one(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
    )
    latest_inspection = latest_inspections.get(extinguisher["id"]) or fetch_one(
        select(monthly_inspections)
        .where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(monthly_inspections.c.inspection_date), desc(monthly_inspections.c.id))
        .limit(1)
    )

    def coerce_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, "strftime"):
            return value
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except Exception:
            try:
                return datetime.strptime(str(value), "%d.%m.%Y").date()
            except Exception:
                return None

    control_date_candidates = []
    rows = []
    for index, row in enumerate(category_assets, start=1):
        inspection = latest_inspections.get(row["id"])
        last_log = latest_logs.get(row["id"])
        checks = []
        for key, label in asset_profile["monthly_control_items"]:
            if "servis etiketi ekipmana" in label.lower():
                checks.append("✓")
            elif inspection:
                checks.append("✓" if inspection.get(key) else "X")
            else:
                checks.append("-")
        for candidate in [
            row.get("last_service_date"),
            (inspection or {}).get("inspection_date"),
            (last_log or {}).get("service_date"),
        ]:
            parsed = coerce_date(candidate)
            if parsed:
                control_date_candidates.append(parsed)
        rows.append(
            {
                "device_no": index,
                "category": row.get("fire_class") or row.get("extinguisher_type") or asset_profile["label"],
                "serial_number": row.get("serial_number") or "-",
                "manufacturer": row.get("manufacturer") or "-",
                "service_date": row.get("last_service_date") or "-",
                "hydrostatic_test_date": row.get("hydrostatic_test_date") or "-",
                "location_detail": row.get("location_detail") or "-",
                "company_contact": row.get("company_contact") or "-",
                "checks": checks,
            }
        )

    starter_dates = [
        coerce_date(extinguisher.get("last_service_date")),
        coerce_date((latest_inspection or {}).get("inspection_date")),
        coerce_date((latest_log or {}).get("service_date")),
    ]
    starter_dates = [value for value in starter_dates if value]
    control_date = (
        starter_dates[0].strftime("%d.%m.%Y")
        if starter_dates
        else max(control_date_candidates).strftime("%d.%m.%Y")
        if control_date_candidates
        else datetime.now().strftime("%d.%m.%Y")
    )

    return {
        "company_name": extinguisher["company_name"],
        "company_address": extinguisher.get("company_address") or "-",
        "company_contact": extinguisher.get("company_contact") or "-",
        "periodic_control_person": (latest_log or {}).get("technician_name") or (latest_inspection or {}).get("inspector_name") or "-",
        "approval_name": "Mustafa Kilic",
        "asset_category": category,
        "asset_profile": asset_profile,
        "form_code": config["form_code"],
        "subject": config["subject"],
        "section_title": config["section_title"],
        "control_date": control_date,
        "method_text": config["method_text"],
        "rows": rows,
        "check_headers": [label for _key, label in asset_profile["monthly_control_items"]],
        "notes": config["notes"],
    }


def build_special_category_company_form_pdf(document_data: dict) -> io.BytesIO:
    regular_font_path = resolve_system_font(
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    )
    bold_font_path = resolve_system_font(
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    )
    if "VestaPDF" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDF", regular_font_path))
    if "VestaPDFBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDFBold", bold_font_path))

    styles = getSampleStyleSheet()
    body_style = styles["BodyText"].clone("special_body")
    body_style.fontName = "VestaPDF"
    body_style.fontSize = 7
    body_style.leading = 8

    tiny_style = styles["BodyText"].clone("special_tiny")
    tiny_style.fontName = "VestaPDF"
    tiny_style.fontSize = 5
    tiny_style.leading = 6

    tiny_bold_style = styles["BodyText"].clone("special_tiny_bold")
    tiny_bold_style.fontName = "VestaPDFBold"
    tiny_bold_style.fontSize = 5
    tiny_bold_style.leading = 6

    compact_style = styles["BodyText"].clone("special_compact")
    compact_style.fontName = "VestaPDF"
    compact_style.fontSize = 5.0
    compact_style.leading = 5.5
    compact_style.spaceBefore = 0
    compact_style.spaceAfter = 0

    compact_date_style = styles["BodyText"].clone("special_compact_date")
    compact_date_style.fontName = "VestaPDF"
    compact_date_style.fontSize = 4.6
    compact_date_style.leading = 5.0
    compact_date_style.spaceBefore = 0
    compact_date_style.spaceAfter = 0
    compact_date_style.alignment = 1

    compact_category_style = styles["BodyText"].clone("special_category")
    compact_category_style.fontName = "VestaPDF"
    compact_category_style.fontSize = 4.5
    compact_category_style.leading = 4.9
    compact_category_style.spaceBefore = 0
    compact_category_style.spaceAfter = 0
    compact_category_style.alignment = 1

    compact_location_style = styles["BodyText"].clone("special_location")
    compact_location_style.fontName = "VestaPDF"
    compact_location_style.fontSize = 4.2
    compact_location_style.leading = 4.6
    compact_location_style.spaceBefore = 0
    compact_location_style.spaceAfter = 0
    compact_location_style.alignment = 1

    target_width = 280 * mm

    def split_check_header(label: str) -> str:
        parts = label.split(" ", 1)
        clean = parts[1] if len(parts) > 1 else label
        replacements = {
            "Yuz Maskesi": "Yuz<br/>Maskesi",
            "Solunum Valfi": "Solunum<br/>Valfi",
            "Regulator Unitesi": "Regulator<br/>Unitesi",
            "Servis Etiketi Ekipmana Yapistirildi": "Servis<br/>Etiketi<br/>Ekipmana<br/>Yapistirildi",
            "Servis etiketi ekipmana yapistirildi": "Servis<br/>Etiketi<br/>Ekipmana<br/>Yapistirildi",
            "Servis Etiketi Ekipmana Yapıştırıldı": "Servis<br/>Etiketi<br/>Ekipmana<br/>Yapistirildi",
            "Valfi Kontrol Edildi": "Valfi<br/>Kontrol<br/>Edildi",
            "Silindir Kontrol Edildi": "Silindir<br/>Kontrol<br/>Edildi",
            "Gorsel Kumas Kontrol yapildi": "Gorsel<br/>Kumas<br/>Kontrolu",
            "Fonksiyonel Fermuar, Cirt cirtlar ve Dugmeler kontrol edildi": "Fermuar,<br/>Cirt Cirt<br/>ve Dugmeler",
            "Yansitici Bantlar Kontrol Edildi": "Yansitici<br/>Bantlar",
            "Elbiselerin Temizligi Kontrol Edildi": "Elbise<br/>Temizligi",
            "Ic Astarin Durumu, Yalitim Ozelligi ve Dikisler Kontrol Edildi": "Ic Astar,<br/>Yalitim ve<br/>Dikisler",
            "Dis kisimda gorsel kontrol yapildi": "Dis Kisimda<br/>Gorsel<br/>Kontrol",
            "Ic kisimda gorsel kontrol yapildi": "Ic Kisimda<br/>Gorsel<br/>Kontrol",
            "Ayar mekanizmasinda kontrol yapildi": "Ayar<br/>Mekanizmasi",
            "Vizor ve goz korumasinda kontrol yapildi": "Vizor ve<br/>Goz<br/>Koruma",
            "Boyun koruyucu kontrolu yapildi": "Boyun<br/>Koruyucu",
            "Baltanin ahsap veya yalitkan sapi kontrol edildi": "Sap<br/>Kontrolu",
            "Metal kismi kontrol edildi": "Metal Kisim",
            "Agiz kismi kontrol edildi": "Agiz Kismi",
            "Bulundugu yerde ulasilabilir durumda mi": "Ulasilabilir<br/>Durumda",
            "Erisebilirlik Kontrol Edildi (Dolap onu acik mi, istif veya engel malz. var mi)": "Erisebilirlik",
            "Levhalar Kontrol Edildi (Yangin dolabi isareti ve talimati mevcut mu)": "Levhalar",
            "Kapak ve Kilit Kontrolu Yapildi (Kapak rahat aciliyor mu, kilit saglam mi)": "Kapak ve<br/>Kilit",
            "Dolap Dis Yuzey Kontrol Edildi (Paslanma veya boya kabarmasi var mi)": "Dis Yuzey",
            "Makara Kontrol Edildi (Kolayca acilabiliyor mu)": "Makara",
            "Hortum Kontrol Edildi (Catlama, kirilma, sertlesme veya kacak var mi)": "Hortum",
            "Baglanti Rekorlari Kontrol Edildi (Hortum vana ve lans baglantilari siki mi)": "Baglanti<br/>Rekorlari",
            "Dolap ici ve disi temizlik kontrolu yapildi": "Dolap Ici<br/>Temizlik",
            "Vana kontrol edildi (Vana kolu rahat donuyor mu, kacak veya sizdirma var mi)": "Vana",
            "Basinc Kontrol Edildi (Statik ve dinamik basinc degerleri uygun mu) (min.4 bar)": "Basinc",
            "Lans Kontrol Edildi (Jet/Spray/Kapali konumlari islevsel mi)": "Lans",
            "Kopuk Doluluk Orani Kontrol Edildi (Seviye gostergesi kontrolu)": "Kopuk<br/>Doluluk",
            "Kopuk Oranlayici Ayarlari Kontrol Edildi (Mix ayari dogru yuzde mi %1,%3 veya %6)": "Oranlayici<br/>Ayarlari",
            "Vana kontrol edildi (Ana su giris vanasi ve kopuk vanasi islevsel mi)": "Vana",
            "Basinc Kontrol Edildi (Sistem calisma basinci kopuk olusumu icin yeterli mi) (min.5-6 bar)": "Basinc",
            "Kopuk Lans Kontrol Edildi (Kopuk yapici ozel lans saglam mi)": "Kopuk<br/>Lans",
            "Erisebilirlik Kontrol Edildi (Hidrant cevresinde arac, malzeme engeli var mi)": "Erisebilirlik",
            "Gorunurlugu kontrol edildi (Hidrantin kirmizi boyasi canli mi, yonlendirme levhalari var mi)": "Gorunurluk",
            "Kapak Kontrolu Yapildi (Cikis agzindaki kor tapalar/kapaklar takili mi, zincirleri saglam mi)": "Kapak",
            "Genel Dis Yuzey Kontrol Yapildi (Govdede catlak, korozyon veya darbe izi var mi)": "Dis Yuzey",
            "Acma kapama mili kontrol edildi (Hidrant anahtari ile mil rahatca donuyor mu)": "Acma Kapama<br/>Mili",
            "Vana Sizdirmazligi Kontrol Edildi (Hidrant kapaliyken cikis agzindan veya govde altindan su sizintisi var mi)": "Vana<br/>Sizdirmazligi",
            "Cikis agizlari kontrol edildi (Rekor dislerinde veya tirnaklarinda asinma veya deformasyon var mi)": "Cikis<br/>Agizlari",
        }
        return replacements.get(clean, clean.replace(" (", "<br/>(").replace(" Kontrol Edildi", "<br/>Kontrol")).strip()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=6 * mm,
        bottomMargin=6 * mm,
    )
    story = []

    logo_cells = []
    if TSE_HYB_LOGO_PATH.exists():
        logo_cells.append(PdfImage(str(TSE_HYB_LOGO_PATH), width=17 * mm, height=17 * mm))
    else:
        logo_cells.append(Paragraph("<b>TSE-HYB</b>", body_style))
    if VESTA_HEADER_LOGO_PATH.exists():
        logo_cells.append(PdfImage(str(VESTA_HEADER_LOGO_PATH), width=13 * mm, height=16 * mm))
    else:
        logo_cells.append(Paragraph("<b>VESTA</b>", body_style))

    logo_table = PdfTable([logo_cells], colWidths=[20 * mm, 20 * mm], rowHeights=[18 * mm])
    logo_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (-1, -1), "CENTER")]))

    header_table = PdfTable(
        [
            [
                logo_table,
                Paragraph("<b>Onay: Sirket Muduru<br/>Mustafa Kilic</b>", body_style),
                Paragraph("<b>Bolum</b>", body_style),
                Paragraph("<b>Revizyon No</b>", body_style),
                Paragraph("<b>Revizyon Tarihi</b>", body_style),
                Paragraph("<b>Sayfa</b>", body_style),
            ],
            ["", "", document_data["form_code"], "0", "2026-03-01", "1"],
            [
                Paragraph("<b>FORM</b>", body_style),
                Paragraph(f"<b>Konu: {document_data['subject']}</b>", body_style),
                Paragraph("<b>Hazirlayan: Kalite Temsilcisi</b>", body_style),
                "",
                "",
                "",
            ],
        ],
        colWidths=[46 * mm, 100 * mm, 18 * mm, 22 * mm, 42 * mm, 52 * mm],
        hAlign="CENTER",
    )
    header_table.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("SPAN", (0, 0), (0, 1)),
            ("SPAN", (1, 0), (1, 1)),
            ("SPAN", (1, 2), (2, 2)),
            ("SPAN", (3, 2), (5, 2)),
            ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
            ("FONTNAME", (0, 0), (-1, 0), "VestaPDFBold"),
            ("FONTNAME", (0, 2), (-1, 2), "VestaPDFBold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (-1, 1), "CENTER"),
        ])
    )
    story.extend([header_table, Spacer(1, 1 * mm)])

    info_table = PdfTable(
        [
            [Paragraph("<b>FIRMA ADI</b>", body_style), document_data["company_name"], "", "", Paragraph("<b>KONTROL TARIHI</b>", body_style), document_data["control_date"]],
            [Paragraph("<b>MUAYENE ADRESI</b>", body_style), document_data["company_address"], "", "", Paragraph("<b>FIRMA YETKILISI</b>", body_style), document_data["company_contact"]],
            [Paragraph("<b>PERIYODIK KONTROL METODU</b>", body_style), document_data["method_text"], "", "", "", ""],
        ],
        colWidths=[34 * mm, 108 * mm, 2 * mm, 2 * mm, 52 * mm, 82 * mm],
        hAlign="CENTER",
    )
    info_table.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("SPAN", (1, 0), (3, 0)),
            ("SPAN", (1, 1), (3, 1)),
            ("SPAN", (1, 2), (-1, 2)),
            ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
            ("FONTNAME", (0, 0), (0, -1), "VestaPDFBold"),
            ("FONTNAME", (4, 0), (4, 1), "VestaPDFBold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ])
    )
    story.extend([info_table, Spacer(1, 1 * mm)])

    check_headers = document_data["check_headers"]
    check_count = len(check_headers)
    first_col_widths = [12 * mm, 18 * mm, 20 * mm, 18 * mm, 16 * mm]
    if document_data["asset_profile"].get("show_hydrostatic"):
        first_col_widths.append(16 * mm)
    first_col_widths.append(22 * mm)
    remaining_width = target_width - sum(first_col_widths)
    check_width = max(7.2 * mm, remaining_width / max(check_count, 1))
    col_widths = first_col_widths + [check_width] * check_count

    static_col_count = len(first_col_widths)
    header_row_1 = [Paragraph(f"<b>{document_data['section_title']}</b>", body_style)] + [""] * (static_col_count - 1) + [Paragraph("<b>TESPIT VE DEGERLENDIRME</b><br/><font size='6'>(✓: UYGUN, X: UYGUN DEGIL, -: UYGULAMA YOK)</font>", tiny_bold_style)] + [""] * (check_count - 1)
    header_row_2 = [
        Paragraph("<b>CIHAZ NO</b>", tiny_bold_style),
        Paragraph("<b>KATEGORI / CINSI</b>", tiny_bold_style),
        Paragraph("<b>SERI NO</b>", tiny_bold_style),
        Paragraph("<b>MARKA</b>", tiny_bold_style),
        Paragraph(f"<b>{document_data['asset_profile']['last_service_label'].upper()}</b>", tiny_bold_style),
    ]
    if document_data["asset_profile"].get("show_hydrostatic"):
        header_row_2.append(Paragraph("<b>HIDROSTATIK<br/>TEST TARIHI</b>", tiny_bold_style))
    header_row_2.append(Paragraph("<b>BULUNDUGU YER</b>", tiny_bold_style))
    header_row_2.extend([Paragraph(f"<b>{split_check_header(label)}</b>", tiny_style) for label in check_headers])

    data_rows = [header_row_1, header_row_2]
    for row in document_data["rows"]:
        current = [
            row["device_no"],
            Paragraph(str(row["category"]).replace("BASINCLI HAVA KACIS SETI", "BASINCLI<br/>HAVA KACIS<br/>SETI").replace("BASINCLI HAVA SOLUNUM TUPU", "BASINCLI HAVA<br/>SOLUNUM TUPU").replace("Yangin Elbisesi", "Yangin<br/>Elbisesi").replace("Yangin Bareti", "Yangin<br/>Bareti").replace("Yangin Baltasi", "Yangin<br/>Baltasi").replace("Yangin Sondurme Dolabi", "Yangin Sondurme<br/>Dolabi").replace("Kopuklu Yangin Sondurme Dolabi", "Kopuklu Yangin<br/>Sondurme Dolabi").replace("Yangin Hidranti", "Yangin<br/>Hidranti"), compact_category_style),
            Paragraph(str(row["serial_number"]), compact_style),
            Paragraph(str(row["manufacturer"]), compact_style),
            Paragraph(str(row["service_date"]), compact_date_style),
        ]
        if document_data["asset_profile"].get("show_hydrostatic"):
            current.append(Paragraph(str(row["hydrostatic_test_date"]), compact_date_style))
        current.append(Paragraph(str(row["location_detail"]), compact_location_style))
        current.extend(row["checks"])
        data_rows.append(current)

    min_rows = 12
    while len(data_rows) < (2 + min_rows):
        data_rows.append([""] * len(col_widths))

    main_table = PdfTable(
        data_rows,
        repeatRows=2,
        colWidths=col_widths,
        rowHeights=[6 * mm, 18 * mm] + [5.1 * mm] * (len(data_rows) - 2),
        hAlign="CENTER",
    )
    main_table.setStyle(
        TableStyle([
            ("SPAN", (0, 0), (static_col_count - 1, 0)),
            ("SPAN", (static_col_count, 0), (-1, 0)),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f7f7f7")),
            ("FONTNAME", (0, 0), (-1, 1), "VestaPDFBold"),
            ("FONTNAME", (0, 2), (-1, -1), "VestaPDF"),
            ("FONTSIZE", (0, 0), (-1, -1), 5),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    story.extend([main_table, Spacer(1, 1 * mm)])

    note_style = styles["BodyText"].clone("special_note")
    note_style.fontName = "VestaPDF"
    note_style.fontSize = 5
    note_style.leading = 6
    for note in document_data["notes"]:
        story.append(Paragraph(note, note_style))

    footer_table = PdfTable(
        [
            [Paragraph("<b>ACIKLAMALAR</b>", body_style)],
            [""],
            [Paragraph("<b>PERIYODIK KONTROL PERSONELI</b>", tiny_bold_style), Paragraph("<b>FIRMA YETKILISI</b>", tiny_bold_style)],
            [document_data["periodic_control_person"], document_data["company_contact"]],
            [Paragraph("<b>ONAY</b>", body_style), document_data["approval_name"]],
        ],
        colWidths=[target_width / 2, target_width / 2],
        hAlign="CENTER",
    )
    footer_table.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("SPAN", (0, 0), (-1, 0)),
            ("SPAN", (0, 1), (-1, 1)),
            ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
            ("FONTNAME", (0, 0), (-1, 0), "VestaPDFBold"),
            ("FONTNAME", (0, 2), (-1, 2), "VestaPDFBold"),
            ("FONTNAME", (0, 4), (0, 4), "VestaPDFBold"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("MINROWHEIGHT", (0, 1), (-1, 1), 16 * mm),
        ])
    )
    story.extend([Spacer(1, 1 * mm), footer_table])

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_control_form_pdf_reportlab(document_data: dict) -> io.BytesIO:
    main_col_widths = [10 * mm, 14 * mm, 14 * mm, 19 * mm, 16 * mm, 14 * mm, 16 * mm, 24 * mm, 11 * mm, 11 * mm, 11 * mm, 11 * mm, 11 * mm, 11 * mm, 11 * mm]
    form_total_width = sum(main_col_widths)

    regular_font_path = resolve_system_font(
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    )
    bold_font_path = resolve_system_font(
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    )

    if "VestaPDF" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDF", regular_font_path))
    if "VestaPDFBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDFBold", bold_font_path))

    styles = getSampleStyleSheet()
    body_style = styles["BodyText"].clone("vesta_body")
    body_style.fontName = "VestaPDF"
    body_style.fontSize = 7
    body_style.leading = 8

    small_style = styles["BodyText"].clone("vesta_small")
    small_style.fontName = "VestaPDF"
    small_style.fontSize = 6
    small_style.leading = 7

    tiny_style = styles["BodyText"].clone("vesta_tiny")
    tiny_style.fontName = "VestaPDF"
    tiny_style.fontSize = 5
    tiny_style.leading = 6

    tiny_bold_style = styles["BodyText"].clone("vesta_tiny_bold")
    tiny_bold_style.fontName = "VestaPDFBold"
    tiny_bold_style.fontSize = 5
    tiny_bold_style.leading = 6

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=6 * mm,
        bottomMargin=6 * mm,
    )

    story = []

    logo_cells = []
    if TSE_HYB_LOGO_PATH.exists():
        logo_cells.append(PdfImage(str(TSE_HYB_LOGO_PATH), width=17 * mm, height=17 * mm))
    else:
        logo_cells.append(Paragraph("<b>TSE-HYB</b>", body_style))
    if VESTA_HEADER_LOGO_PATH.exists():
        logo_cells.append(PdfImage(str(VESTA_HEADER_LOGO_PATH), width=13 * mm, height=16 * mm))
    else:
        logo_cells.append(Paragraph("<b>VESTA</b>", body_style))

    logo_table = PdfTable([logo_cells], colWidths=[20 * mm, 20 * mm], rowHeights=[18 * mm])
    logo_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )

    header_table = PdfTable(
        [
            [
                logo_table,
                Paragraph("<b>Onay: Şirket Müdürü<br/>Mustafa KİLİÇ</b>", body_style),
                Paragraph("<b>Bölüm</b>", body_style),
                Paragraph("<b>Revizyon No</b>", body_style),
                Paragraph("<b>Revizyon Tarihi</b>", body_style),
                Paragraph("<b>Sayfa</b>", body_style),
            ],
            [
                "",
                "",
                "F-06",
                "00",
                "01.03.2026",
                "1",
            ],
            [
                Paragraph("<b>FORM</b>", body_style),
                Paragraph("<b>Konu: Taşınabilir Yangın Söndürme Cihazı Kontrol Formu</b>", body_style),
                Paragraph("<b>Hazırlayan: Kalite Temsilcisi</b>", body_style),
                "",
                "",
                "",
            ],
        ],
        colWidths=[40 * mm, 60 * mm, 16 * mm, 22 * mm, 32 * mm, 34 * mm],
        hAlign="CENTER",
    )
    header_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("SPAN", (0, 0), (0, 1)),
                ("SPAN", (1, 0), (1, 1)),
                ("SPAN", (1, 2), (2, 2)),
                ("SPAN", (3, 2), (5, 2)),
                ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
                ("FONTNAME", (0, 0), (-1, 0), "VestaPDFBold"),
                ("FONTNAME", (0, 2), (-1, 2), "VestaPDFBold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (2, 0), (-1, 1), "CENTER"),
            ]
        )
    )
    story.extend([header_table, Spacer(1, 1 * mm)])

    compact_small_style = styles["BodyText"].clone("vesta_compact_small")
    compact_small_style.fontName = "VestaPDF"
    compact_small_style.fontSize = 5
    compact_small_style.leading = 6
    compact_small_style.spaceBefore = 0
    compact_small_style.spaceAfter = 0

    compact_cell_style = styles["BodyText"].clone("vesta_compact_cell")
    compact_cell_style.fontName = "VestaPDF"
    compact_cell_style.fontSize = 4.4
    compact_cell_style.leading = 4.8
    compact_cell_style.spaceBefore = 0
    compact_cell_style.spaceAfter = 0
    compact_cell_style.alignment = 1

    method_paragraph = Paragraph(document_data["method_text"], compact_small_style)
    address_cell_style = styles["BodyText"].clone("vesta_address_cell")
    address_cell_style.fontName = "VestaPDF"
    address_cell_style.fontSize = 4.6
    address_cell_style.leading = 5
    address_cell_style.spaceBefore = 0
    address_cell_style.spaceAfter = 0

    info_table = PdfTable(
        [
            [Paragraph("<b>GENEL BİLGİLER</b>", body_style), "", "", "", "", ""],
            ["FIRMA ADI", document_data["company_name"], "", "", "KONTROL TARIHI", document_data["control_date"]],
            ["MUAYENE ADRESI", Paragraph(document_data["company_address"], address_cell_style), "", "", "FIRMA YETKILI KISI", document_data["company_contact"]],
            ["PERIYODIK KONTROL METODU", method_paragraph, "", "", "", ""],
        ],
        colWidths=[56 * mm, 58 * mm, 2 * mm, 2 * mm, 34 * mm, 52 * mm],
        hAlign="CENTER",
        rowHeights=[7 * mm, 7 * mm, 7 * mm, 9 * mm],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("SPAN", (0, 0), (-1, 0)),
                ("SPAN", (1, 1), (3, 1)),
                ("SPAN", (1, 2), (3, 2)),
                ("SPAN", (1, 3), (-1, 3)),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f1f1")),
                ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
                ("FONTNAME", (0, 0), (-1, 0), "VestaPDFBold"),
                ("FONTNAME", (0, 1), (0, -1), "VestaPDFBold"),
                ("FONTNAME", (4, 1), (4, 2), "VestaPDFBold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 0.8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (1, 3), (1, 3), 10),
            ]
        )
    )
    story.extend([info_table, Spacer(1, 1 * mm)])

    header_row_1 = [
        Paragraph("<b>YANGIN SÖNDÜRME CİHAZI (YSC) BİLGİLERİ</b>", body_style),
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        Paragraph(
            "<b>TESPİT VE DEĞERLENDİRME</b><br/><font size='6'>(✓: UYGUN, X: UYGUN DEĞİL, -: UYGULAMA YOK)</font>",
            tiny_bold_style,
        ),
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    rotated_width = 11 * mm
    rotated_height = 30 * mm

    header_row_2 = [
        Paragraph("<b>CİHAZ</b><br/><b>NO</b>", tiny_bold_style),
        Paragraph("<b>YSC CİNSİ</b>", tiny_bold_style),
        Paragraph("<b>YSC SINIFI</b>", tiny_bold_style),
        Paragraph("<b>SERİ NO / KOD</b>", tiny_bold_style),
        Paragraph("<b>YSC ÜRETİCİ</b>", tiny_bold_style),
        Paragraph("<b>DOLUM TARİHİ</b>", tiny_bold_style),
        Paragraph("<b>HİDROSTATİK</b><br/><b>TEST TARİHİ</b>", tiny_bold_style),
        Paragraph("<b>BULUNDUĞU YER</b>", tiny_bold_style),
        *[
            RotatedParagraph(f"<b>{label}</b>", tiny_style, rotated_width, rotated_height)
            for label in document_data["check_headers"]
        ],
    ]

    data_rows = [header_row_1, header_row_2]
    for row in document_data["rows"]:
        extinguisher_type = str(row["extinguisher_type"]).replace("Kuru Kimyevi Toz", "Kuru<br/>Kimyevi Toz")
        data_rows.append(
            [
                row["device_no"],
                Paragraph(extinguisher_type, compact_cell_style),
                Paragraph(str(row["fire_class"]), compact_cell_style),
                Paragraph(str(row["serial_number"]), compact_cell_style),
                Paragraph(str(row["manufacturer"]), compact_cell_style),
                Paragraph(str(row["service_date"]), compact_cell_style),
                Paragraph(str(row["hydrostatic_test_date"]), compact_cell_style),
                Paragraph(str(row["location_detail"]), compact_cell_style),
                *row["checks"],
            ]
        )

    while len(data_rows) < 17:
        data_rows.append([""] * 15)

    main_table = PdfTable(
        data_rows,
        repeatRows=2,
        colWidths=main_col_widths,
        rowHeights=[6 * mm, 30 * mm] + [4.8 * mm] * (len(data_rows) - 2),
        hAlign="CENTER",
    )
    main_table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (7, 0)),
                ("SPAN", (8, 0), (14, 0)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f7f7f7")),
                ("FONTNAME", (0, 0), (-1, 1), "VestaPDFBold"),
                ("FONTNAME", (0, 2), (-1, -1), "VestaPDF"),
                ("FONTSIZE", (0, 0), (-1, -1), 5),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.extend([main_table, Spacer(1, 1 * mm)])

    note_style = styles["BodyText"].clone("vesta_note")
    note_style.fontName = "VestaPDF"
    note_style.fontSize = 5
    note_style.leading = 6
    note_style.leftIndent = 0
    note_style.spaceBefore = 0
    note_style.spaceAfter = 0

    notes_table = PdfTable(
        [[Paragraph(note, note_style)] for note in document_data["notes"]],
        colWidths=[form_total_width],
        hAlign="CENTER",
    )
    notes_table.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(notes_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_company_category_report_pdf(document_data: dict) -> io.BytesIO:
    regular_font_path = resolve_system_font(
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    )
    bold_font_path = resolve_system_font(
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    )
    if "VestaPDF" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDF", regular_font_path))
    if "VestaPDFBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDFBold", bold_font_path))

    styles = getSampleStyleSheet()
    title_style = styles["Title"].clone("category_report_title")
    title_style.fontName = "VestaPDFBold"
    title_style.fontSize = 16
    title_style.leading = 19
    title_style.textColor = colors.HexColor("#3f2319")

    body_style = styles["BodyText"].clone("category_report_body")
    body_style.fontName = "VestaPDF"
    body_style.fontSize = 8
    body_style.leading = 10

    cell_style = styles["BodyText"].clone("category_report_cell")
    cell_style.fontName = "VestaPDF"
    cell_style.fontSize = 7
    cell_style.leading = 8

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    story = [
        Paragraph(
            f"{document_data['company_name']} - Firma {document_data['asset_profile']['label']} Kayıt Raporu",
            title_style,
        ),
        Spacer(1, 3 * mm),
    ]

    info_table = PdfTable(
        [
            ["Firma", document_data["company_name"], "Urun Grubu", document_data["asset_category"]],
            ["Muayene Adresi", document_data["company_address"], "Kayit Sayisi", str(len(document_data["rows"]))],
            ["Rapor Olusturma Tarihi", document_data["generated_at"], "Aciklama", "Bu rapor aynı firmadaki aynı kategori kayıtlarını listeler."],
        ],
        colWidths=[34 * mm, 92 * mm, 34 * mm, 112 * mm],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c8b2a5")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1e1d9")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f1e1d9")),
                ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
                ("FONTNAME", (0, 0), (0, -1), "VestaPDFBold"),
                ("FONTNAME", (2, 0), (2, -1), "VestaPDFBold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.extend([info_table, Spacer(1, 4 * mm)])

    headers = [Paragraph(f"<b>{column['label']}</b>", cell_style) for column in document_data["columns"]]
    table_rows = [headers]
    for row in document_data["rows"]:
        table_rows.append(
            [
                Paragraph(str(row.get(column["key"], "-")), cell_style)
                for column in document_data["columns"]
            ]
        )

    data_table = PdfTable(
        table_rows,
        repeatRows=1,
        colWidths=[column["width"] for column in document_data["columns"]],
    )
    data_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#c8b2a5")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9d6cf")),
                ("FONTNAME", (0, 0), (-1, 0), "VestaPDFBold"),
                ("FONTNAME", (0, 1), (-1, -1), "VestaPDF"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(data_table)
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph("Bu rapor, firma portalında seçilen kategoriye ait tüm kayıtları tek dosyada sunmak için otomatik oluşturulmuştur.", body_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_scba_company_form_pdf(document_data: dict) -> io.BytesIO:
    regular_font_path = resolve_system_font(
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    )
    bold_font_path = resolve_system_font(
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    )
    if "VestaPDF" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDF", regular_font_path))
    if "VestaPDFBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VestaPDFBold", bold_font_path))

    styles = getSampleStyleSheet()
    body_style = styles["BodyText"].clone("scba_body")
    body_style.fontName = "VestaPDF"
    body_style.fontSize = 7
    body_style.leading = 8

    tiny_style = styles["BodyText"].clone("scba_tiny")
    tiny_style.fontName = "VestaPDF"
    tiny_style.fontSize = 5
    tiny_style.leading = 6

    tiny_bold_style = styles["BodyText"].clone("scba_tiny_bold")
    tiny_bold_style.fontName = "VestaPDFBold"
    tiny_bold_style.fontSize = 5
    tiny_bold_style.leading = 6

    compact_style = styles["BodyText"].clone("scba_compact")
    compact_style.fontName = "VestaPDF"
    compact_style.fontSize = 5.6
    compact_style.leading = 6.2
    compact_style.spaceBefore = 0
    compact_style.spaceAfter = 0

    compact_location_style = styles["BodyText"].clone("scba_compact_location")
    compact_location_style.fontName = "VestaPDF"
    compact_location_style.fontSize = 4.5
    compact_location_style.leading = 5.0
    compact_location_style.spaceBefore = 0
    compact_location_style.spaceAfter = 0

    compact_date_style = styles["BodyText"].clone("scba_compact_date")
    compact_date_style.fontName = "VestaPDF"
    compact_date_style.fontSize = 5.0
    compact_date_style.leading = 5.5
    compact_date_style.spaceBefore = 0
    compact_date_style.spaceAfter = 0
    compact_date_style.alignment = 1

    form_total_width = 210 * mm

    def scba_check_header_text(label: str) -> str:
        mapping = {
            "17.S.1001.A.1 Yüz Maskesi Kontrol Edildi": "Yüz<br/>Maskesi<br/>Kontrol<br/>Edildi",
            "17.S.1001.A.2 Solunum Valfi Kontrol Edildi": "Solunum<br/>Valfi<br/>Kontrol<br/>Edildi",
            "17.S.1001.A.3 Regülatör Ünitesi Kontrol Edildi": "Regülatör<br/>Ünitesi<br/>Kontrol<br/>Edildi",
            "17.S.1001.A.4 Kemer Kontrol Edildi": "Kemer<br/>Kontrol<br/>Edildi",
            "17.S.1001.A.5 Silindir Kontrol Edildi": "Silindir<br/>Kontrol<br/>Edildi",
            "17.S.1001.A.6 NOZUL UYGUNLUĞU (PASLANMA VB.)": "NOZUL<br/>UYGUNLUĞU<br/>(PASLANMA<br/>VB.)",
            "17.S.1001.A.7 Servis Etiketi Ekipmana Yapıştırıldı": "Servis<br/>Etiketi<br/>Ekipmana<br/>Yapıştırıldı",
        }
        return mapping.get(label, label)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=6 * mm,
        bottomMargin=6 * mm,
    )
    story = []

    logo_cells = []
    if TSE_HYB_LOGO_PATH.exists():
        logo_cells.append(PdfImage(str(TSE_HYB_LOGO_PATH), width=17 * mm, height=17 * mm))
    else:
        logo_cells.append(Paragraph("<b>TSE-HYB</b>", body_style))
    if VESTA_HEADER_LOGO_PATH.exists():
        logo_cells.append(PdfImage(str(VESTA_HEADER_LOGO_PATH), width=13 * mm, height=16 * mm))
    else:
        logo_cells.append(Paragraph("<b>VESTA</b>", body_style))

    logo_table = PdfTable([logo_cells], colWidths=[20 * mm, 20 * mm], rowHeights=[18 * mm])
    logo_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (-1, -1), "CENTER")]))

    header_table = PdfTable(
        [
            [
                logo_table,
                Paragraph("<b>Onay: Şirket Müdürü<br/>Mustafa KİLİÇ</b>", body_style),
                Paragraph("<b>Bölüm</b>", body_style),
                Paragraph("<b>Revizyon No</b>", body_style),
                Paragraph("<b>Revizyon Tarihi</b>", body_style),
                Paragraph("<b>Sayfa</b>", body_style),
            ],
            ["", "", "F-17", "0", "2026-03-01", "1"],
            [
                Paragraph("<b>FORM</b>", body_style),
                Paragraph("<b>Konu: BAĞIMSIZ SOLUNUM CİHAZI (SELF-CONTAINED BREATHING APPARATUS) Kontrol Formu</b>", body_style),
                Paragraph("<b>Hazırlayan: Kalite Temsilcisi</b>", body_style),
                "",
                "",
                "",
            ],
        ],
        colWidths=[40 * mm, 76 * mm, 16 * mm, 24 * mm, 34 * mm, 20 * mm],
        hAlign="CENTER",
    )
    header_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("SPAN", (0, 0), (0, 1)),
                ("SPAN", (1, 0), (1, 1)),
                ("SPAN", (1, 2), (2, 2)),
                ("SPAN", (3, 2), (5, 2)),
                ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
                ("FONTNAME", (0, 0), (-1, 0), "VestaPDFBold"),
                ("FONTNAME", (0, 2), (-1, 2), "VestaPDFBold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (2, 0), (-1, 1), "CENTER"),
            ]
        )
    )
    story.extend([header_table, Spacer(1, 1 * mm)])

    info_table = PdfTable(
        [
            [Paragraph("<b>FİRMA ADI</b>", body_style), document_data["company_name"], "", "", Paragraph("<b>KONTROL TARİHİ</b>", body_style), document_data["control_date"]],
            [Paragraph("<b>MUAYENE ADRESİ</b>", body_style), document_data["company_address"], "", "", Paragraph("<b>FİRMA YETKİLİ KİŞİ</b>", body_style), document_data["company_contact"]],
            [Paragraph("<b>PERİYODİK KONTROL METODU</b>", body_style), Paragraph(document_data["method_text"], compact_style), "", "", "", ""],
        ],
        colWidths=[35 * mm, 75 * mm, 2 * mm, 2 * mm, 38 * mm, 58 * mm],
        hAlign="CENTER",
        rowHeights=[8 * mm, 8 * mm, 10 * mm],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("SPAN", (1, 0), (3, 0)),
                ("SPAN", (1, 1), (3, 1)),
                ("SPAN", (1, 2), (-1, 2)),
                ("FONTNAME", (0, 0), (-1, -1), "VestaPDF"),
                ("FONTNAME", (0, 0), (0, -1), "VestaPDFBold"),
                ("FONTNAME", (4, 0), (4, 1), "VestaPDFBold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.extend([info_table, Spacer(1, 1 * mm)])

    header_row_1 = [
        Paragraph("<b>BAĞIMSIZ SOLUNUM CİHAZI (SELF-CONTAINED BREATHING APPARATUS) BİLGİLERİ</b>", body_style),
        "", "", "", "", "", "", "",
        Paragraph("<b>TESPİT VE DEĞERLENDİRME</b><br/><font size='6'>(✔: UYGUN; X: U. DEĞİL; ─: UYGULAMA YOK)</font>", tiny_bold_style),
        "", "", "", "", "", "",
    ]
    header_row_2 = [
        Paragraph("<b>CIHAZ</b><br/><b>NO</b>", tiny_bold_style),
        Paragraph("<b>KATEGORİ / CİNSİ</b>", tiny_bold_style),
        "",
        Paragraph("<b>SERİ NO / KOD</b>", tiny_bold_style),
        Paragraph("<b>MARKA</b>", tiny_bold_style),
        Paragraph("<b>DOLUM TARİHİ</b>", tiny_bold_style),
        Paragraph("<b>HİDROSTATİK TEST TARİHİ</b>", tiny_bold_style),
        Paragraph("<b>BULUNDUĞU YER</b>", tiny_bold_style),
        *[Paragraph(f"<b>{scba_check_header_text(label)}</b>", tiny_style) for _key, label in SCBA_CONTROL_ITEMS],
    ]
    data_rows = [header_row_1, header_row_2]
    for row in document_data["rows"]:
        data_rows.append(
            [
                row["device_no"],
                Paragraph(str(row["category"]), compact_style),
                "",
                Paragraph(str(row["serial_number"]), compact_style),
                Paragraph(str(row["manufacturer"]), compact_style),
                Paragraph(str(row["service_date"]).replace("-", "&#8209;"), compact_date_style),
                Paragraph(str(row["hydrostatic_test_date"]).replace("-", "&#8209;"), compact_date_style),
                Paragraph(str(row["location_detail"]), compact_location_style),
                *row["checks"],
            ]
        )
    target_rows = 10 if len(document_data["rows"]) <= 10 else len(document_data["rows"])
    while len(data_rows) < (target_rows + 2):
        data_rows.append([""] * 15)

    main_table = PdfTable(
        data_rows,
        repeatRows=2,
        colWidths=[12 * mm, 6 * mm, 19 * mm, 19 * mm, 17 * mm, 16 * mm, 19 * mm, 22 * mm, 10 * mm, 10 * mm, 10 * mm, 10 * mm, 10 * mm, 15 * mm, 15 * mm],
        rowHeights=[7 * mm, 18 * mm] + [6 * mm] * (len(data_rows) - 2),
        hAlign="CENTER",
    )
    main_table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (7, 0)),
                ("SPAN", (8, 0), (14, 0)),
                ("SPAN", (1, 1), (2, 1)),
                ("SPAN", (1, 2), (2, -1)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f7f7f7")),
                ("FONTNAME", (0, 0), (-1, 1), "VestaPDFBold"),
                ("FONTNAME", (0, 2), (-1, -1), "VestaPDF"),
                ("FONTSIZE", (0, 0), (-1, -1), 5),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.extend([main_table, Spacer(1, 1 * mm)])

    note_style = styles["BodyText"].clone("scba_note")
    note_style.fontName = "VestaPDF"
    note_style.fontSize = 5
    note_style.leading = 6
    notes_table = PdfTable([[Paragraph(note, note_style)] for note in document_data["notes"]], colWidths=[form_total_width], hAlign="CENTER")
    notes_table.setStyle(TableStyle([("LEFTPADDING", (0,0), (-1,-1), 0), ("RIGHTPADDING", (0,0), (-1,-1), 0), ("TOPPADDING", (0,0), (-1,-1), 0), ("BOTTOMPADDING", (0,0), (-1,-1), 0)]))
    story.append(notes_table)

    add_second_page = len(document_data["rows"]) > 10
    if add_second_page:
        story.append(PageBreak())
        story.append(header_table)
        story.append(Spacer(1, 2 * mm))
    else:
        story.append(Spacer(1, 4 * mm))

    story.append(PdfTable([[Paragraph("<b>AÇIKLAMALAR</b>", body_style)]], colWidths=[form_total_width], hAlign="CENTER", rowHeights=[7 * mm]))
    story.append(PdfTable([[""]], colWidths=[form_total_width], hAlign="CENTER", rowHeights=[28 * mm], style=TableStyle([("GRID",(0,0),(-1,-1),0.6,colors.black)])))
    signature = PdfTable(
        [
            [Paragraph("<b>PERIYODIK KONTROL PERSONELI</b>", body_style), "", Paragraph("<b>FIRMA YETKILISI</b>", body_style), ""],
            [Paragraph(document_data.get("periodic_control_person") or "-", body_style), "", Paragraph(document_data.get("company_contact") or "-", body_style), ""],
        ],
        colWidths=[52.5 * mm, 52.5 * mm, 52.5 * mm, 52.5 * mm],
        rowHeights=[7 * mm, 10 * mm],
        hAlign="CENTER",
    )
    signature.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.6,colors.black), ("SPAN",(0,0),(1,0)), ("SPAN",(2,0),(3,0)), ("SPAN",(0,1),(1,1)), ("SPAN",(2,1),(3,1)), ("ALIGN",(0,0),(-1,-1),"CENTER"), ("VALIGN",(0,0),(-1,-1),"MIDDLE"), ("FONTNAME",(0,0),(-1,0),"VestaPDFBold"), ("FONTNAME",(0,1),(-1,1),"VestaPDF")]))
    story.append(signature)
    story.append(PdfTable([["ONAY"], [document_data.get("approval_name") or "Mustafa Kilic"]], colWidths=[form_total_width], hAlign="CENTER", rowHeights=[8 * mm, 18 * mm], style=TableStyle([("GRID",(0,0),(-1,-1),0.6,colors.black), ("ALIGN",(0,0),(-1,-1),"LEFT"), ("VALIGN",(0,0),(-1,-1),"TOP"), ("FONTNAME",(0,0),(0,0),"VestaPDFBold"), ("FONTNAME",(0,1),(0,1),"VestaPDF"), ("LEFTPADDING",(0,0),(-1,-1),4), ("TOPPADDING",(0,0),(-1,-1),4)])))
    story.append(Spacer(1, 2 * mm))
    if add_second_page:
        notes2 = PdfTable([[Paragraph(note, note_style)] for note in document_data["notes"]], colWidths=[form_total_width], hAlign="CENTER")
        notes2.setStyle(TableStyle([("LEFTPADDING", (0,0), (-1,-1), 0), ("RIGHTPADDING", (0,0), (-1,-1), 0), ("TOPPADDING", (0,0), (-1,-1), 0), ("BOTTOMPADDING", (0,0), (-1,-1), 0)]))
        story.append(notes2)

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_control_form_excel_pdf(
    company_name: str,
    extinguishers_for_company: list[dict],
    latest_inspections: dict[int, dict],
) -> io.BytesIO | None:
    if os.name != "nt" or not CONTROL_FORM_EXCEL_TEMPLATE_PATH.exists():
        return None

    try:
        import pythoncom
        import win32com.client  # type: ignore[import-not-found]
    except ImportError:
        return None

    workbook = load_workbook(CONTROL_FORM_EXCEL_TEMPLATE_PATH)
    worksheet = workbook["KONTROL FORMU"]
    first = extinguishers_for_company[0] if extinguishers_for_company else {}

    worksheet["C4"] = company_name
    worksheet["N4"] = datetime.now().strftime("%d.%m.%Y")
    worksheet["C5"] = first.get("company_address") or "-"
    worksheet["N5"] = first.get("company_contact") or "-"

    start_row = 10
    max_rows = 14
    for offset in range(max_rows):
        row_no = start_row + offset
        extinguisher = extinguishers_for_company[offset] if offset < len(extinguishers_for_company) else None
        if extinguisher is None:
            for column in "ABCDEFGHIJKLMNO":
                worksheet[f"{column}{row_no}"] = None
            continue

        inspection = latest_inspections.get(extinguisher["id"], {})
        worksheet[f"A{row_no}"] = offset + 1
        worksheet[f"B{row_no}"] = pdf_equipment_label(extinguisher.get("extinguisher_type"))
        worksheet[f"C{row_no}"] = extinguisher.get("fire_class") or "-"
        worksheet[f"D{row_no}"] = extinguisher.get("serial_number") or "-"
        worksheet[f"E{row_no}"] = extinguisher.get("manufacturer") or "-"
        worksheet[f"F{row_no}"] = extinguisher.get("last_service_date") or "-"
        worksheet[f"G{row_no}"] = extinguisher.get("hydrostatic_test_date") or "-"
        worksheet[f"H{row_no}"] = extinguisher.get("location_detail") or "-"
        for idx, column in enumerate("IJKLMNO"):
            key = CONTROL_FORM_ITEMS[idx][0]
            if inspection:
                worksheet[f"{column}{row_no}"] = "✔" if inspection.get(key) else "X"
            else:
                worksheet[f"{column}{row_no}"] = "-"

    temp_dir = Path(tempfile.mkdtemp(prefix="vesta-control-form-"))
    xlsx_path = temp_dir / "control-form.xlsx"
    pdf_path = temp_dir / "control-form.pdf"
    workbook.save(xlsx_path)
    workbook.close()

    pythoncom.CoInitialize()
    excel = None
    excel_workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel_workbook = excel.Workbooks.Open(str(xlsx_path))
        worksheet_com = excel_workbook.Worksheets("KONTROL FORMU")
        worksheet_com.ExportAsFixedFormat(0, str(pdf_path))
        excel_workbook.Close(False)
        excel.Quit()
        pdf_bytes = pdf_path.read_bytes()
    finally:
        if excel_workbook is not None:
            try:
                excel_workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    buffer = io.BytesIO(pdf_bytes)
    buffer.seek(0)
    return buffer


def build_pdf_from_html(html: str) -> io.BytesIO:
    buffer = io.BytesIO()
    result = pisa.CreatePDF(src=html, dest=buffer, encoding="utf-8")
    if result.err:
        raise RuntimeError("HTML PDF olusturma basarisiz oldu.")
    buffer.seek(0)
    return buffer


def build_electrical_report_document_data(public_id: str) -> dict:
    extinguisher = get_extinguisher(public_id)
    latest_log = fetch_one(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
    )
    notes = parse_structured_notes(extinguisher.get("notes"))
    summary = parse_electrical_operation_summary(latest_log.get("operation_summary") if latest_log else "")
    return {
        "extinguisher": extinguisher,
        "notes": notes,
        "summary": summary,
        "latest_log": latest_log,
    }


def build_electrical_report_pdf(public_id: str) -> io.BytesIO:
    if not ELECTRICAL_REPORT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Elektrik raporu sablonu bulunamadi: {ELECTRICAL_REPORT_TEMPLATE_PATH}")

    document_data = build_electrical_report_document_data(public_id)
    extinguisher = document_data["extinguisher"]
    notes = document_data["notes"]
    summary = document_data["summary"]

    def draw_box(page, rect, value, *, size=7, align=0, fontname="helv"):
        text_value = str(value or "-").strip() or "-"
        page.insert_textbox(
            fitz.Rect(*rect),
            text_value,
            fontname=fontname,
            fontsize=size,
            color=(0, 0, 0),
            align=align,
        )

    def mark_checkbox(page, x, y):
        page.insert_text((x, y), "X", fontname="helv", fontsize=8, color=(0, 0, 0))

    def yes_no_positions(yes_xy, no_xy, value):
        normalized = (value or "").strip().lower()
        if normalized in {"var", "evet", "uygun"}:
            return yes_xy
        if normalized in {"yok", "hayir", "uygun degil"}:
            return no_xy
        return None

    template = fitz.open(ELECTRICAL_REPORT_TEMPLATE_PATH)

    # Page 1
    page = template[0]
    draw_box(page, (92, 122, 305, 138), extinguisher.get("company_name"), size=7)
    draw_box(page, (499, 122, 586, 138), extinguisher.get("serial_number"), size=7, align=1)
    draw_box(page, (92, 140, 305, 193), extinguisher.get("company_address"), size=7)
    draw_box(page, (499, 140, 586, 156), extinguisher.get("last_service_date"), size=7, align=1)
    draw_box(page, (499, 158, 586, 174), notes.get("ISG-KATIP Sozlesme ID"), size=7, align=1)
    draw_box(page, (499, 176, 586, 192), summary.get("control_start"), size=6, align=1)
    draw_box(page, (499, 194, 586, 210), summary.get("control_end"), size=6, align=1)
    draw_box(page, (92, 194, 305, 210), notes.get("SGK Sicil Numarasi"), size=7)
    draw_box(page, (499, 212, 586, 228), extinguisher.get("next_service_date"), size=7, align=1)
    draw_box(page, (92, 212, 586, 268), notes.get("Periyodik Kontrol Metodu ve Kapsami"), size=6)

    draw_box(page, (124, 281, 271, 296), extinguisher.get("manufacturer"), size=7)
    draw_box(page, (123, 306, 197, 321), extinguisher.get("fire_class"), size=7)
    draw_box(page, (123, 331, 197, 346), notes.get("Yapi cinsi"), size=7)
    draw_box(page, (256, 306, 399, 321), notes.get("Tesise ait proje var mi"), size=7)
    draw_box(page, (451, 306, 585, 321), notes.get("Tek hat semasi var mi"), size=7)
    draw_box(page, (123, 331, 197, 370), notes.get("Kontrol nedeni"), size=7)
    draw_box(page, (256, 331, 399, 370), notes.get("Topraklayici tipi"), size=7)
    draw_box(page, (256, 370, 399, 395), summary.get("equipment_usage_purpose"), size=7)
    draw_box(page, (451, 370, 585, 395), notes.get("Son kontrol tarihi"), size=7)
    draw_box(page, (123, 396, 197, 460), notes.get("Faz iletkenlerinin sayisi ve tipi"), size=6)
    draw_box(page, (330, 396, 399, 412), notes.get("Temel topraklama direnci"), size=7)
    draw_box(page, (330, 412, 585, 428), notes.get("Ilave topraklama elektrotu detaylari"), size=6)
    draw_box(page, (330, 428, 585, 444), notes.get("Sistem topraklama iletkeni ve kesiti"), size=6)
    draw_box(page, (330, 444, 585, 460), notes.get("Ana espotansiyel iletkeni ve kesiti"), size=6)
    draw_box(page, (123, 462, 399, 515), notes.get("Besleme kaynagi karakteristikleri"), size=6)
    draw_box(page, (504, 462, 585, 489), notes.get("Ana RCD anma akimi"), size=7)
    draw_box(page, (123, 516, 399, 541), notes.get("Ana kesici karakteristikleri"), size=6)
    draw_box(page, (504, 516, 585, 541), notes.get("Ana RCD test akimi ve suresi"), size=6)

    checkbox = yes_no_positions((274, 557), (314, 557), notes.get("Tesisatta kapsamli degisiklik var mi (>%20)"))
    if checkbox:
        mark_checkbox(page, *checkbox)
    checkbox = yes_no_positions((274, 573), (314, 573), notes.get("Asiri gerilim koruma cihazlari kullanilmis mi"))
    if checkbox:
        mark_checkbox(page, *checkbox)
    checkbox = yes_no_positions((274, 650), (314, 650), notes.get("Bir onceki periyodik kontrol etiketi var mi"))
    if checkbox:
        mark_checkbox(page, *checkbox)
    draw_box(page, (272, 589, 585, 645), notes.get("Dogrudan dokunmaya karsi koruma onlemleri"), size=5.5)
    draw_box(page, (120, 589, 272, 645), notes.get("Tespit edilen bilgiler"), size=5.5)
    draw_box(page, (120, 683, 271, 755), notes.get("Termal Kamera 1"), size=6)
    draw_box(page, (272, 683, 585, 755), notes.get("Termal Kamera 2"), size=6)

    # Page 2
    page = template[1]
    draw_box(page, (120, 112, 271, 184), notes.get("Olcum Aleti 1"), size=6)
    draw_box(page, (272, 112, 585, 184), notes.get("Olcum Aleti 2"), size=6)
    draw_box(page, (182, 216, 585, 474), notes.get("Kontrol Kriterleri ve Testler"), size=6)
    draw_box(page, (119, 474, 271, 528), notes.get("Termal fotograf tarihi"), size=6)
    draw_box(page, (424, 474, 585, 528), notes.get("Termal fotograf no"), size=6)
    draw_box(page, (271, 474, 424, 501), notes.get("Kontak gevsakligi isinmasi"), size=6)
    draw_box(page, (271, 501, 424, 528), notes.get("Asiri yuk isinmasi"), size=6)
    draw_box(page, (122, 576, 585, 593), notes.get("Olcum ve Dogrulama Metodu"), size=7)
    draw_box(page, (119, 631, 585, 701), notes.get("6.1 Notlari"), size=6)

    # Page 3
    page = template[2]
    draw_box(page, (28, 294, 569, 408), notes.get("6.2 Notlari"), size=6)
    draw_box(page, (28, 438, 569, 493), notes.get("6.3 Notlari"), size=6)
    draw_box(page, (28, 523, 569, 598), notes.get("Kusur Aciklamalari"), size=6)
    draw_box(page, (28, 621, 569, 736), notes.get("Ekipman Fotograflari"), size=6)
    draw_box(page, (28, 758, 569, 808), notes.get("Genel Notlar"), size=6)

    # Page 4
    page = template[3]
    draw_box(page, (26, 104, 569, 329), notes.get("Sonuc ve Kanaat"), size=6)
    draw_box(page, (120, 783, 455, 797), notes.get("Yetkili Kisi"), size=7)
    draw_box(page, (120, 797, 455, 811), notes.get("Yetkili Kisi"), size=7)
    draw_box(page, (120, 811, 455, 825), notes.get("Yetkili Kisi"), size=7)
    draw_box(page, (25, 823, 270, 836), f"Bu rapor {notes.get('Nusha Sayisi') or '2'} (yazi/rakam) nusha olarak hazirlanmistir.", size=5)

    output = io.BytesIO(template.tobytes(garbage=4, deflate=True))
    output.seek(0)
    template.close()
    return output


def pdf_page_to_data_uri(pdf_path: Path, page_number: int) -> str:
    document = fitz.open(pdf_path)
    try:
        page = document[page_number]
        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        encoded = base64.b64encode(pixmap.tobytes("png")).decode("ascii")
        return f"data:image/png;base64,{encoded}"
    finally:
        document.close()


def build_electrical_report_html_context(public_id: str) -> dict:
    document_data = build_electrical_report_document_data(public_id)
    extinguisher = document_data["extinguisher"]
    notes = document_data["notes"]
    summary = document_data["summary"]

    backgrounds = [
        pdf_page_to_data_uri(ELECTRICAL_REPORT_TEMPLATE_PATH, idx)
        for idx in range(4)
    ]

    def box(left, top, width, height, value, size=12, align="left", weight="normal"):
        return {
            "left": left * 2,
            "top": top * 2,
            "width": width * 2,
            "height": height * 2,
            "value": str(value or "-").strip() or "-",
            "size": size,
            "align": align,
            "weight": weight,
        }

    page_boxes = {
        0: [
            box(92, 122, 213, 16, extinguisher.get("company_name")),
            box(499, 122, 87, 16, extinguisher.get("serial_number"), align="center"),
            box(92, 140, 213, 53, extinguisher.get("company_address")),
            box(499, 140, 87, 16, extinguisher.get("last_service_date"), align="center"),
            box(499, 158, 87, 16, notes.get("ISG-KATIP Sozlesme ID"), align="center"),
            box(499, 176, 87, 16, summary.get("control_start"), size=10, align="center"),
            box(499, 194, 87, 16, summary.get("control_end"), size=10, align="center"),
            box(92, 194, 213, 16, notes.get("SGK Sicil Numarasi")),
            box(499, 212, 87, 16, extinguisher.get("next_service_date"), align="center"),
            box(92, 212, 494, 56, notes.get("Periyodik Kontrol Metodu ve Kapsami"), size=10),
            box(124, 281, 147, 15, extinguisher.get("manufacturer")),
            box(123, 306, 74, 15, extinguisher.get("fire_class")),
            box(256, 306, 143, 15, notes.get("Tesise ait proje var mi")),
            box(451, 306, 134, 15, notes.get("Tek hat semasi var mi")),
            box(123, 331, 74, 39, notes.get("Kontrol nedeni")),
            box(256, 331, 143, 39, notes.get("Topraklayici tipi")),
            box(123, 370, 74, 26, notes.get("Yapi cinsi")),
            box(256, 370, 143, 25, summary.get("equipment_usage_purpose")),
            box(451, 370, 134, 25, notes.get("Son kontrol tarihi")),
            box(123, 396, 74, 64, notes.get("Faz iletkenlerinin sayisi ve tipi"), size=10),
            box(330, 396, 69, 16, notes.get("Temel topraklama direnci")),
            box(330, 412, 255, 16, notes.get("Ilave topraklama elektrotu detaylari"), size=10),
            box(330, 428, 255, 16, notes.get("Sistem topraklama iletkeni ve kesiti"), size=10),
            box(330, 444, 255, 16, notes.get("Ana espotansiyel iletkeni ve kesiti"), size=10),
            box(123, 462, 276, 53, notes.get("Besleme kaynagi karakteristikleri"), size=10),
            box(504, 462, 81, 27, notes.get("Ana RCD anma akimi"), align="center"),
            box(123, 516, 276, 25, notes.get("Ana kesici karakteristikleri"), size=10),
            box(504, 516, 81, 25, notes.get("Ana RCD test akimi ve suresi"), size=10, align="center"),
            box(121, 554, 147, 16, notes.get("Tesisatta kapsamli degisiklik var mi (>%20)")),
            box(121, 570, 147, 18, notes.get("Asiri gerilim koruma cihazlari kullanilmis mi"), size=10),
            box(271, 589, 314, 56, notes.get("Dogrudan dokunmaya karsi koruma onlemleri"), size=9),
            box(121, 589, 150, 56, notes.get("Tespit edilen bilgiler"), size=9),
            box(121, 647, 147, 16, notes.get("Bir onceki periyodik kontrol etiketi var mi")),
            box(120, 683, 151, 72, notes.get("Termal Kamera 1"), size=10),
            box(272, 683, 313, 72, notes.get("Termal Kamera 2"), size=10),
        ],
        1: [
            box(120, 112, 151, 72, notes.get("Olcum Aleti 1"), size=10),
            box(272, 112, 313, 72, notes.get("Olcum Aleti 2"), size=10),
            box(182, 216, 403, 258, notes.get("Kontrol Kriterleri ve Testler"), size=9),
            box(119, 474, 152, 54, notes.get("Termal fotograf tarihi"), size=10),
            box(271, 474, 153, 27, notes.get("Kontak gevsakligi isinmasi"), size=10),
            box(271, 501, 153, 27, notes.get("Asiri yuk isinmasi"), size=10),
            box(424, 474, 161, 54, notes.get("Termal fotograf no"), size=10),
            box(122, 576, 463, 17, notes.get("Olcum ve Dogrulama Metodu")),
            box(119, 631, 466, 70, notes.get("6.1 Notlari"), size=9),
        ],
        2: [
            box(28, 294, 541, 114, notes.get("6.2 Notlari"), size=9),
            box(28, 438, 541, 55, notes.get("6.3 Notlari"), size=9),
            box(28, 523, 541, 75, notes.get("Kusur Aciklamalari"), size=9),
            box(28, 621, 541, 115, notes.get("Ekipman Fotograflari"), size=9),
            box(28, 758, 541, 50, notes.get("Genel Notlar"), size=9),
        ],
        3: [
            box(26, 104, 543, 225, notes.get("Sonuc ve Kanaat"), size=9),
            box(120, 783, 335, 14, notes.get("Yetkili Kisi"), size=10),
            box(120, 797, 335, 14, notes.get("Yetkili Kisi"), size=10),
            box(120, 811, 335, 14, notes.get("Yetkili Kisi"), size=10),
            box(25, 823, 245, 13, f"Bu rapor {notes.get('Nusha Sayisi') or '2'} (yazi/rakam) nusha olarak hazirlanmistir.", size=8),
        ],
    }

    return {
        "extinguisher": extinguisher,
        "page_backgrounds": backgrounds,
        "page_boxes": page_boxes,
    }


def build_electrical_report_pdf_html(public_id: str) -> io.BytesIO:
    context = build_electrical_report_html_context(public_id)
    html = render_template("electrical_report_pdf.html", **context)
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                executable_path=playwright.chromium.executable_path,
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-setuid-sandbox",
                    "--disable-dev-shm-usage",
                ],
            )
            page = browser.new_page(viewport={"width": 1191, "height": 1684})
            page.set_content(html, wait_until="load")
            pdf_bytes = page.pdf(
                format="A4",
                print_background=True,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
            )
            browser.close()
        buffer = io.BytesIO(pdf_bytes)
        buffer.seek(0)
        return buffer
    except Exception:
        return build_pdf_from_html(html)


@app.route("/extinguishers/<public_id>/electrical-report")
@login_required
def electrical_report_preview(public_id: str):
    extinguisher = get_extinguisher(public_id)
    if extinguisher.get("asset_category") != "Elektrik Ic Tesisati":
        abort(404)
    context = build_electrical_report_html_context(public_id)
    return render_template("electrical_report_pdf.html", **context)


@app.route("/public/<public_id>/electrical-report")
def public_electrical_report_preview(public_id: str):
    extinguisher = get_extinguisher(public_id)
    if extinguisher.get("asset_category") != "Elektrik Ic Tesisati":
        abort(404)
    context = build_electrical_report_html_context(public_id)
    return render_template("electrical_report_pdf.html", **context)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = fetch_one(select(users).where(users.c.username == username))
        if user and bool(user["is_active"]) and check_password_hash(user["password_hash"], password):
            session["authenticated"] = True
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["full_name"] = user["full_name"]
            session["is_admin"] = bool(user["is_admin"])
            flash("Giris yapildi.", "success")
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        flash("Kullanici adi veya sifre hatali.", "error")
    return render_template("login.html")


@app.route("/health")
def health():
    due_assets = process_due_soon_notifications()
    return {"status": "ok", "due_soon": len(due_assets)}, 200


@app.route("/logout")
def logout():
    session.clear()
    flash("Cikis yapildi.", "success")
    return redirect(url_for("login"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    current_user = fetch_one(select(users).where(users.c.id == session.get("user_id")))
    if current_user is None:
        session.clear()
        return redirect(url_for("login"))

    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not check_password_hash(current_user["password_hash"], current_password):
            flash("Mevcut sifre hatali.", "error")
            return redirect(url_for("profile"))
        if not new_password:
            flash("Yeni sifre bos olamaz.", "error")
            return redirect(url_for("profile"))
        if new_password != confirm_password:
            flash("Yeni sifreler eslesmiyor.", "error")
            return redirect(url_for("profile"))

        now = datetime.now().isoformat(timespec="seconds")
        with engine.begin() as connection:
            connection.execute(
                update(users)
                .where(users.c.id == current_user["id"])
                .values(
                    password_hash=generate_password_hash(new_password),
                    updated_at=now,
                )
            )
        flash("Sifren guncellendi.", "success")
        return redirect(url_for("profile"))

    return render_template("profile.html", current_user=current_user)


@app.route("/users", methods=["GET"])
@admin_required
def user_management():
    user_rows = fetch_all(
        select(
            users.c.id,
            users.c.username,
            users.c.full_name,
            users.c.is_admin,
            users.c.is_active,
            users.c.created_at,
        ).order_by(users.c.full_name)
    )
    return render_template("user_management.html", users=user_rows)


@app.route("/users/create", methods=["POST"])
@admin_required
def create_user():
    username = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    password = request.form.get("password", "")
    is_admin = request.form.get("is_admin") == "on"

    if not username or not full_name or not password:
        flash("Kullanici adi, ad soyad ve sifre gerekli.", "error")
        return redirect(url_for("user_management"))

    now = datetime.now().isoformat(timespec="seconds")
    try:
        with engine.begin() as connection:
            connection.execute(
                insert(users).values(
                    username=username,
                    full_name=full_name,
                    password_hash=generate_password_hash(password),
                    is_admin=is_admin,
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                )
            )
    except IntegrityError:
        flash("Bu kullanici adi zaten mevcut.", "error")
        return redirect(url_for("user_management"))

    flash("Kullanici olusturuldu.", "success")
    return redirect(url_for("user_management"))


@app.route("/users/<int:user_id>/password", methods=["POST"])
@admin_required
def update_user_password(user_id: int):
    password = request.form.get("password", "")
    if not password:
        flash("Yeni sifre bos olamaz.", "error")
        return redirect(url_for("user_management"))

    now = datetime.now().isoformat(timespec="seconds")
    with engine.begin() as connection:
        connection.execute(
            update(users)
            .where(users.c.id == user_id)
            .values(
                password_hash=generate_password_hash(password),
                updated_at=now,
            )
        )
    flash("Sifre guncellendi.", "success")
    return redirect(url_for("user_management"))


@app.route("/users/<int:user_id>/toggle-active", methods=["POST"])
@admin_required
def toggle_user_active(user_id: int):
    target_user = fetch_one(select(users).where(users.c.id == user_id))
    if not target_user:
        abort(404)
    if target_user["username"] == session.get("username"):
        flash("Kendi hesabini pasife alamazsin.", "error")
        return redirect(url_for("user_management"))

    now = datetime.now().isoformat(timespec="seconds")
    with engine.begin() as connection:
        connection.execute(
            update(users)
            .where(users.c.id == user_id)
            .values(
                is_active=not bool(target_user["is_active"]),
                updated_at=now,
            )
        )
    flash("Kullanici durumu guncellendi.", "success")
    return redirect(url_for("user_management"))


@app.route("/companies", methods=["GET"])
@login_required
def company_management():
    company_rows = get_company_choices()
    selected_company_id = request.args.get("company_id", type=int)
    selected_company = None
    brand_options = []
    location_options = []
    last_date_options = []
    next_date_options = []
    if company_rows:
        selected_company = next((row for row in company_rows if row["id"] == selected_company_id), company_rows[0])
    company_assets = []
    if selected_company:
        company_assets = fetch_all(
            select(extinguishers)
            .where(extinguishers.c.company_id == selected_company["id"])
            .order_by(extinguishers.c.asset_category, extinguishers.c.location_detail, extinguishers.c.serial_number)
        )
        brand_options = sorted(
            {
                (asset["manufacturer"] or "").strip()
                for asset in company_assets
                if (asset["manufacturer"] or "").strip()
            }
        )
        location_options = sorted(
            {
                (asset["location_detail"] or "").strip()
                for asset in company_assets
                if (asset["location_detail"] or "").strip()
            }
        )
        last_date_options = sorted(
            {
                str(asset["last_service_date"])
                for asset in company_assets
                if asset["last_service_date"]
            }
        )
        next_date_options = sorted(
            {
                str(asset["next_service_date"])
                for asset in company_assets
                if asset["next_service_date"]
            }
        )
    return render_template(
        "company_management.html",
        companies=company_rows,
        selected_company=selected_company,
        company_assets=company_assets,
        asset_categories=get_asset_category_choices(),
        brand_options=brand_options,
        location_options=location_options,
        last_date_options=last_date_options,
        next_date_options=next_date_options,
    )


@app.route("/companies/new", methods=["GET"])
@admin_required
def new_company():
    return render_template("company_create.html")


@app.route("/companies/create", methods=["POST"])
@admin_required
def create_company():
    name = request.form.get("name", "").strip()
    address = request.form.get("address", "").strip()
    contact_name = request.form.get("contact_name", "").strip() or "-"
    phone = request.form.get("phone", "").strip() or "-"
    email = request.form.get("email", "").strip() or "-"
    raw_slug = request.form.get("slug", "").strip()
    if not name or not address:
        flash("Firma adi ve adres gerekli.", "error")
        return redirect(url_for("new_company"))

    now = datetime.now().isoformat(timespec="seconds")
    try:
        with engine.begin() as connection:
            result = connection.execute(
                insert(companies).values(
                    public_id=uuid.uuid4().hex[:12],
                    slug=resolve_company_slug(connection, raw_slug, name),
                    name=name,
                    address=address,
                    contact_name=contact_name,
                    phone=phone,
                    email=email,
                    created_at=now,
                    updated_at=now,
                )
            )
    except IntegrityError:
        flash("Bu firma zaten mevcut.", "error")
        return redirect(url_for("new_company"))

    flash("Musteri kaydi olusturuldu.", "success")
    company_id = result.inserted_primary_key[0] if result.inserted_primary_key else None
    return redirect(url_for("company_management", company_id=company_id) if company_id else url_for("company_management"))


@app.route("/companies/<int:company_id>/update", methods=["POST"])
@admin_required
def update_company(company_id: int):
    company = get_company(company_id)
    name = request.form.get("name", "").strip()
    address = request.form.get("address", "").strip()
    contact_name = request.form.get("contact_name", "").strip() or "-"
    phone = request.form.get("phone", "").strip() or "-"
    email = request.form.get("email", "").strip() or "-"
    raw_slug = request.form.get("slug", "").strip()
    if not name or not address:
        flash("Firma adi ve adres gerekli.", "error")
        return redirect(url_for("company_management"))

    now = datetime.now().isoformat(timespec="seconds")
    try:
        with engine.begin() as connection:
            slug = resolve_company_slug(connection, raw_slug, name, exclude_id=company_id)
            connection.execute(
                update(companies)
                .where(companies.c.id == company_id)
                .values(
                    public_id=company.get("public_id") or uuid.uuid4().hex[:12],
                    slug=slug,
                    name=name,
                    address=address,
                    contact_name=contact_name,
                    phone=phone,
                    email=email,
                    updated_at=now,
                )
            )
            connection.execute(
                update(extinguishers)
                .where(extinguishers.c.company_id == company_id)
                .values(
                    company_name=name,
                    company_address=address,
                    updated_at=now,
                )
            )
    except IntegrityError:
        flash("Bu firma adi baska bir caride kullaniliyor.", "error")
        return redirect(url_for("company_management"))

    flash(f"{company['name']} guncellendi.", "success")
    return redirect(url_for("company_management", company_id=company_id))


@app.route("/")
@login_required
def index():
    due_soon_assets = get_due_soon_assets()
    latest_log_date = (
        select(service_logs.c.service_date)
        .where(service_logs.c.extinguisher_id == extinguishers.c.id)
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
        .scalar_subquery()
    )
    statement = (
        select(extinguishers, latest_log_date.label("latest_log_date"))
        .order_by(desc(extinguishers.c.updated_at))
    )
    return render_template("index.html", extinguishers=fetch_all(statement), due_soon_assets=due_soon_assets)


@app.route("/export/extinguishers.xlsx")
@login_required
def export_extinguishers():
    rows = fetch_all(
        select(
            extinguishers.c.serial_number,
            extinguishers.c.company_name,
            extinguishers.c.location_detail,
            extinguishers.c.weight_kg,
            extinguishers.c.extinguisher_type,
            extinguishers.c.pressure_status,
            extinguishers.c.last_service_date,
            extinguishers.c.next_service_date,
            extinguishers.c.notes,
            extinguishers.c.created_at,
            extinguishers.c.updated_at,
        ).order_by(desc(extinguishers.c.updated_at))
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tupler"
    headers = [
        "Seri No",
        "Firma",
        "Konum",
        "Kg",
        "Tip",
        "Basinc",
        "Son Bakim",
        "Sonraki Bakim",
        "Notlar",
        "Kayit Tarihi",
        "Guncelleme Tarihi",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row.values()))

    autosize_worksheet(sheet)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="yangin-tupleri.xlsx",
    )


@app.route("/export/service-logs.xlsx")
@login_required
def export_service_logs():
    statement = (
        select(
            extinguishers.c.serial_number,
            extinguishers.c.company_name,
            extinguishers.c.location_detail,
            service_logs.c.service_date,
            service_logs.c.technician_name,
            service_logs.c.operation_summary,
            service_logs.c.pressure_status,
            service_logs.c.notes,
            service_logs.c.created_at,
        )
        .join(service_logs, service_logs.c.extinguisher_id == extinguishers.c.id)
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
    )
    rows = fetch_all(statement)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bakim Gecmisi"
    headers = [
        "Seri No",
        "Firma",
        "Konum",
        "Bakim Tarihi",
        "Teknisyen",
        "Yapilan Islem",
        "Basinc",
        "Notlar",
        "Kayit Zamani",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row.values()))

    autosize_worksheet(sheet)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bakim-gecmisi.xlsx",
    )


@app.route("/extinguishers/new", methods=["GET", "POST"])
@login_required
def create_extinguisher():
    company_choices = get_company_choices()
    asset_categories = get_asset_category_choices()
    if request.method == "POST":
        submit_action = request.form.get("submit_action", "save_print")
        form = parse_required_form(request.form)
        form["technician_name"] = form.get("technician_name") or current_user_full_name()
        try:
            form, selected_company = sync_company_payload_from_selection(form)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_extinguisher.html",
                form=form,
                monthly_control_items=MONTHLY_CONTROL_ITEMS,
                equipment_options=EQUIPMENT_OPTIONS,
                equipment_presets=EQUIPMENT_PRESETS,
                companies=company_choices,
                asset_categories=asset_categories,
            )
        required_fields = {
            "serial_number": "Seri numarasi",
            "company_id": "Cari secimi",
            "asset_category": "Urun grubu",
            "location_detail": "Firma ici konum",
            "weight_kg": "Kg bilgisi",
            "extinguisher_type": "Tup tipi",
            "fire_class": "YSC sinifi",
            "manufacturer": "YSC uretici",
            "last_service_date": "Son bakim tarihi",
            "hydrostatic_test_date": "Hidrostatik test tarihi",
            "next_service_date": "Sonraki bakim tarihi",
            "technician_name": "Teknisyen",
            "operation_summary": "Yapilan islem",
        }
        missing = [label for key, label in required_fields.items() if not form.get(key)]
        if missing:
            flash(f"Eksik alanlar: {', '.join(missing)}", "error")
            return render_template(
                "create_extinguisher.html",
                form=form,
                monthly_control_items=MONTHLY_CONTROL_ITEMS,
                equipment_options=EQUIPMENT_OPTIONS,
                equipment_presets=EQUIPMENT_PRESETS,
                companies=company_choices,
                asset_categories=asset_categories,
            )

        try:
            weight_kg = parse_float(form["weight_kg"], "Kg")
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_extinguisher.html",
                form=form,
                monthly_control_items=MONTHLY_CONTROL_ITEMS,
                equipment_options=EQUIPMENT_OPTIONS,
                equipment_presets=EQUIPMENT_PRESETS,
                companies=company_choices,
                asset_categories=asset_categories,
            )

        now = datetime.now().isoformat(timespec="seconds")
        public_id = uuid.uuid4().hex[:12]
        inspection_values = build_monthly_inspection_values(request.form)
        control_values = derive_ysc_control_form_values(
            inspection_values,
            build_control_form_values(request.form),
        )

        try:
            with engine.begin() as connection:
                result = connection.execute(
                    insert(extinguishers).values(
                        public_id=public_id,
                        serial_number=form["serial_number"],
                        company_id=selected_company["id"],
                        company_name=form["company_name"],
                        company_address=form["company_address"],
                        company_contact=form["company_contact"],
                        asset_category=form["asset_category"],
                        location_detail=form["location_detail"],
                        weight_kg=weight_kg,
                        extinguisher_type=form["extinguisher_type"],
                        fire_class=form["fire_class"],
                        manufacturer=form["manufacturer"],
                        hydrostatic_test_date=form["hydrostatic_test_date"],
                        pressure_status=form.get("pressure_status"),
                        notes=form.get("notes"),
                        last_service_date=form["last_service_date"],
                        next_service_date=form["next_service_date"],
                        created_at=now,
                        updated_at=now,
                    )
                )
                extinguisher_id = result.inserted_primary_key[0]
                connection.execute(
                    insert(service_logs).values(
                        extinguisher_id=extinguisher_id,
                        service_date=form["last_service_date"],
                        technician_name=form["technician_name"],
                        operation_summary=form["operation_summary"],
                        pressure_status=form.get("pressure_status"),
                        notes=form.get("notes"),
                        created_at=now,
                    )
                )
                save_monthly_inspection(
                    connection=connection,
                    extinguisher_id=extinguisher_id,
                    inspection_date=form["last_service_date"],
                    inspector_name=form["technician_name"],
                    notes=form.get("notes"),
                    inspection_values=inspection_values,
                    control_values=control_values,
                    created_at=now,
                )
        except IntegrityError:
            flash("Bu seri numarasi zaten kayitli.", "error")
            return render_template(
                "create_extinguisher.html",
                form=form,
                monthly_control_items=MONTHLY_CONTROL_ITEMS,
                equipment_options=EQUIPMENT_OPTIONS,
                equipment_presets=EQUIPMENT_PRESETS,
                companies=company_choices,
                asset_categories=asset_categories,
            )

        if submit_action == "save":
            flash("Tup kaydedildi.", "success")
            return redirect(url_for("extinguisher_detail", public_id=public_id))

        flash("Tup kaydedildi. Etiket yazdirmaya yonlendirildiniz.", "success")
        return redirect(url_for("extinguisher_label", public_id=public_id))

    return render_template(
        "create_extinguisher.html",
        form={
            "technician_name": current_user_full_name(),
            "asset_category": DEFAULT_ASSET_CATEGORY,
        },
        monthly_control_items=MONTHLY_CONTROL_ITEMS,
        equipment_options=EQUIPMENT_OPTIONS,
        equipment_presets=EQUIPMENT_PRESETS,
        companies=company_choices,
        asset_categories=asset_categories,
    )


@app.route("/records/new/yangin-elbisesi", methods=["GET", "POST"])
@login_required
def create_fire_suit():
    return render_profile_record_form("yangin-elbisesi")


@app.route("/records/new")
@login_required
def record_group_picker():
    return render_template(
        "record_group_picker.html",
        groups=get_registration_groups(),
    )


@app.route("/records/new/<group_slug>", methods=["GET", "POST"])
@login_required
def record_group_entry(group_slug: str):
    group = get_registration_group(group_slug)
    if group_slug == "elektrik-ic-tesisati":
        return create_electrical_installation()
    if group["status"] == "active":
        return render_profile_record_form(group_slug)
    return render_template(
        "record_group_placeholder.html",
        group=group,
    )


@app.route("/extinguishers/<public_id>/edit", methods=["GET", "POST"])
@login_required
def edit_extinguisher_record(public_id: str):
    extinguisher = get_extinguisher(public_id)
    category = extinguisher.get("asset_category") or DEFAULT_ASSET_CATEGORY
    asset_profile = get_asset_profile(category)
    company_choices = get_company_choices()
    latest_log = fetch_one(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
    )
    latest_inspection = fetch_one(
        select(monthly_inspections)
        .where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(monthly_inspections.c.inspection_date), desc(monthly_inspections.c.id))
        .limit(1)
    )

    if category == "Elektrik Ic Tesisati":
        group = get_registration_group("elektrik-ic-tesisati")
        if request.method == "POST":
            form = parse_required_form(request.form)
            form["technician_name"] = current_user_full_name()
            form["report_number"] = extinguisher["serial_number"]
            try:
                form, selected_company = sync_company_payload_from_selection(form)
            except ValueError as exc:
                flash(str(exc), "error")
                return render_template(
                    "create_electrical_installation.html",
                    form=form,
                    companies=company_choices,
                    group=group,
                    asset_profile=asset_profile,
                    form_title=f"{group['label']} Kaydini Duzenle",
                    form_description="Bu ekranda mevcut elektrik ic tesisati kaydini guncelleyebilirsin.",
                    submit_label="Kaydi Guncelle",
                )

            required_fields = {
                "company_id": "Cari secimi",
                "company_address": "Periyodik kontrol adresi",
                "report_date": "Rapor tarihi",
                "control_start": "Periyodik kontrol baslangic tarihi ve saati",
                "control_end": "Periyodik kontrol bitis tarihi ve saati",
                "next_service_date": "Bir sonraki periyodik kontrol tarihi",
                "energy_provider": "Enerji saglayan kurulus",
                "grid_type": "Sebeke tipi",
                "grid_voltage": "Sebeke gerilimi",
                "equipment_usage_purpose": "Ekipmanin kullanim amaci",
                "final_conclusion_status": "Sonuc ve kanaat",
            }
            missing = [label for key, label in required_fields.items() if not form.get(key)]
            if missing:
                flash(f"Eksik alanlar: {', '.join(missing)}", "error")
                return render_template(
                    "create_electrical_installation.html",
                    form=form,
                    companies=company_choices,
                    group=group,
                    asset_profile=asset_profile,
                    form_title=f"{group['label']} Kaydini Duzenle",
                    form_description="Bu ekranda mevcut elektrik ic tesisati kaydini guncelleyebilirsin.",
                    submit_label="Kaydi Guncelle",
                )

            structured_notes = build_electrical_structured_notes(form)
            note_lines = ["Elektrik Ic Tesisati Tam Form Baslangic Kaydi", *build_electrical_note_lines(structured_notes)]
            now = datetime.now().isoformat(timespec="seconds")
            with engine.begin() as connection:
                connection.execute(
                    update(extinguishers)
                    .where(extinguishers.c.id == extinguisher["id"])
                    .values(
                        company_id=selected_company["id"],
                        company_name=form["company_name"],
                        company_address=form["company_address"],
                        company_contact=form["report_number"],
                        asset_category=group["label"],
                        location_detail=form["company_address"],
                        weight_kg=0.0,
                        extinguisher_type="Elektrik Ic Tesisati",
                        fire_class=form["grid_type"],
                        manufacturer=form["energy_provider"],
                        hydrostatic_test_date=None,
                        pressure_status=None,
                        notes="\n".join(note_lines),
                        last_service_date=form["report_date"],
                        next_service_date=form["next_service_date"],
                        updated_at=now,
                    )
                )
            flash("Kayit guncellendi.", "success")
            return redirect(url_for("extinguisher_detail", public_id=public_id))

        return render_template(
            "create_electrical_installation.html",
            form=build_electrical_form_state(extinguisher, latest_log),
            companies=company_choices,
            group=group,
            asset_profile=asset_profile,
            form_title=f"{group['label']} Kaydini Duzenle",
            form_description="Bu ekranda mevcut elektrik ic tesisati kaydini guncelleyebilirsin.",
            submit_label="Kaydi Guncelle",
        )

    if category == DEFAULT_ASSET_CATEGORY:
        asset_categories = [get_asset_category(label=category) or {"label": category}]
        if request.method == "POST":
            form = parse_required_form(request.form)
            form["technician_name"] = current_user_full_name()
            form["asset_category"] = category
            try:
                form, selected_company = sync_company_payload_from_selection(form)
            except ValueError as exc:
                flash(str(exc), "error")
                return render_template(
                    "create_extinguisher.html",
                    form=form,
                    monthly_control_items=MONTHLY_CONTROL_ITEMS,
                    equipment_options=EQUIPMENT_OPTIONS,
                    equipment_presets=EQUIPMENT_PRESETS,
                    companies=company_choices,
                    asset_categories=asset_categories,
                    form_title="Kaydi Duzenle",
                    form_badge="Mevcut YSC kaydini guncelle",
                    primary_submit_label="Kaydi Guncelle",
                    show_secondary_submit=False,
                )
            required_fields = {
                "serial_number": "Seri numarasi",
                "company_id": "Cari secimi",
                "location_detail": "Firma ici konum",
                "weight_kg": "Kg bilgisi",
                "extinguisher_type": "Tup tipi",
                "fire_class": "YSC sinifi",
                "manufacturer": "YSC uretici",
                "last_service_date": "Son bakim tarihi",
                "hydrostatic_test_date": "Hidrostatik test tarihi",
                "next_service_date": "Sonraki bakim tarihi",
            }
            missing = [label for key, label in required_fields.items() if not form.get(key)]
            if missing:
                flash(f"Eksik alanlar: {', '.join(missing)}", "error")
                return render_template(
                    "create_extinguisher.html",
                    form=form,
                    monthly_control_items=MONTHLY_CONTROL_ITEMS,
                    equipment_options=EQUIPMENT_OPTIONS,
                    equipment_presets=EQUIPMENT_PRESETS,
                    companies=company_choices,
                    asset_categories=asset_categories,
                    form_title="Kaydi Duzenle",
                    form_badge="Mevcut YSC kaydini guncelle",
                    primary_submit_label="Kaydi Guncelle",
                    show_secondary_submit=False,
                )
            try:
                weight_kg = parse_float(form["weight_kg"], "Kg")
            except ValueError as exc:
                flash(str(exc), "error")
                return render_template(
                    "create_extinguisher.html",
                    form=form,
                    monthly_control_items=MONTHLY_CONTROL_ITEMS,
                    equipment_options=EQUIPMENT_OPTIONS,
                    equipment_presets=EQUIPMENT_PRESETS,
                    companies=company_choices,
                    asset_categories=asset_categories,
                    form_title="Kaydi Duzenle",
                    form_badge="Mevcut YSC kaydini guncelle",
                    primary_submit_label="Kaydi Guncelle",
                    show_secondary_submit=False,
                )
            existing_same_category = fetch_one(
                select(extinguishers.c.id)
                .where(extinguishers.c.asset_category == category)
                .where(extinguishers.c.serial_number == form["serial_number"])
                .where(extinguishers.c.id != extinguisher["id"])
            )
            if existing_same_category:
                flash("Bu seri numarasi bu urun grubunda zaten kayitli.", "error")
                return render_template(
                    "create_extinguisher.html",
                    form=form,
                    monthly_control_items=MONTHLY_CONTROL_ITEMS,
                    equipment_options=EQUIPMENT_OPTIONS,
                    equipment_presets=EQUIPMENT_PRESETS,
                    companies=company_choices,
                    asset_categories=asset_categories,
                    form_title="Kaydi Duzenle",
                    form_badge="Mevcut YSC kaydini guncelle",
                    primary_submit_label="Kaydi Guncelle",
                    show_secondary_submit=False,
                )
            now = datetime.now().isoformat(timespec="seconds")
            with engine.begin() as connection:
                connection.execute(
                    update(extinguishers)
                    .where(extinguishers.c.id == extinguisher["id"])
                    .values(
                        serial_number=form["serial_number"],
                        company_id=selected_company["id"],
                        company_name=form["company_name"],
                        company_address=form["company_address"],
                        company_contact=form.get("company_contact"),
                        asset_category=category,
                        location_detail=form["location_detail"],
                        weight_kg=weight_kg,
                        extinguisher_type=form["extinguisher_type"],
                        fire_class=form["fire_class"],
                        manufacturer=form["manufacturer"],
                        hydrostatic_test_date=form["hydrostatic_test_date"],
                        pressure_status=form.get("pressure_status"),
                        notes=form.get("notes"),
                        last_service_date=form["last_service_date"],
                        next_service_date=form["next_service_date"],
                        updated_at=now,
                    )
                )
            flash("Kayit guncellendi.", "success")
            return redirect(url_for("extinguisher_detail", public_id=public_id))

        form = {
            "serial_number": extinguisher.get("serial_number") or "",
            "company_id": str(extinguisher.get("company_id") or ""),
            "asset_category": category,
            "company_address": extinguisher.get("company_address") or "",
            "company_contact": extinguisher.get("company_contact") or "",
            "location_detail": extinguisher.get("location_detail") or "",
            "weight_kg": str(extinguisher.get("weight_kg") or ""),
            "extinguisher_type": extinguisher.get("extinguisher_type") or "",
            "pressure_status": extinguisher.get("pressure_status") or "",
            "fire_class": extinguisher.get("fire_class") or "",
            "manufacturer": extinguisher.get("manufacturer") or "",
            "last_service_date": extinguisher.get("last_service_date") or "",
            "hydrostatic_test_date": extinguisher.get("hydrostatic_test_date") or "",
            "next_service_date": extinguisher.get("next_service_date") or "",
            "technician_name": current_user_full_name(),
            "operation_summary": (latest_log or {}).get("operation_summary") or "",
            "notes": extinguisher.get("notes") or "",
        }
        if latest_inspection:
            for key, _label in MONTHLY_CONTROL_ITEMS:
                form[key] = bool(latest_inspection.get(key))
            for key, _label in CONTROL_FORM_ITEMS:
                form[key] = bool(latest_inspection.get(key))

        return render_template(
            "create_extinguisher.html",
            form=form,
            monthly_control_items=MONTHLY_CONTROL_ITEMS,
            equipment_options=EQUIPMENT_OPTIONS,
            equipment_presets=EQUIPMENT_PRESETS,
            companies=company_choices,
            asset_categories=asset_categories,
            form_title="Kaydi Duzenle",
            form_badge="Mevcut YSC kaydini guncelle",
            primary_submit_label="Kaydi Guncelle",
            show_secondary_submit=False,
        )

    group = get_registration_group_by_label(category)
    if group is None:
        abort(404)

    if request.method == "POST":
        form = parse_required_form(request.form)
        form["technician_name"] = current_user_full_name()
        form["asset_category"] = group["label"]
        form["extinguisher_type"] = asset_profile.get("fixed_type") or group["label"]
        try:
            form, selected_company = sync_company_payload_from_selection(form)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
                form_title=f"{group['label']} Kaydini Duzenle",
                form_description="Bu kayitta yanlis girilen alanlari guncelleyebilirsin.",
                primary_submit_label="Kaydi Guncelle",
                show_secondary_submit=False,
            )

        required_fields = {
            "serial_number": "Seri numarasi",
            "company_id": "Cari secimi",
            "company_contact": asset_profile["owner_label"],
            "location_detail": "Bulundugu yer",
            "fire_class": asset_profile["class_label"],
            "manufacturer": asset_profile["brand_label"],
            "last_service_date": asset_profile["service_input_label"],
            "next_service_date": asset_profile["next_service_input_label"],
        }
        if asset_profile["show_weight"]:
            required_fields["weight_kg"] = "Kg"
        if asset_profile["show_hydrostatic"]:
            required_fields["hydrostatic_test_date"] = "Hidrostatik test tarihi"
        missing = [label for key, label in required_fields.items() if not form.get(key)]
        if missing:
            flash(f"Eksik alanlar: {', '.join(missing)}", "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
                form_title=f"{group['label']} Kaydini Duzenle",
                form_description="Bu kayitta yanlis girilen alanlari guncelleyebilirsin.",
                primary_submit_label="Kaydi Guncelle",
                show_secondary_submit=False,
            )
        try:
            weight_kg = parse_float(form["weight_kg"], "Kg") if asset_profile["show_weight"] else 0.0
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
                form_title=f"{group['label']} Kaydini Duzenle",
                form_description="Bu kayitta yanlis girilen alanlari guncelleyebilirsin.",
                primary_submit_label="Kaydi Guncelle",
                show_secondary_submit=False,
            )
        existing_same_category = fetch_one(
            select(extinguishers.c.id)
            .where(extinguishers.c.asset_category == group["label"])
            .where(extinguishers.c.serial_number == form["serial_number"])
            .where(extinguishers.c.id != extinguisher["id"])
        )
        if existing_same_category:
            flash("Bu seri numarasi bu urun grubunda zaten kayitli.", "error")
            return render_template(
                "create_asset_profile.html",
                form=form,
                companies=company_choices,
                asset_profile=asset_profile,
                group=group,
                form_title=f"{group['label']} Kaydini Duzenle",
                form_description="Bu kayitta yanlis girilen alanlari guncelleyebilirsin.",
                primary_submit_label="Kaydi Guncelle",
                show_secondary_submit=False,
            )
        now = datetime.now().isoformat(timespec="seconds")
        with engine.begin() as connection:
            connection.execute(
                update(extinguishers)
                .where(extinguishers.c.id == extinguisher["id"])
                .values(
                    serial_number=form["serial_number"],
                    company_id=selected_company["id"],
                    company_name=form["company_name"],
                    company_address=form["company_address"],
                    company_contact=form["company_contact"],
                    asset_category=group["label"],
                    location_detail=form["location_detail"],
                    weight_kg=weight_kg,
                    extinguisher_type=form["extinguisher_type"],
                    fire_class=form["fire_class"],
                    manufacturer=form["manufacturer"],
                    hydrostatic_test_date=form.get("hydrostatic_test_date") if asset_profile["show_hydrostatic"] else None,
                    pressure_status=None,
                    notes=form.get("notes"),
                    last_service_date=form["last_service_date"],
                    next_service_date=form["next_service_date"],
                    updated_at=now,
                )
            )
        flash("Kayit guncellendi.", "success")
        return redirect(url_for("extinguisher_detail", public_id=public_id))

    form = {
        "serial_number": extinguisher.get("serial_number") or "",
        "company_id": str(extinguisher.get("company_id") or ""),
        "company_address": extinguisher.get("company_address") or "",
        "company_contact": extinguisher.get("company_contact") or "",
        "location_detail": extinguisher.get("location_detail") or "",
        "weight_kg": str(extinguisher.get("weight_kg") or ""),
        "last_service_date": extinguisher.get("last_service_date") or "",
        "next_service_date": extinguisher.get("next_service_date") or "",
        "hydrostatic_test_date": extinguisher.get("hydrostatic_test_date") or "",
        "fire_class": extinguisher.get("fire_class") or "",
        "manufacturer": extinguisher.get("manufacturer") or "",
        "extinguisher_type": asset_profile.get("fixed_type") or group["label"],
        "technician_name": current_user_full_name(),
        "operation_summary": (latest_log or {}).get("operation_summary") or "",
        "notes": extinguisher.get("notes") or "",
    }
    if latest_inspection:
        for key, _label in asset_profile["monthly_control_items"]:
            form[key] = bool(latest_inspection.get(key))

    return render_template(
        "create_asset_profile.html",
        form=form,
        companies=company_choices,
        asset_profile=asset_profile,
        group=group,
        form_title=f"{group['label']} Kaydini Duzenle",
        form_description="Bu kayitta yanlis girilen alanlari guncelleyebilirsin.",
        primary_submit_label="Kaydi Guncelle",
        show_secondary_submit=False,
    )


@app.route("/extinguishers/<public_id>")
@login_required
def extinguisher_detail(public_id: str):
    extinguisher = get_extinguisher(public_id)
    asset_profile = get_asset_profile(extinguisher.get("asset_category"))
    electrical_sections = (
        build_electrical_note_sections(extinguisher.get("notes"))
        if extinguisher.get("asset_category") == "Elektrik Ic Tesisati"
        else []
    )
    company_portal_url = None
    if extinguisher.get("company_id"):
        company = get_company(extinguisher["company_id"])
        company_portal_url = url_for(
            "public_company_portal",
            company_slug=company["slug"],
            _external=True,
        )
    service_history = fetch_all(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
    )
    monthly_history = with_monthly_control_labels(
        fetch_all(
            select(monthly_inspections)
            .where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
            .order_by(
                desc(monthly_inspections.c.inspection_date),
                desc(monthly_inspections.c.id),
            )
        ),
        asset_profile["monthly_control_items"],
    )
    return render_template(
        "extinguisher_detail.html",
        extinguisher=extinguisher,
        service_logs=service_history,
        monthly_inspections=monthly_history,
        monthly_control_items=asset_profile["monthly_control_items"],
        equipment_preset=get_equipment_preset(extinguisher.get("extinguisher_type")),
        company_portal_url=company_portal_url,
        asset_profile=asset_profile,
        electrical_sections=electrical_sections,
        can_delete=is_admin_user(),
    )


@app.route("/extinguishers/<public_id>/delete", methods=["POST"])
@admin_required
def delete_extinguisher(public_id: str):
    extinguisher = get_extinguisher(public_id)
    with engine.begin() as connection:
        connection.execute(
            service_logs.delete().where(service_logs.c.extinguisher_id == extinguisher["id"])
        )
        connection.execute(
            monthly_inspections.delete().where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
        )
        connection.execute(
            extinguishers.delete().where(extinguishers.c.id == extinguisher["id"])
        )
    flash("Kayit silindi.", "success")
    return redirect(url_for("index"))


@app.route("/extinguishers/<public_id>/service/<int:log_id>/delete", methods=["POST"])
@admin_required
def delete_service_log(public_id: str, log_id: int):
    extinguisher = get_extinguisher(public_id)
    target_log = fetch_one(
        select(service_logs).where(
            service_logs.c.id == log_id,
            service_logs.c.extinguisher_id == extinguisher["id"],
        )
    )
    if target_log is None:
        abort(404)
    with engine.begin() as connection:
        connection.execute(
            service_logs.delete().where(service_logs.c.id == log_id)
        )
    flash("Bakim kaydi silindi.", "success")
    return redirect(url_for("extinguisher_detail", public_id=public_id))


@app.route("/extinguishers/<public_id>/monthly-inspection/<int:inspection_id>/delete", methods=["POST"])
@admin_required
def delete_monthly_inspection(public_id: str, inspection_id: int):
    extinguisher = get_extinguisher(public_id)
    target_inspection = fetch_one(
        select(monthly_inspections).where(
            monthly_inspections.c.id == inspection_id,
            monthly_inspections.c.extinguisher_id == extinguisher["id"],
        )
    )
    if target_inspection is None:
        abort(404)
    with engine.begin() as connection:
        connection.execute(
            monthly_inspections.delete().where(monthly_inspections.c.id == inspection_id)
        )
    flash("Aylik kontrol kaydi silindi.", "success")
    return redirect(url_for("extinguisher_detail", public_id=public_id))


@app.route("/extinguishers/<public_id>/control-form.pdf")
@login_required
def control_form_pdf(public_id: str):
    document_data = build_control_form_document_data(public_id)
    pdf_buffer = build_control_form_pdf_reportlab(document_data)
    filename = f"{build_company_filename(document_data['company_name'])}-kontrol-formu.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/public/<public_id>/control-form.pdf")
def public_control_form_pdf(public_id: str):
    document_data = build_control_form_document_data(public_id)
    pdf_buffer = build_control_form_pdf_reportlab(document_data)
    filename = f"{build_company_filename(document_data['company_name'])}-kontrol-formu.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


def build_category_report_bundle(public_id: str) -> tuple[dict, io.BytesIO]:
    extinguisher = get_extinguisher(public_id)
    asset_category = extinguisher.get("asset_category")
    try:
        if asset_category in SPECIAL_CATEGORY_FORM_CONFIGS:
            document_data = build_special_category_company_document_data(public_id)
            pdf_buffer = build_special_category_company_form_pdf(document_data)
        else:
            document_data = build_company_category_report_document_data(public_id)
            pdf_buffer = build_company_category_report_pdf(document_data)
        return document_data, pdf_buffer
    except Exception as exc:
        print(f"[category-report-fallback] {asset_category} {public_id}: {exc}")
        document_data = build_company_category_report_document_data(public_id)
        pdf_buffer = build_company_category_report_pdf(document_data)
        return document_data, pdf_buffer


@app.route("/extinguishers/<public_id>/category-report.pdf")
@login_required
def category_report_pdf(public_id: str):
    document_data, pdf_buffer = build_category_report_bundle(public_id)
    filename = (
        f"{build_company_filename(document_data['company_name'])}-"
        f"{build_company_filename(document_data['asset_category'])}-kategori-raporu.pdf"
    )
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/public/<public_id>/category-report.pdf")
def public_category_report_pdf(public_id: str):
    document_data, pdf_buffer = build_category_report_bundle(public_id)
    filename = (
        f"{build_company_filename(document_data['company_name'])}-"
        f"{build_company_filename(document_data['asset_category'])}-kategori-raporu.pdf"
    )
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/extinguishers/<public_id>/electrical-report.pdf")
@login_required
def electrical_report_pdf(public_id: str):
    extinguisher = get_extinguisher(public_id)
    if extinguisher.get("asset_category") != "Elektrik Ic Tesisati":
        abort(404)
    pdf_buffer = build_electrical_report_pdf_html(public_id)
    filename = f"{build_company_filename(extinguisher['company_name'])}-elektrik-ic-tesisati-raporu.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/public/<public_id>/electrical-report.pdf")
def public_electrical_report_pdf(public_id: str):
    extinguisher = get_extinguisher(public_id)
    if extinguisher.get("asset_category") != "Elektrik Ic Tesisati":
        abort(404)
    pdf_buffer = build_electrical_report_pdf_html(public_id)
    filename = f"{build_company_filename(extinguisher['company_name'])}-elektrik-ic-tesisati-raporu.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/firma/<company_slug>")
def public_company_portal(company_slug: str):
    company = get_company_by_slug(company_slug)
    sections = build_company_portal_sections(company["id"])
    return render_template(
        "public_company_portal.html",
        company=company,
        sections=sections,
        selected_category=None,
        selected_assets=[],
    )


@app.route("/firma/<company_slug>/<category_slug>")
def public_company_assets(company_slug: str, category_slug: str):
    company = get_company_by_slug(company_slug)
    selected_category = get_asset_category(slug=category_slug)
    if selected_category is None:
        abort(404)
    sections = build_company_portal_sections(company["id"])
    selected_section = next(
        (section for section in sections if section["slug"] == category_slug),
        None,
    )
    return render_template(
        "public_company_portal.html",
        company=company,
        sections=sections,
        selected_category=selected_category,
        selected_assets=selected_section["items"] if selected_section else [],
    )


@app.route("/extinguishers/<public_id>/control-form")
@login_required
def control_form_preview(public_id: str):
    document_data = build_control_form_document_data(public_id)
    return render_template("control_form_preview.html", **document_data)


@app.route("/extinguishers/<public_id>/service", methods=["GET", "POST"])
@login_required
def add_service_log(public_id: str):
    extinguisher = get_extinguisher(public_id)
    company_choices = get_company_choices()
    asset_categories = get_asset_category_choices()
    asset_profile = get_asset_profile(extinguisher.get("asset_category"))
    if request.method == "POST":
        form = parse_required_form(request.form)
        form["technician_name"] = form.get("technician_name") or current_user_full_name()
        try:
            form, selected_company = sync_company_payload_from_selection(form)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "service_log_form.html",
                extinguisher=extinguisher,
                form=form,
                monthly_control_items=MONTHLY_CONTROL_ITEMS,
                equipment_options=EQUIPMENT_OPTIONS,
                equipment_presets=EQUIPMENT_PRESETS,
                companies=company_choices,
                asset_categories=asset_categories,
                asset_profile=asset_profile,
            )
        required_fields = {
            "service_date": asset_profile["last_service_label"],
            "next_service_date": asset_profile["next_service_label"],
            "technician_name": "Teknisyen",
            "company_id": "Cari secimi",
            "asset_category": "Urun grubu",
            "location_detail": "Bulundugu yer" if asset_profile["label"] != "Yangin Sondurme Cihazi" else "Firma ici konum",
            "fire_class": asset_profile["class_label"],
            "manufacturer": asset_profile["brand_label"],
            "operation_summary": "Yapilan islem",
            "company_contact": asset_profile["owner_label"],
        }
        if asset_profile["show_weight"]:
            required_fields["weight_kg"] = "Kg"
        if asset_profile["label"] == "Yangin Sondurme Cihazi":
            required_fields["extinguisher_type"] = "Tup tipi"
        if asset_profile["show_hydrostatic"]:
            required_fields["hydrostatic_test_date"] = "Hidrostatik test tarihi"
        missing = [label for key, label in required_fields.items() if not form.get(key)]
        if missing:
            flash(f"Eksik alanlar: {', '.join(missing)}", "error")
            return render_template(
                "service_log_form.html",
                extinguisher=extinguisher,
                form=form,
                monthly_control_items=MONTHLY_CONTROL_ITEMS,
                equipment_options=EQUIPMENT_OPTIONS,
                equipment_presets=EQUIPMENT_PRESETS,
                companies=company_choices,
                asset_categories=asset_categories,
                asset_profile=asset_profile,
            )

        if asset_profile["show_weight"]:
            try:
                weight_kg = parse_float(
                    form.get("weight_kg") or str(extinguisher["weight_kg"]),
                    "Kg",
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return render_template(
                    "service_log_form.html",
                    extinguisher=extinguisher,
                    form=form,
                    monthly_control_items=MONTHLY_CONTROL_ITEMS,
                    equipment_options=EQUIPMENT_OPTIONS,
                    equipment_presets=EQUIPMENT_PRESETS,
                    companies=company_choices,
                    asset_categories=asset_categories,
                    asset_profile=asset_profile,
                )
        else:
            weight_kg = extinguisher.get("weight_kg") or 0.0

        now = datetime.now().isoformat(timespec="seconds")
        inspection_values = build_monthly_inspection_values(request.form, asset_profile["monthly_control_items"])
        control_values = (
            derive_ysc_control_form_values(
                inspection_values,
                build_control_form_values(request.form),
            )
            if asset_profile["control_form_items"]
            else build_control_form_values({})
        )
        with engine.begin() as connection:
            connection.execute(
                insert(service_logs).values(
                    extinguisher_id=extinguisher["id"],
                    service_date=form["service_date"],
                    technician_name=form["technician_name"],
                    operation_summary=form["operation_summary"],
                    pressure_status=form.get("pressure_status"),
                    notes=form.get("notes"),
                    created_at=now,
                )
            )
            connection.execute(
                update(extinguishers)
                .where(extinguishers.c.id == extinguisher["id"])
                .values(
                    company_id=selected_company["id"],
                    company_name=form.get("company_name") or extinguisher["company_name"],
                    company_address=form.get("company_address") or extinguisher.get("company_address"),
                    company_contact=form.get("company_contact") or extinguisher.get("company_contact"),
                    asset_category=form.get("asset_category") or extinguisher.get("asset_category") or DEFAULT_ASSET_CATEGORY,
                    location_detail=form.get("location_detail")
                    or extinguisher["location_detail"],
                    weight_kg=weight_kg,
                    extinguisher_type=form.get("extinguisher_type")
                    or extinguisher["extinguisher_type"],
                    fire_class=form.get("fire_class") or extinguisher.get("fire_class"),
                    manufacturer=form.get("manufacturer") or extinguisher.get("manufacturer"),
                    hydrostatic_test_date=(form.get("hydrostatic_test_date") or extinguisher.get("hydrostatic_test_date")) if asset_profile["show_hydrostatic"] else None,
                    pressure_status=form.get("pressure_status") if asset_profile["show_pressure"] else None,
                    notes=form.get("notes"),
                    last_service_date=form["service_date"],
                    next_service_date=form["next_service_date"],
                    updated_at=now,
                )
            )
            save_monthly_inspection(
                connection=connection,
                extinguisher_id=extinguisher["id"],
                inspection_date=form["service_date"],
                inspector_name=form["technician_name"],
                notes=form.get("notes"),
                inspection_values=inspection_values,
                control_values=control_values,
                created_at=now,
            )

        flash("Bakim kaydi eklendi.", "success")
        return redirect(url_for("extinguisher_detail", public_id=public_id))

    return render_template(
        "service_log_form.html",
        extinguisher=extinguisher,
        form={
            "technician_name": current_user_full_name(),
            "company_id": str(extinguisher.get("company_id") or ""),
            "company_name": extinguisher.get("company_name") or "",
            "company_address": extinguisher.get("company_address") or "",
            "company_contact": extinguisher.get("company_contact") or "",
            "asset_category": extinguisher.get("asset_category") or DEFAULT_ASSET_CATEGORY,
        },
        monthly_control_items=MONTHLY_CONTROL_ITEMS,
        equipment_options=EQUIPMENT_OPTIONS,
        equipment_presets=EQUIPMENT_PRESETS,
        companies=company_choices,
        asset_categories=asset_categories,
        asset_profile=asset_profile,
    )


@app.route("/public/<public_id>")
def public_detail(public_id: str):
    extinguisher = get_extinguisher(public_id)
    asset_profile = get_asset_profile(extinguisher.get("asset_category"))
    electrical_sections = (
        build_electrical_note_sections(extinguisher.get("notes"))
        if extinguisher.get("asset_category") == "Elektrik Ic Tesisati"
        else []
    )
    company_portal_url = None
    if extinguisher.get("company_id"):
        company_portal_url = url_for(
            "public_company_portal",
            company_slug=get_company(extinguisher["company_id"])["slug"],
        )
    latest_log = fetch_one(
        select(service_logs)
        .where(service_logs.c.extinguisher_id == extinguisher["id"])
        .order_by(desc(service_logs.c.service_date), desc(service_logs.c.id))
        .limit(1)
    )
    latest_monthly_inspection_raw = fetch_one(
        select(monthly_inspections)
        .where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
        .order_by(
            desc(monthly_inspections.c.inspection_date),
            desc(monthly_inspections.c.id),
        )
        .limit(1)
    )
    monthly_history_raw = fetch_all(
        select(monthly_inspections)
        .where(monthly_inspections.c.extinguisher_id == extinguisher["id"])
        .order_by(
            desc(monthly_inspections.c.inspection_date),
            desc(monthly_inspections.c.id),
        )
    )
    latest_monthly_inspection = (
        with_monthly_control_labels([latest_monthly_inspection_raw], asset_profile["monthly_control_items"])[0]
        if latest_monthly_inspection_raw
        else None
    )
    return render_template(
        "public_detail.html",
        extinguisher=extinguisher,
        latest_log=latest_log,
        latest_monthly_inspection=latest_monthly_inspection,
        equipment_preset=get_equipment_preset(extinguisher.get("extinguisher_type")),
        monthly_table=build_monthly_table(monthly_history_raw, asset_profile["monthly_control_items"]),
        company_portal_url=company_portal_url,
        asset_profile=asset_profile,
        electrical_sections=electrical_sections,
    )


@app.route("/extinguishers/<public_id>/monthly-inspection", methods=["GET", "POST"])
@login_required
def add_monthly_inspection(public_id: str):
    extinguisher = get_extinguisher(public_id)
    asset_profile = get_asset_profile(extinguisher.get("asset_category"))
    if request.method == "POST":
        form = parse_required_form(request.form)
        form["inspector_name"] = form.get("inspector_name") or current_user_full_name()
        required_fields = {
            "inspection_date": "Kontrol tarihi",
            "inspector_name": "Kontrol eden",
        }
        missing = [label for key, label in required_fields.items() if not form.get(key)]
        if missing:
            flash(f"Eksik alanlar: {', '.join(missing)}", "error")
            return render_template(
                "monthly_inspection_form.html",
                extinguisher=extinguisher,
                monthly_control_items=asset_profile["monthly_control_items"],
                control_form_items=asset_profile["control_form_items"],
                form=form,
                asset_profile=asset_profile,
            )

        now = datetime.now().isoformat(timespec="seconds")
        inspection_values = build_monthly_inspection_values(request.form, asset_profile["monthly_control_items"])
        control_values = (
            derive_ysc_control_form_values(
                inspection_values,
                build_control_form_values(request.form),
            )
            if asset_profile["control_form_items"]
            else build_control_form_values({})
        )
        with engine.begin() as connection:
            save_monthly_inspection(
                connection=connection,
                extinguisher_id=extinguisher["id"],
                inspection_date=form["inspection_date"],
                inspector_name=form["inspector_name"],
                notes=form.get("notes"),
                inspection_values=inspection_values,
                control_values=control_values,
                created_at=now,
            )

        flash("Aylık kontrol kaydı eklendi.", "success")
        return redirect(url_for("extinguisher_detail", public_id=public_id))

    return render_template(
        "monthly_inspection_form.html",
        extinguisher=extinguisher,
        monthly_control_items=asset_profile["monthly_control_items"],
        control_form_items=asset_profile["control_form_items"],
        form={"inspector_name": current_user_full_name()},
        asset_profile=asset_profile,
    )


@app.route("/extinguishers/<public_id>/qr")
@login_required
def extinguisher_qr(public_id: str):
    get_extinguisher(public_id)
    public_url = url_for("public_detail", public_id=public_id, _external=True)
    buffer = build_branded_qr(public_url)
    return send_file(buffer, mimetype="image/png", download_name=f"{public_id}.png")


@app.route("/extinguishers/<public_id>/label-qr")
@login_required
def extinguisher_label_qr(public_id: str):
    get_extinguisher(public_id)
    public_url = url_for("public_detail", public_id=public_id, _external=True)
    buffer = build_branded_qr(public_url, label_mode=True)
    return send_file(buffer, mimetype="image/png", download_name=f"{public_id}-label.png")


@app.route("/extinguishers/<public_id>/label")
@login_required
def extinguisher_label(public_id: str):
    get_extinguisher(public_id)
    return render_template("label.html", public_id=public_id)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=True)
